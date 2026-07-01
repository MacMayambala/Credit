import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import io
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .services import DataMigrationService


def generate_migration_template_http() -> HttpResponse:
    """
    Constructs a dynamic openpyxl spreadsheet buffer loaded with standard headers,
    color-coded required metrics, instructions, and data validations.
    """

    wb = openpyxl.Workbook()

    # =========================
    # Worksheet Setup
    # =========================
    ws = wb.active
    ws.title = "Data Migration Upload"
    ws.sheet_view.showGridLines = True

    # =========================
    # Styles
    # =========================
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    instruct_font = Font(name="Segoe UI", size=10, italic=True, color="333333")

    required_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    optional_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # =========================
    # Instructions Row
    # =========================
    ws.append([
        "INSTRUCTIONS: Fill mandatory columns carefully. Do not change headers. "
        "Use YYYY-MM-DD for dates. Leave loan fields empty if not applicable."
    ])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=50)

    instr_cell = ws.cell(row=1, column=1)
    instr_cell.font = instruct_font
    instr_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # =========================
    # DYNAMIC HEADERS (SINGLE SOURCE OF TRUTH)
    # =========================
    RAW_HEADERS = DataMigrationService.TEMPLATE_HEADERS_ORDERED

    REQUIRED_FIELDS = {
        "member_number",
        "first_name",
        "last_name",
        "gender",
        "date_of_birth",
        "nin",
        "phone_number",
        "savings_balance",
    }

    headers = []
    for h in RAW_HEADERS:
        style_fill = required_fill if h in REQUIRED_FIELDS else optional_fill
        headers.append((h, style_fill))

    # =========================
    # Header Row (Row 2)
    # =========================
    for col_idx, (header_text, style_fill) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = style_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[2].height = 28

    # =========================
    # Data Validations
    # =========================
    gender_dv = DataValidation(
        type="list",
        formula1='"Male,Female,Other"',
        allow_blank=False
    )
    gender_dv.error = "Invalid gender selection"
    gender_dv.errorTitle = "Gender Error"
    ws.add_data_validation(gender_dv)

    if "gender" in RAW_HEADERS:
        col = RAW_HEADERS.index("gender") + 1
        gender_dv.add(f"{get_column_letter(col)}3:{get_column_letter(col)}5000")

    product_dv = DataValidation(
        type="list",
        formula1='"personal,business,salary,group,emergency,asset"',
        allow_blank=True
    )
    product_dv.error = "Invalid product type"
    product_dv.errorTitle = "Product Error"
    ws.add_data_validation(product_dv)

    if "product_type" in RAW_HEADERS:
        col = RAW_HEADERS.index("product_type") + 1
        product_dv.add(f"{get_column_letter(col)}3:{get_column_letter(col)}5000")

    # =========================
    # Sample Row (DYNAMIC SAFE)
    # =========================
    sample_row = []
    for h in RAW_HEADERS:
        if h == "member_number":
            sample_row.append("KAL001")
        elif h == "first_name":
            sample_row.append("John")
        elif h == "last_name":
            sample_row.append("Doe")
        elif h == "gender":
            sample_row.append("Male")
        elif h == "date_of_birth":
            sample_row.append("1990-01-01")
        elif h == "nin":
            sample_row.append("CMXXXXXXXXXXXX")
        elif h == "phone_number":
            sample_row.append("0772000111")
        elif h == "savings_balance":
            sample_row.append("250000")
        elif h == "loan_reference":
            sample_row.append("LN-2026-001")
        elif h == "principal_amount":
            sample_row.append("1000000")
        elif h == "interest_rate":
            sample_row.append("12")
        elif h == "product_type":
            sample_row.append("personal")
        else:
            sample_row.append("")

    for col_idx, value in enumerate(sample_row, start=1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.border = thin_border
        cell.font = Font(name="Segoe UI", size=10, color="555555")

    # =========================
    # Auto column sizing
    # =========================
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 40)

    # =========================
    # Output buffer
    # =========================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Customer_migration_template.xlsx"'

    return response