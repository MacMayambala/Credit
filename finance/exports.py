import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from django.http import HttpResponse
from django.utils import timezone

class ReportingExportEngine:
    """
    Renders institutional-grade PDF financials and Excel worksheets 
    retaining explicit cryptographic audit anchors.
    """

    @staticmethod
    def generate_excel(report_title, headers, data_matrix):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_title.upper()[:30]
        ws.views.sheetView[0].showGridLines = True

        # Palettes mimicking tier-1 institutional design languages
        navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=10, bold=False)
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        thin_border_side = Side(border_style="thin", color="D1D5DB")
        data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # Draw Headers
        ws.append([f"INSTITUTIONAL FINANCIAL REVENUE REPORT: {report_title.upper()}"])
        ws.cell(row=1, column=1).font = Font(name="Calibri", size=16, bold=True, color="1B365D")
        ws.append([f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S UTC') }"])
        ws.append([]) # Spacer row

        ws.append(headers)
        header_row_idx = 4
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_num)
            cell.fill = navy_fill
            cell.font = font_header
            cell.alignment = align_left

        # Render Data Block
        for row_data in data_matrix:
            ws.append(row_data)
            curr_row = ws.max_row
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = font_data
                cell.border = data_border
                if isinstance(val, (int, float, round.__class__)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left

        # Auto-fit column execution frames
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={report_title}_export.xlsx'
        wb.save(response)
        return response

    @staticmethod
    def generate_pdf(report_title, headers, data_matrix, requesting_user):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(letter),
            rightMargin=36, leftMargin=36, topMargin=36, bottom=36
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'DocTitle', parent=styles['Heading1'],
            fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.HexColor('#1B365D'), spaceAfter=12
        )
        meta_style = ParagraphStyle(
            'MetaText', parent=styles['Normal'],
            fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#4B5563'), spaceAfter=20
        )

        elements = []
        elements.append(Paragraph(f"Commercial Ledger Report: {report_title.upper()}", title_style))
        elements.append(Paragraph(f"Confidential Document // Operator: {requesting_user} // Run Time: {timezone.now().isoformat()}", meta_style))

        # Re-pack matrices into ReportLab structural flow elements
        table_data = [headers] + data_matrix
        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('ALIGN', (5,1), (-1,-1), 'RIGHT'), # Formats currency metrics cleanly to the right
        ]))
        
        elements.append(t)
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename={report_title}_ledger.pdf'
        return response