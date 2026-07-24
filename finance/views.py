from enum import member

from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.utils import timezone
from decimal import Decimal
from .models import (
    Installment, Loan, Member, Repayment, SavingsAccount, SystemSetting, Transaction, TransactionReversal, 
    process_repayment, generate_schedule
)
from .models import Loan, LoanPenaltyRule, Member, Installment
# finance/views.py
import json
import random
import string
import logging
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta

from .models import (
    Loan,
    Member,
    Installment,
    LoanPenaltyRule,          # <-- add this
    generate_schedule,
           # if you have these helpers
)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum, F, Q, DecimalField, ExpressionWrapper, Window
from django.db.models.functions import Coalesce, ExtractMonth, TruncMonth  # Add TruncMonth here
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from django.contrib.auth.models import User
from django_celery_beat.models import PeriodicTask

from .models import (
    Installment, Loan, Member, SavingsAccount, Transaction, 
    Repayment, TransactionReversal, SystemSetting, GeneralLedger,
    ChartOfAccount, AutoRepaymentSetting, AutoRepaymentLog, DailyRepaymentSummary
)
from .services import (
    FinancialTransactionService, process_repayment, generate_transaction_ref,
    LoanRepaymentEngineService
)
from .forms import AutoRepaymentSettingForm
from .utils import send_bulk_arrears_reminders, generate_schedule

logger = logging.getLogger(__name__)

from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.utils import timezone
from django.contrib import messages
from decimal import Decimal
from django.db.models import Sum
from .models import (
    Installment, Loan, Member, SavingsAccount, Transaction, 
    process_repayment, generate_schedule
)

from django.db.models import Sum, F
from .models import Loan, Member, SavingsAccount

import json
from django.shortcuts import render
from django.db.models import Sum
from django.db.models.functions import ExtractMonth
from .models import Loan, Member, SavingsAccount, Transaction

import json
from django.db.models import Sum
from django.db.models.functions import ExtractMonth
from django.shortcuts import render

from .models import Member, SavingsAccount, Loan, Transaction


import json
from django.db.models.functions import ExtractMonth
from django.db.models import Sum
from django.shortcuts import render

from .models import Loan, Transaction


import json
from django.db.models import Sum, DecimalField
from django.db.models.functions import ExtractMonth, Cast
from django.shortcuts import render
from .models import Member, Loan, SavingsAccount, Transaction

import json
from decimal import Decimal
from datetime import date
from dateutil.relativedelta import relativedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.utils import timezone
from django.contrib import messages
from django.db.models import Sum, F, Q
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db.models.functions import ExtractMonth

from .models import (
    Installment, Loan, Member, SavingsAccount, Transaction, 
    Repayment, process_repayment, generate_schedule
)
from .utils import send_bulk_arrears_reminders

# ========================
# DASHBOARD & REGISTRY
# ========================
from django.core.exceptions import PermissionDenied
from django.shortcuts import redirect

import string
import random
import json
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import date
from dateutil.relativedelta import relativedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.db.models import Sum, F, Q, DecimalField
from django.db.models.functions import ExtractMonth, Cast
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied, ValidationError

from .models import (
    Installment, Loan, Member, SavingsAccount, Transaction, 
    Repayment, TransactionReversal, SystemSetting,
    process_repayment, generate_schedule
)
from .utils import send_bulk_arrears_reminders, generate_transaction_ref

# ========================
# UTILITIES & DECORATORS
# ========================

def allowed_users(allowed_roles=[]):
    """
    RBAC View Decorator to check user groups.
    """
    def decorator(view_func):
        def wrapper_func(request, *args, **kwargs):
            group = None
            if request.user.groups.exists():
                group = request.user.groups.all()[0].name

            if group in allowed_roles or request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            else:
                raise PermissionDenied  # Shows 403 Forbidden
        return wrapper_func
    return decorator


def generate_loan_ref(length=10):
    """Generates a random uppercase alphanumeric string for loans"""
    chars = string.ascii_uppercase + string.digits
    return ''.join(random.choices(chars, k=length))


# ========================
# CORE CORE SACCO VIEWS
# ========================

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from decimal import Decimal

from finance.models import SavingsAccount, Loan, Transaction, Member
from django.db.models.functions import ExtractMonth
# finance/views.py – dashboard view (fully updated)
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from django.db.models.functions import ExtractMonth

from finance.models import (
    SavingsAccount, Loan, Transaction, Member,
    GeneralLedger, ChartOfAccount
)
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from django.db.models.functions import ExtractMonth

from finance.models import (
    SavingsAccount, Loan, Transaction, Member,
    GeneralLedger, ChartOfAccount, Installment
)

@login_required
def dashboard(request):
    """
    Main SACCO Dashboard with updated split-balance aggregation
    and a true Interest Yield percentage.
    """
    today = timezone.now().date()
    one_year_ago = today - relativedelta(years=1)

    # ============================================================
    # 1. SUMMARY WIDGETS
    # ============================================================

    # Total Savings
    total_savings = SavingsAccount.objects.aggregate(total=Sum('balance'))['total'] or 0

    # Total Active Loans (principal_balance + interest_balance)
    loan_stats = Loan.objects.filter(is_active=True).aggregate(
        p_bal=Sum('principal_balance'),
        i_bal=Sum('interest_balance')
    )
    total_loans = (loan_stats['p_bal'] or 0) + (loan_stats['i_bal'] or 0)

    # ============================================================
    # Interest Earned in the last 12 months (from GeneralLedger)
    # ============================================================
    try:
        # Your ChartOfAccount code for Interest Income is '2100' (as seen in Repayment.save())
        interest_account = ChartOfAccount.objects.get(code='2100')
        interest_last_year = GeneralLedger.objects.filter(
            account=interest_account,
            date__gte=one_year_ago
        ).aggregate(total=Sum('credit'))['total'] or 0
    except ChartOfAccount.DoesNotExist:
        # Fallback: sum interest_paid from Installment model
        interest_last_year = Installment.objects.filter(
            loan__is_active=True,
            due_date__gte=one_year_ago
        ).aggregate(total=Sum('interest_paid'))['total'] or 0

    # ============================================================
    # Interest Yield (annualized) = (interest_last_year / total_loans) * 100
    # ============================================================
    if total_loans > 0:
        interest_yield = (interest_last_year / total_loans) * 100
    else:
        interest_yield = 0

    # Keep the raw amount for display if needed (e.g., in a tooltip)
    total_interest = interest_last_year

    # Counts
    total_members = Member.objects.count()
    active_loans_count = Loan.objects.filter(is_active=True).count()
    recent_loans = Loan.objects.select_related('member').order_by('-start_date')[:10]

    # ============================================================
    # 2. CHART DATA (Monthly trends for last 12 months)
    # ============================================================

    labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    savings_trend = [0] * 12
    loan_trend = [0] * 12

    # Savings deposits per month
    savings_data = Transaction.objects.filter(
        type='deposit',
        timestamp__gte=one_year_ago
    ).annotate(month=ExtractMonth('timestamp')).values('month').annotate(total=Sum('amount'))

    for entry in savings_data:
        if entry['month'] and 1 <= entry['month'] <= 12:
            savings_trend[entry['month'] - 1] = float(entry['total'])

    # Loan disbursements per month
    loan_data = Loan.objects.filter(
        start_date__gte=one_year_ago,
        status='approved'
    ).annotate(month=ExtractMonth('start_date')).values('month').annotate(total=Sum('principal_amount'))

    for entry in loan_data:
        if entry['month'] and 1 <= entry['month'] <= 12:
            loan_trend[entry['month'] - 1] = float(entry['total'])

    # ============================================================
    # 3. CONTEXT
    # ============================================================

    context = {
        'total_savings': total_savings,
        'total_loans': total_loans,
        'total_interest': total_interest,          # UGX amount earned last year
        'interest_yield': interest_yield,          # percentage
        'total_members': total_members,
        'active_loans_count': active_loans_count,
        'loans': recent_loans,
        'chart_labels': labels,
        'chart_savings': savings_trend,
        'chart_loans': loan_trend,
        'today': today,
    }

    return render(request, 'finance/dashboard.html', context)


@login_required
def loan_list(request):
    """Master Loan Registry with split-balance analytics"""
    loans = Loan.objects.select_related('member').all().order_by('-id')
    total_disbursed = loans.aggregate(Sum('principal_amount'))['principal_amount__sum'] or 0
    
    portfolio_stats = loans.aggregate(
        p_total=Sum('principal_balance'),
        i_total=Sum('interest_balance')
    )
    total_outstanding = (portfolio_stats['p_total'] or 0) + (portfolio_stats['i_total'] or 0)
    active_count = loans.filter(is_active=True).count()

    context = {
        'loans': loans,
        'total_disbursed': total_disbursed,
        'total_outstanding': total_outstanding,
        'active_count': active_count,
    }
    return render(request, 'finance/loan_list.html', context)


from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from decimal import Decimal

from django.db.models import Sum
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum
from decimal import Decimal
from dateutil.relativedelta import relativedelta

from .models import Loan, Installment, Repayment, ManualPenalty
from .utils import generate_schedule  # assuming you have this utility

# finance/views.py
# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from .models import Loan, ManualPenalty

# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from .models import Loan, ManualPenalty
from .utils import generate_schedule   # if you have it; otherwise import from models

# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta

from .models import Loan, ManualPenalty
from .utils import generate_schedule   # if you have this in utils.py; adjust if needed


# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta

from .models import Loan, ManualPenalty
from .utils import generate_schedule   # adjust import if needed

# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta

from .models import Loan, ManualPenalty
from .utils import generate_schedule   # adjust if you have it elsewhere


# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta
import math   # for ceil in penalty calculation

from .models import Loan, ManualPenalty
from .utils import generate_schedule   # or from .models import generate_schedule
from finance.penalties import calculate_penalty   # ensure this import works
# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta

from .models import Loan, ManualPenalty
from .utils import generate_schedule   # or import from wherever you have it
from finance.penalties import calculate_penalty   # your penalty calculator

# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from .models import Loan, ManualPenalty
from .utils import generate_schedule
from finance.penalties import calculate_penalty

# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta

from .models import Loan, ManualPenalty
from .utils import generate_schedule
from finance.penalties import calculate_penalty


# finance/views.py
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta

from .models import Loan, ManualPenalty
from .utils import generate_schedule
from finance.penalties import calculate_penalty
from .models import SMSConfig


@login_required
def loan_detail(request, pk):
    loan = get_object_or_404(Loan.objects.select_related('member', 'officer'), pk=pk)
    today = timezone.now().date()

    # Generate schedule if missing
    if not loan.installments.exists():
        generate_schedule(loan)

    # Due amounts for banner
    active_due = loan.installments.filter(paid=False, due_date__lte=today).aggregate(
        total_interest=Sum('interest_portion'),
        total_principal=Sum('principal_portion')
    )
    interest_due = active_due['total_interest'] or Decimal('0.00')
    principal_due = active_due['total_principal'] or Decimal('0.00')
    total_due_now = (interest_due + principal_due).quantize(Decimal('0.01'))

    total_paid = loan.repayments.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    total_payable = loan.total_payable or Decimal('0')

    disbursement_date = loan.disbursed_date or loan.start_date
    end_date = disbursement_date + relativedelta(months=loan.period_months) if disbursement_date else None

    # --- Build enriched schedule (combines calculated + manual penalties) ---
    # We build a list of dictionaries with the same keys as the original installment object,
    # so the template can still use inst.xxx without changes.
    schedule = []
    for inst in loan.installments.all().order_by('due_date'):
        # Calculated penalty from the rule
        calc_penalty = calculate_penalty(inst) or Decimal('0.00')

        # Manual penalties (not waived) for this installment
        manual_total = inst.manual_penalties.filter(is_waived=False).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        total_penalty = calc_penalty + manual_total

        # Balances
        principal_bal = inst.principal_balance
        interest_bal = inst.interest_balance
        total_balance = principal_bal + interest_bal + total_penalty

        schedule.append({
            'id': inst.id,
            'due_date': inst.due_date,
            'principal_portion': inst.principal_portion,
            'interest_portion': inst.interest_portion,
            'penalty_amount': total_penalty,          # <-- this will show combined penalty
            'balance': total_balance,                 # <-- this will include penalty
            'paid': inst.paid,
            'is_overdue': inst.is_overdue,
        })

    # Manual penalties (for the card)
    manual_penalties = loan.manual_penalties.filter(is_waived=False).order_by('-applied_date')
    total_manual_penalty = manual_penalties.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    sms_config = SMSConfig.objects.first()
    sms_balance = sms_config.balance if sms_config else 0

    context = {
        'loan': loan,
        'sms_balance': sms_balance,
        'principal_balance': loan.principal_balance,
        'interest_due': interest_due,
        'principal_due': principal_due,
        'total_due_now': total_due_now,
        'schedule': schedule,                         # <-- reusing the same variable name
        'repayments': loan.repayments.all().order_by('-date_paid'),
        'total_paid': total_paid.quantize(Decimal('0.01')),
        'total_payable': total_payable,
        'disbursement_date': disbursement_date,
        'end_date': end_date,
        'today': today,
        'manual_penalties': manual_penalties,
        'total_manual_penalty': total_manual_penalty,
    }

    return render(request, 'finance/loan_detail.html', context)
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

# ========================
# Local Imports
# ========================
from .models import Loan, generate_schedule
from .utils import generate_loan_ref   # Make sure this exists in finance/utils.py


from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

# Local imports
from .models import Loan, generate_schedule
from .utils import generate_loan_ref


from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from .models import Loan, generate_schedule
from .utils import generate_loan_ref


@login_required
def apply_loan(request, member_id=None):
    """
    Clean & Modern Loan Application View - Uses term_value + repayment_frequency
    Integrated with LoanPenaltyRule for flexible penalties.
    """
    members = Member.objects.all().order_by('first_name')
    selected_member = None

    if member_id:
        selected_member = get_object_or_404(Member, id=member_id)

    if request.method == "POST":
        try:
            member = get_object_or_404(Member, id=request.POST.get('member') or member_id)

            # === Required Fields ===
            principal = Decimal(request.POST.get('principal_amount') or '0')
            interest_rate = Decimal(request.POST.get('interest_rate') or '0')
            term_value = int(request.POST.get('term_value') or 1)
            repayment_frequency = request.POST.get('repayment_frequency', 'monthly')

            # Validation
            if principal <= 0:
                messages.error(request, "Principal amount must be greater than zero.")
                return render(request, 'finance/apply_loan.html', {
                    'members': members, 'selected_member': selected_member
                })

            if term_value < 1:
                messages.error(request, "Term Value must be at least 1.")
                return render(request, 'finance/apply_loan.html', {
                    'members': members, 'selected_member': selected_member
                })

            start_date_str = request.POST.get('start_date')
            start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else timezone.now().date()

            with transaction.atomic():
                ref_code = generate_loan_ref()

                # --- FIXED: Flat interest calculation (no multiplication by term_value) ---
                # Interest = Principal × (Rate / 100)
                total_interest = (principal * (interest_rate / Decimal('100'))).quantize(Decimal('0.01'))
                total_payable = principal + total_interest

                # --- Create Loan (deprecating old penalty fields) ---
                loan = Loan(
                    member=member,
                    officer=request.user,
                    loan_reference=ref_code,

                    principal_amount=principal,
                    interest_rate=interest_rate,
                    term_value=term_value,
                    repayment_frequency=repayment_frequency,
                    start_date=start_date,

                    total_payable=total_payable,
                    principal_balance=principal,
                    interest_balance=total_interest,

                    status='pending',
                    is_active=False,

                    # Other fields
                    product_type=request.POST.get('product_type', 'personal'),
                    purpose=request.POST.get('purpose', ''),

                    guarantor_1_name=request.POST.get('guarantor_1_name', ''),
                    guarantor_1_phone=request.POST.get('guarantor_1_phone', ''),
                    guarantor_2_name=request.POST.get('guarantor_2_name') or None,
                    guarantor_2_phone=request.POST.get('guarantor_2_phone') or None,

                    collateral_type=request.POST.get('collateral_type', ''),
                    collateral_value=Decimal(request.POST.get('collateral_value') or '0'),
                    collateral_description=request.POST.get('collateral_description', ''),
                    location=request.POST.get('location', ''),
                    contact_person=request.POST.get('contact_person', ''),
                    contact_phone=request.POST.get('contact_phone', ''),

                    # Old penalty fields (will be overridden by LoanPenaltyRule)
                    penalty_type=request.POST.get('penalty_type', 'daily_flat'),
                    penalty_rate=Decimal(request.POST.get('penalty_rate') or '1.0'),
                    penalty_flat_amount=Decimal(request.POST.get('penalty_flat_amount') or '1000'),
                    penalty_grace_days=int(request.POST.get('penalty_grace_days') or '0'),

                    notes=request.POST.get('notes', ''),
                )

                loan.save()

                # --- Create the Penalty Rule ---
                penalty_type = request.POST.get('penalty_type', 'daily_flat')
                penalty_rate = Decimal(request.POST.get('penalty_rate') or '0')
                penalty_flat_amount = Decimal(request.POST.get('penalty_flat_amount') or '1000')
                penalty_grace_days = int(request.POST.get('penalty_grace_days') or '0')
                max_penalty_cap = Decimal(request.POST.get('max_penalty_cap') or '0')
                compound = request.POST.get('compound') == 'true'

                # Map period (default to monthly, but could be derived from frequency)
                frequency_to_period = {
                    'monthly': 'monthly',
                    'weekly': 'weekly',
                    'daily': 'daily',
                    'manual': 'monthly',  # fallback
                }
                period = frequency_to_period.get(repayment_frequency, 'monthly')

                # Map penalty_type to the rule's choices
                rule_penalty_type = penalty_type
                if penalty_type == 'compound':
                    rule_penalty_type = 'percentage'  # but set compound True

                penalty_rule = LoanPenaltyRule.objects.create(
                    loan=loan,
                    penalty_type=rule_penalty_type,
                    period=period,
                    fixed_amount=penalty_flat_amount,
                    percentage_rate=penalty_rate,
                    grace_period_days=penalty_grace_days,
                    max_penalty_cap=max_penalty_cap,
                    compound=compound,
                )

                # Generate installments (splits total interest equally across term)
                generate_schedule(loan)

                messages.success(request, f"Loan {ref_code} created successfully!")
                return redirect('dashboard')

        except ValueError as ve:
            messages.error(request, f"Invalid data entered: {str(ve)}")
        except Exception as e:
            messages.error(request, f"Error creating loan: {str(e)}")

    # GET Request
    context = {
        'members': members,
        'selected_member': selected_member,
        'product_choices': Loan.PRODUCT_CHOICES,
    }
    return render(request, 'finance/apply_loan.html', context)
@login_required
@transaction.atomic
def approve_loan(request, pk, action):
    """
    Handles atomic execution of loan validation, state transitions, and immediate savings disbursement.
    """
    loan = get_object_or_404(Loan.objects.select_for_update(), pk=pk)
    
    if action == 'approve':
        if loan.status == 'approved' or loan.is_active:
            messages.info(request, "This loan has already been approved and disbursed.")
            return redirect('loan_detail', pk=loan.id)

        try:
            savings = loan.member.savings  
            savings = type(savings).objects.select_for_update().get(id=savings.id)
            principal = Decimal(str(loan.principal_amount))

            # Update status safely
            loan.status = 'approved'
            loan.is_active = True
            if not loan.disbursed_date:
                loan.disbursed_date = timezone.now().date()
            loan.save()

            # Execute underlying calculations
            if not loan.installments.exists():
               generate_schedule(loan)
            # Route currency pool to member portfolio
            savings.balance += principal
            savings.save()

            ref = generate_transaction_ref("DSB")
            Transaction.objects.create(
                member=loan.member,
                amount=principal,
                type='disbursement',  
                reference=ref
            )

            messages.success(
                request, 
                f"Loan {loan.id} approved successfully. Reference {ref}: UGX {principal:,.0f} disbursed to savings."
            )

        except AttributeError:
            messages.error(request, "Approval failed: Member has no active savings account.")
            return redirect('loan_detail', pk=loan.id)
        except Exception as e:
            messages.error(request, f"Error during approval and disbursement: {str(e)}")
            return redirect('loan_detail', pk=loan.id)

    elif action == 'reject':
        if loan.status != 'pending':
            messages.error(request, "Only pending loans can be rejected.")
            return redirect('loan_detail', pk=loan.id)
            
        loan.status = 'rejected'
        loan.is_active = False
        loan.save()
        messages.warning(request, f"Loan {loan.id} has been rejected.")

    return redirect('loan_detail', pk=loan.id)


from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Member, SavingsAccount, SystemSetting, Loan
from .services import FinancialTransactionService # Ensure this is created
from .utils import generate_transaction_ref


# finance/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction as db_transaction
from django.utils import timezone
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import uuid
import logging

from .models import Member, Loan, SystemSetting, Transaction, SavingsAccount
from .services import FinancialTransactionService, process_repayment, generate_transaction_ref

logger = logging.getLogger(__name__)


# finance/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction as db_transaction
from django.utils import timezone
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import logging

from .models import Member, Loan, SystemSetting, Transaction, SavingsAccount, Installment
from .services import FinancialTransactionService, process_repayment, generate_transaction_ref

logger = logging.getLogger(__name__)

# finance/views.py

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone

from .models import (
    Member, SavingsAccount, Transaction, SystemSetting, 
    Company, Loan, Installment
)
from .services import (
    FinancialTransactionService, 
    generate_transaction_ref, 
    process_repayment
)

# finance/views.py

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone

from .models import (
    Member, SavingsAccount, Transaction, SystemSetting, 
    Company, Loan, Installment
)
from .services import (
    FinancialTransactionService, 
    generate_transaction_ref, 
    process_repayment
)


# ============================================================
# RECEIPT VIEW (new)
# ============================================================

@login_required
def view_receipt(request):
    """
    Display the receipt stored in session.
    Expects 'deposit_receipt' or 'withdrawal_receipt' in session.
    """
    receipt = request.session.get('deposit_receipt') or request.session.get('withdrawal_receipt')
    
    if not receipt or not receipt.get('show'):
        messages.warning(request, "No receipt to display.")
        return redirect('dashboard')  # fallback URL
    
    # Clear the session flag after displaying
    if 'deposit_receipt' in request.session:
        del request.session['deposit_receipt']
    if 'withdrawal_receipt' in request.session:
        del request.session['withdrawal_receipt']
    
    context = {
        'receipt': receipt['data'],
        'company': Company.get_company(),  # ensure this method exists
    }
    return render(request, 'finance/receipt.html', context)


# ============================================================
# DEPOSIT VIEW
# ============================================================

@login_required
@transaction.atomic
def deposit_savings(request, member_id):
    """
    Handles member deposits with high-precision Decimal math,
    atomic ledger updates, and auto-sweep recovery.
    """
    print("=" * 60)
    print("DEPOSIT SAVINGS FUNCTION STARTED")
    print(f"Member ID: {member_id}")
    print("=" * 60)
    
    member = get_object_or_404(Member, id=member_id)
    print(f"Member found: {member.get_full_name()} (ID: {member.id})")
    
    backdate_allowed = SystemSetting.is_backdate_allowed()
    print(f"Backdate allowed: {backdate_allowed}")
    
    # Get previous balance for receipt
    previous_balance = Decimal('0')
    if hasattr(member, 'savings') and member.savings:
        previous_balance = member.savings.balance
        print(f"Previous balance: {previous_balance}")
    else:
        print("No savings account found for member")
    
    # ---- FETCH COMPANY ONCE ----
    company = Company.get_company()
    
    if request.method == "POST":
        print("\n--- POST Request Received ---")
        print(f"POST data: {request.POST}")
        
        amount_raw = request.POST.get('amount', '0').strip()
        custom_date = request.POST.get('back_date')
        print(f"Amount raw: {amount_raw}")
        print(f"Custom date: {custom_date}")
        
        # Validate amount
        if not amount_raw:
            print("ERROR: No amount provided")
            messages.error(request, "Please enter a deposit amount.")
            return redirect('deposit_savings', member_id=member.id)
        
        try:
            amount = Decimal(amount_raw)
            print(f"Amount converted to Decimal: {amount}")
        except (ValueError, InvalidOperation, TypeError) as e:
            print(f"ERROR converting amount: {e}")
            messages.error(request, "Please enter a valid number for the deposit amount.")
            return redirect('deposit_savings', member_id=member.id)
        
        # Quantize to 2 decimal places
        amount = amount.quantize(Decimal('1.00'), rounding=ROUND_HALF_UP)
        print(f"Amount quantized: {amount}")
        
        if amount <= 0:
            print("ERROR: Amount is zero or negative")
            messages.error(request, "Deposit amount must be greater than zero.")
            return redirect('deposit_savings', member_id=member.id)
        
        try:
            # Generate transaction reference
            ref = generate_transaction_ref("DEP")
            print(f"Generated reference: {ref}")
            
            txn_timestamp = custom_date if (backdate_allowed and custom_date) else timezone.now()
            print(f"Transaction timestamp: {txn_timestamp}")
            
            # Execute Atomic Transaction via Service Layer
            print("Calling FinancialTransactionService.record_deposit...")
            transaction_obj = FinancialTransactionService.record_deposit(
                member=member,
                amount=amount,
                reference=ref,
                date=txn_timestamp,
                created_by=request.user
            )
            print(f"Deposit recorded successfully. Transaction ID: {transaction_obj.id}")
            
            # Refresh member to get updated balance
            member.refresh_from_db()
            new_balance = Decimal('0')
            if hasattr(member, 'savings') and member.savings:
                new_balance = member.savings.balance
                print(f"New balance: {new_balance}")
            
            # Check for Arrears/Auto-Sweep Recovery
            arrears_cleared = Decimal('0')
            active_loan = Loan.objects.filter(member=member, is_active=True).first()
            
            if active_loan:
                print(f"Active loan found: {active_loan.loan_reference}")
                overdue_installments = Installment.objects.filter(
                    loan=active_loan,
                    paid=False,
                    due_date__lte=timezone.now().date()
                )
                overdue_exists = overdue_installments.exists()
                print(f"Overdue installments exist: {overdue_exists}")
                
                if overdue_exists:
                    print("Processing repayment...")
                    result = process_repayment(active_loan.id)
                    print(f"Repayment result: {result}")
                    arrears_cleared = amount
                    messages.info(request, f"Deposit {ref} recorded. Arrears detected; auto-repayment triggered.")
                else:
                    print("No overdue installments found")
                    messages.success(request, f"Deposit {ref} of UGX {amount:,.0f} processed successfully.")
            else:
                print("No active loan found")
                messages.success(request, f"Deposit {ref} of UGX {amount:,.0f} processed successfully.")
            
            # Prepare receipt data (adds extra fields for receipt template)
            receipt_data = {
                'receipt_id': str(ref),
                'date': txn_timestamp.strftime('%d %b, %Y %H:%M'),
                'member_name': f"{member.first_name} {member.last_name}",
                'member_id': str(member.member_number or member.id),
                'member_pk': member.id,                     # for back button
                'processed_by': request.user.get_full_name() or request.user.username,
                'amount': str(amount),
                'prev_balance': str(previous_balance),
                'new_balance': str(new_balance),
                'arrears_cleared': str(arrears_cleared),
                'payment_method': 'Cash Deposit',
                'status': 'COMPLETED',
                'type': 'deposit',
                'timestamp': txn_timestamp.isoformat(),
                'description': 'Savings Deposit' + (' (with auto-repayment)' if arrears_cleared > 0 else ''),
            }
            
            print(f"Receipt data prepared: {receipt_data}")
            
            # Store receipt in session
            request.session['deposit_receipt'] = {
                'data': receipt_data,
                'show': True
            }
            request.session.modified = True
            print("Receipt stored in session")
            
            print("=" * 60)
            print("DEPOSIT COMPLETED SUCCESSFULLY - Redirecting to receipt")
            print("=" * 60)
            
            # Redirect to receipt view instead of profile
            return redirect('view_receipt')

        except Exception as e:
            print(f"\n!!! EXCEPTION OCCURRED !!!")
            print(f"Error type: {type(e).__name__}")
            print(f"Error message: {str(e)}")
            import traceback
            print(f"Traceback: {traceback.format_exc()}")
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('deposit_savings', member_id=member.id)
    
    # GET request – render the deposit form
    print("Rendering deposit form (GET request)")
    return render(request, 'finance/deposit.html', {
        'member': member,
        'backdate_allowed': backdate_allowed,
        'company': company,
    })


# ============================================================
# WITHDRAWAL VIEW
# ============================================================

@login_required
@transaction.atomic
def withdraw_savings(request, member_id):
    """
    Handles member withdrawals with atomic ledger updates.
    """
    member = get_object_or_404(Member, id=member_id)
    savings = SavingsAccount.objects.select_for_update().get_or_create(member=member)[0]
    backdate_allowed = SystemSetting.is_backdate_allowed()
    previous_balance = savings.balance  # capture before withdrawal

    if request.method == 'POST':
        amount_raw = request.POST.get('amount', '0').strip()
        custom_date = request.POST.get('back_date')
        
        try:
            amount = Decimal(amount_raw)
            if amount <= 0:
                messages.error(request, "Withdrawal amount must be greater than zero.")
            elif savings.balance < amount:
                messages.error(request, f"Insufficient funds. Current balance: UGX {savings.balance:,.0f}")
            else:
                ref = generate_transaction_ref("WTH")
                txn_timestamp = custom_date if (backdate_allowed and custom_date) else timezone.now()

                # Delegate to the Service Layer for Ledger integrity
                FinancialTransactionService.record_withdrawal(
                    member=member,
                    amount=amount,
                    reference=ref,
                    date=txn_timestamp
                )

                # Refresh to get new balance
                member.refresh_from_db()
                new_balance = member.savings.balance if hasattr(member, 'savings') else Decimal('0')

                # Prepare receipt data
                receipt_data = {
                    'receipt_id': ref,
                    'date': txn_timestamp.strftime('%d %b, %Y %H:%M'),
                    'member_name': f"{member.first_name} {member.last_name}",
                    'member_id': str(member.member_number or member.id),
                    'member_pk': member.id,
                    'processed_by': request.user.get_full_name() or request.user.username,
                    'amount': str(amount),
                    'prev_balance': str(previous_balance),
                    'new_balance': str(new_balance),
                    'arrears_cleared': '0',   # no arrears cleared on withdrawal
                    'payment_method': 'Cash Withdrawal',
                    'status': 'COMPLETED',
                    'type': 'withdrawal',
                    'timestamp': txn_timestamp.isoformat(),
                    'description': 'Savings Withdrawal',
                }

                # Store receipt in session
                request.session['withdrawal_receipt'] = {
                    'data': receipt_data,
                    'show': True
                }
                request.session.modified = True

                messages.success(request, f"Withdrawal {ref} of UGX {amount:,.0f} successful.")
                # Redirect to receipt view
                return redirect('view_receipt')

        except Exception as e:
            messages.error(request, f"Error processing withdrawal: {str(e)}")

    return render(request, 'finance/withdraw_form.html', {
        'member': member,
        'savings': savings,
        'backdate_allowed': backdate_allowed
    })


@login_required
def member_statement(request, member_id):
    """Detailed individual ledger statement"""
    member = get_object_or_404(Member, id=member_id)
    transactions = Transaction.objects.filter(member=member).order_by('-timestamp')
    savings = SavingsAccount.objects.filter(member=member).first()
    
    return render(request, 'finance/statement.html', {
        'member': member,
        'transactions': transactions,
        'savings': savings
    })


@login_required
@transaction.atomic
def receive_payment(request, loan_id):
    """
    Accepts external user payment injections, writes ledger records,
    and runs the allocation calculation module.
    """
    if request.method != "POST":
        return redirect('loan_detail', pk=loan_id)

    loan = get_object_or_404(Loan.objects.select_for_update(), id=loan_id)
    backdate_allowed = SystemSetting.is_backdate_allowed()

    if loan.status not in ['approved', 'arrears']:
        messages.error(request, "Repayments are only accepted for active loans.")
        return redirect('loan_detail', pk=loan_id)

    try:
        current_balance = Decimal(str(loan.principal_balance + loan.interest_balance))
        if current_balance <= 0:
            messages.warning(request, "This loan is already fully paid.")
            return redirect('loan_detail', pk=loan_id)

        principal = Decimal(request.POST.get('principal', '0').strip() or '0')
        interest = Decimal(request.POST.get('interest', '0').strip() or '0')
        penalty = Decimal(request.POST.get('penalty', '0').strip() or '0')
        custom_date = request.POST.get('back_date')
        notes = request.POST.get('notes', '').strip()

        total_payment = principal + interest + penalty

        if total_payment <= 0:
            messages.error(request, "Total payment must be greater than zero.")
            return redirect('loan_detail', pk=loan_id)

        txn_timestamp = timezone.now()
        if backdate_allowed and custom_date:
            txn_timestamp = custom_date

        ref = generate_transaction_ref("PAY")

        # 1. Instantiation of Repayment tracking model
        Repayment.objects.create(
            loan=loan,
            amount_paid=total_payment,
            receipt_number=ref,
            date_paid=txn_timestamp, 
            notes=notes if notes else None,
        )

        # 2. Append general financial audit line
        Transaction.objects.create(
            member=loan.member,
            amount=total_payment,
            type='repayment',
            reference=ref,
            timestamp=txn_timestamp
        )

        # 3. CRITICAL ENGINE TRIGGER: Run split balance allocation calculations 
        process_repayment(loan.id)

        messages.success(request, f"Payment {ref} recorded and allocated successfully.")

    except (ValueError, InvalidOperation):
        messages.error(request, "Invalid payment amounts entered.")
    except Exception as e:
        messages.error(request, f"Error processing payment execution: {str(e)}")

    return redirect('loan_detail', pk=loan_id)


@login_required
def arrears_report(request):
    """Portfolio at Risk (PAR) Report"""
    today = timezone.now().date()
    overdue_installments = Installment.objects.filter(
        paid=False, 
        due_date__lt=today
    ).select_related('loan__member').order_by('due_date')

    total_at_risk = overdue_installments.aggregate(
        total_sum=Sum(F('principal_portion') + F('interest_portion') + F('penalty_amount'))
    )['total_sum'] or 0

    return render(request, 'finance/arrears.html', {
        'overdue': overdue_installments,
        'total_at_risk': total_at_risk,
        'today': today
    })
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from .models import Loan
from members.models import Member
from .models import SavingsAccount # Assuming this is your savings model path
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from .models import Member, Loan

import string
import random
from django.contrib.auth.decorators import login_required

import string
import random
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required
from .models import Member, Loan

import string
import random
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required
from .models import Member, Loan

def generate_loan_ref(length=10):
    """Generates a random uppercase alphanumeric string"""
    chars = string.ascii_uppercase + string.digits
    return ''.join(random.choices(chars, k=length))

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from decimal import Decimal
from .models import Loan, Member
# Ensure you have your generate_loan_ref helper imported



from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Loan



from datetime import date
from dateutil.relativedelta import relativedelta
from decimal import Decimal

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from .models import Loan, Repayment
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from .models import Loan, Repayment, SavingsAccount

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from .models import Loan, Repayment, SavingsAccount

from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from decimal import Decimal

from .models import Loan, SavingsAccount


from decimal import Decimal
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.db.models import Sum
from decimal import Decimal, InvalidOperation
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.db.models import Sum
from decimal import Decimal, InvalidOperation
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.db.models import Sum

from decimal import Decimal, InvalidOperation
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.db.models import Sum
from dateutil.relativedelta import relativedelta   # Add this if not installed: pip install python-dateutil

from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta

from .models import Loan, ManualPenalty
from .utils import generate_schedule
from finance.penalties import calculate_penalty

from decimal import Decimal
from datetime import timedelta
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta

from .models import Loan, ManualPenalty
from .utils import generate_schedule
from finance.penalties import calculate_penalty

@login_required
def loan_detail(request, pk):
    """
    Comprehensive loan detail view – displays full financial summary,
    amortization schedule with combined penalties, manual penalties list,
    and repayment history.
    """
    loan = get_object_or_404(Loan.objects.select_related('member', 'officer'), pk=pk)
    today = timezone.now().date()

    # ---- 1. Generate schedule if missing ----
    if not loan.installments.exists():
        generate_schedule(loan)

    # ---- 2. Compute due amounts for the banner ----
    active_due = loan.installments.filter(paid=False, due_date__lte=today).aggregate(
        total_interest=Sum('interest_portion'),
        total_principal=Sum('principal_portion')
    )
    interest_due = active_due['total_interest'] or Decimal('0.00')
    principal_due = active_due['total_principal'] or Decimal('0.00')
    total_due_now = (interest_due + principal_due).quantize(Decimal('0.01'))

    # ---- 3. Total paid & total payable ----
    total_paid = loan.repayments.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    total_payable = loan.total_payable or Decimal('0')

    # ---- 4. Dates ----
    disbursement_date = loan.disbursed_date or loan.start_date
    term_value = getattr(loan, 'term_value', 0)
    frequency = getattr(loan, 'repayment_frequency', 'monthly')
    period_months = getattr(loan, 'period_months', 0)

    # Compute end date based on repayment frequency
    if disbursement_date and term_value > 0:
        if frequency == 'monthly':
            end_date = disbursement_date + relativedelta(months=term_value)
        elif frequency == 'weekly':
            end_date = disbursement_date + timedelta(weeks=term_value)
        elif frequency == 'daily':
            end_date = disbursement_date + timedelta(days=term_value)
        else:  # manual or fallback
            if period_months:
                end_date = disbursement_date + relativedelta(months=period_months)
            else:
                end_date = None
    else:
        end_date = None

    # ---- 5. Build schedule_data (combines calculated + manual penalties) ----
    schedule_data = []
    for inst in loan.installments.all().order_by('due_date'):
        calc_penalty = calculate_penalty(inst) or Decimal('0.00')
        manual_total = inst.manual_penalties.filter(is_waived=False).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        total_penalty = calc_penalty + manual_total

        principal_bal = inst.principal_balance
        interest_bal = inst.interest_balance
        total_balance = principal_bal + interest_bal + total_penalty

        schedule_data.append({
            'id': inst.id,
            'due_date': inst.due_date,
            'principal_portion': inst.principal_portion,
            'interest_portion': inst.interest_portion,
            'penalty_amount': total_penalty,
            'balance': total_balance,
            'paid': inst.paid,
            'is_overdue': inst.is_overdue,
        })

    # ---- 6. Manual penalties (active, not waived) ----
    manual_penalties = loan.manual_penalties.filter(is_waived=False).order_by('-applied_date')
    total_manual_penalty = manual_penalties.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # ---- 7. Recent repayments ----
    repayments = loan.repayments.all().order_by('-date_paid')[:10]

    # ---- 8. Additional context (savings, address) ----
    savings_balance = getattr(loan.member, 'savings', None)
    savings_balance = savings_balance.balance if savings_balance else Decimal('0.00')
    member = loan.member
    member_address = f"{member.village}, {member.parish}, {member.district}".strip(', ')

    # ---- 9. Build context ----
    context = {
        'loan': loan,
        'principal_balance': loan.principal_balance,
        'interest_balance': loan.interest_balance,
        'interest_due': interest_due,
        'principal_due': principal_due,
        'total_due_now': total_due_now,
        'schedule_data': schedule_data,
        'repayments': repayments,
        'total_paid': total_paid.quantize(Decimal('0.01')),
        'total_payable': total_payable,
        'disbursement_date': disbursement_date,
        'end_date': end_date,
        'today': today,
        'manual_penalties': manual_penalties,
        'total_manual_penalty': total_manual_penalty,
        'savings_balance': savings_balance,
        'member_address': member_address,
        'officer': loan.officer,
        'period_months': period_months,  # fallback value
    }

    return render(request, 'finance/loan_detail.html', context)

from django.shortcuts import render
from django.db.models import Sum
from .models import Loan

def loan_list(request):
    """Master Loan Registry with split-balance analytics"""
    loans = Loan.objects.select_related('member').all().order_by('-id')
    total_disbursed = loans.aggregate(Sum('principal_amount'))['principal_amount__sum'] or 0
    
    portfolio_stats = loans.aggregate(
        p_total=Sum('principal_balance'),
        i_total=Sum('interest_balance')
    )
    total_outstanding = (portfolio_stats['p_total'] or 0) + (portfolio_stats['i_total'] or 0)
    active_count = loans.filter(is_active=True).count()

    context = {
        'loans': loans,
        'total_disbursed': total_disbursed,
        'total_outstanding': total_outstanding,
        'active_count': active_count,
    }
    return render(request, 'finance/loan_list.html', context)

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.db import transaction
from .models import Loan, Repayment

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from decimal import Decimal

import decimal
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from .models import Member, Loan, Repayment

from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.core.exceptions import ValidationError

from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.core.exceptions import ValidationError

from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from.utils import generate_transaction_ref

@login_required
@transaction.atomic
def receive_payment(request, loan_id):
    if request.method != "POST":
        return redirect('loan_detail', pk=loan_id)

    loan = get_object_or_404(Loan.objects.select_for_update(), id=loan_id)
    
    # Permission check: Only staff can backdate, and only if global setting is ON
    backdate_allowed = SystemSetting.is_backdate_allowed()

    if loan.status != 'approved' and loan.status != 'arrears':
        messages.error(request, "Repayments are only accepted for active loans.")
        return redirect('loan_detail', pk=loan_id)

    try:
        current_balance = Decimal(str(loan.principal_balance + loan.interest_balance))
        if current_balance <= 0:
            messages.warning(request, "This loan is already fully paid.")
            return redirect('loan_detail', pk=loan_id)

        # Get form data
        principal = Decimal(request.POST.get('principal', '0').strip() or '0')
        interest = Decimal(request.POST.get('interest', '0').strip() or '0')
        penalty = Decimal(request.POST.get('penalty', '0').strip() or '0')
        custom_date = request.POST.get('back_date')
        notes = request.POST.get('notes', '').strip()

        total_payment = principal + interest + penalty

        if total_payment <= 0:
            messages.error(request, "Total payment must be greater than zero.")
            return redirect('loan_detail', pk=loan_id)

        # Determine Payment Date
        txn_timestamp = timezone.now()
        if backdate_allowed and custom_date:
            txn_timestamp = custom_date

        ref = generate_transaction_ref("PAY")

        # 1. Create the repayment record 
        # (Pass date_paid to override the default now())
        Repayment.objects.create(
            loan=loan,
            amount_paid=total_payment,
            receipt_number=ref,
            date_paid=txn_timestamp, 
            notes=notes if notes else None,
        )

        # 2. Create Transaction log
        Transaction.objects.create(
            member=loan.member,
            amount=total_payment,
            type='repayment',
            reference=ref,
            timestamp=txn_timestamp
        )

        messages.success(request, f"Payment {ref} recorded for date: {txn_timestamp}.")

    except (ValueError, InvalidOperation):
        messages.error(request, "Invalid payment amounts entered.")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect('loan_detail', pk=loan_id)


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from .models import Loan, Transaction, SavingsAccount

@transaction.atomic
def update_loan_status(request, pk, action):
    """
    Handles approval or rejection of a loan with unique disbursement referencing.
    """
    loan = get_object_or_404(Loan, pk=pk)
    
    if action == 'approve':
        if loan.status == 'approved':
            messages.info(request, "This loan has already been approved.")
            return redirect('loan_detail', pk=loan.id)

        try:
            with transaction.atomic():
                # 1. Update Loan Status
                loan.status = 'approved'
                loan.is_active = True
                if not loan.disbursed_date:
                    loan.disbursed_date = timezone.now().date()
                
                loan.save()

                # 2. Generate Repayment Schedule
                # This ensures installments start with the correct amount_remaining
                generate_schedule(loan)

                # 3. Disburse principal to member's savings
                savings = loan.member.savings
                principal = Decimal(str(loan.principal_amount))
                
                # Using select_for_update here is a good safety measure if not already in the model
                savings.balance += principal
                savings.save()

                # 4. Record disbursement transaction with UNIQUE REFERENCE
                # We use the 'DSB' prefix for disbursements
                ref = generate_transaction_ref("DSB")
                Transaction.objects.create(
                    member=loan.member,
                    amount=principal,
                    type='disbursement',
                    reference=ref
                )

                messages.success(
                    request, 
                    f"Loan {loan.id} approved successfully. "
                    f"Reference {ref}: UGX {principal:,.0f} disbursed to savings."
                )

        except AttributeError:
            messages.error(request, "Approval failed: Member has no savings account.")
            return redirect('loan_detail', pk=loan.id)

        except Exception as e:
            messages.error(request, f"Error during approval: {str(e)}")
            return redirect('loan_detail', pk=loan.id)

    elif action == 'reject':
        loan.status = 'rejected'
        loan.is_active = False
        loan.save()
        messages.warning(request, f"Loan {loan.id} has been rejected.")

    return redirect('loan_detail', pk=loan.id)

from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.contrib import messages
from .utils import send_bulk_arrears_reminders


@login_required
def reject_loan(request, loan_id):
    """
    Rejects a pending loan application.
    """
    loan = get_object_or_404(Loan, id=loan_id)
    if loan.status == 'pending':
        loan.status = 'rejected'
        loan.is_active = False
        loan.save()
        messages.error(request, f"Loan application #LN-{loan.id} has been rejected.")
    else:
        messages.warning(request, "Only pending loans can be rejected.")
        
    return redirect('member_profile', member_id=loan.member.id)



from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.contrib import messages
from .utils import send_bulk_arrears_reminders


@login_required
def bulk_sms_reminder_view(request):
    """Bulk send arrears reminders"""
    sent, failed = send_bulk_arrears_reminders(request)

    if sent > 0:
        messages.success(request, f"Successfully sent {sent} arrears reminders.")
    if failed > 0:
        messages.warning(request, f"Failed to send {failed} reminders.")
    if sent == 0 and failed == 0:
        messages.info(request, "No members with arrears were found.")

    return redirect('arrears_report')   # Change to your preferred redirect



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from .models import SavingsAccount, Transaction, Member
from decimal import Decimal, ROUND_HALF_UP

from decimal import Decimal, InvalidOperation
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone
from .models import Member, SavingsAccount, Transaction, SystemSetting
# Assuming these utility functions exist in your project
# from .utils import generate_transaction_ref 


# finance/views.py
from django.shortcuts import render
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Q
from django.db.models.functions import Coalesce
from datetime import datetime
from dateutil.relativedelta import relativedelta
from .models import Loan, Member, SavingsAccount, Transaction, Installment, Repayment, ChartOfAccount
from django.contrib.auth.decorators import login_required

from django.shortcuts import render
from django.db.models import Sum, F, Q
from django.contrib.auth.models import User
from decimal import Decimal
from datetime import date
from .models import Loan, SavingsAccount, Transaction, Installment


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.utils import timezone
from decimal import Decimal
from datetime import timedelta

from finance.models import (
    Loan, Installment, SavingsAccount, Company
)

@login_required
def reports_dashboard(request):
    """
    Central Reports Dashboard – Financial Intelligence.
    Computes real-time KPIs from the database.
    """
    today = timezone.now().date()
    company = Company.get_company()

    # ============================================================
    # 1. Gross Loan Portfolio (active loans only)
    # ============================================================
    loan_stats = Loan.objects.filter(is_active=True).aggregate(
        p_bal=Sum('principal_balance'),
        i_bal=Sum('interest_balance')
    )
    total_loan_portfolio = (loan_stats['p_bal'] or 0) + (loan_stats['i_bal'] or 0)

    # ============================================================
    # 2. Active Loan Accounts
    # ============================================================
    active_loans_count = Loan.objects.filter(is_active=True).count()

    # ============================================================
    # 3. Portfolio at Risk (PAR > 30 days)
    # ============================================================
    cutoff_date = today - timedelta(days=30)
    overdue_installments = Installment.objects.filter(
        loan__is_active=True,
        due_date__lt=cutoff_date,
        paid=False
    )

    # Sum the outstanding balance across overdue installments
    # Using the same formula as Installment.balance property
    par_30_balance = overdue_installments.aggregate(
        total=Coalesce(
            Sum(
                F('principal_portion') - F('principal_paid') +
                F('interest_portion') - F('interest_paid') +
                F('penalty_amount') - F('penalty_paid')
            ),
            Value(Decimal('0.00'), output_field=DecimalField())
        )
    )['total'] or 0

    if total_loan_portfolio > 0:
        par_30 = (par_30_balance / total_loan_portfolio) * 100
    else:
        par_30 = 0

    # ============================================================
    # 4. PAR Change (month-over-month) – placeholder (0 for now)
    # ============================================================
    par_change = 0

    # ============================================================
    # 5. Liquidity Ratio (proxy: total savings / total loans)
    # ============================================================
    total_savings = SavingsAccount.objects.aggregate(
        total=Sum('balance')
    )['total'] or 0

    if total_loan_portfolio > 0:
        liquidity_ratio = (total_savings / total_loan_portfolio) * 100
    else:
        liquidity_ratio = 0

    # ============================================================
    # 6. Context – ready for the template
    # ============================================================
    context = {
        'par_30': par_30,
        'par_change': par_change,
        'liquidity_ratio': liquidity_ratio,
        'active_loans_count': active_loans_count,
        'total_loan_portfolio': total_loan_portfolio,
        'company': company,
        'today': today,
    }

    return render(request, 'finance/reports/reports_dashboard.html', context)


from django.db.models import Sum, F
from decimal import Decimal
from decimal import Decimal
from django.shortcuts import render
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from datetime import date
from django.contrib.auth.models import User

def loan_portfolio_report(request):
    """Fixed Loan Portfolio Report"""

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    officer_id = request.GET.get('officer')

    # Base Query
    loans = Loan.objects.select_related('member', 'officer').filter(
        status__in=['approved', 'active', 'closed']
    ).order_by('-disbursed_date', '-start_date')

    # Apply filters
    if start_date:
        loans = loans.filter(disbursed_date__gte=start_date)
    if end_date:
        loans = loans.filter(disbursed_date__lte=end_date)
    if officer_id:
        loans = loans.filter(officer_id=officer_id)

    today = date.today()
    report_data = []

    for loan in loans:
        p_bal = Decimal(str(loan.principal_balance or 0))
        i_bal = Decimal(str(loan.interest_balance or 0))

        # Arrears (overdue)
        overdue = loan.installments.filter(paid=False, due_date__lt=today)
        principal_in_arrears = overdue.aggregate(
            total=Coalesce(Sum('principal_portion'), Decimal('0'))
        )['total']

        # Total Due (including today)
        total_due_today = loan.installments.filter(
            paid=False, due_date__lte=today
        ).aggregate(
            total=Coalesce(Sum(F('principal_portion') + F('interest_portion')), Decimal('0'))
        )['total']

        penalty_due = overdue.aggregate(
            total=Coalesce(Sum('penalty_amount'), Decimal('0'))
        )['total']

        report_data.append({
            'borrower': f"{loan.member.first_name} {loan.member.last_name}",
            'officer': loan.officer.get_full_name() if loan.officer else 'System',
            'account_number': loan.member.member_number,
            'contact': loan.member.phone_number,
            'loan_disbursed': Decimal(str(loan.principal_amount or 0)),
            'disbursement_date': loan.disbursed_date or loan.start_date,
            'principal_balance': p_bal,
            'interest_balance': i_bal,
            'principal_in_arrears': principal_in_arrears,
            'total_dues': total_due_today + penalty_due,
            'par': p_bal if principal_in_arrears > 0 else Decimal('0'),
        })

    # ====================== CALCULATE TOTALS ======================
    total_disbursed = sum(item['loan_disbursed'] for item in report_data)
    total_outstanding = sum(item['principal_balance'] + item['interest_balance'] for item in report_data)
    total_par = sum(item['par'] for item in report_data)

    context = {
        'report_data': report_data,
        'today': today,
        'officers': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),

        # Summary Cards
        'total_disbursed': total_disbursed,
        'total_outstanding': total_outstanding,
        'total_par': total_par,

        # For backward compatibility with your old template if needed
        'total_p_bal': sum(item['principal_balance'] for item in report_data),
    }

    return render(request, 'finance/reports/loan_portfolio.html', context)

from django.db.models import Sum, Q, F
from decimal import Decimal
from datetime import date
from django.shortcuts import render
from django.db.models import Sum, F
from django.utils import timezone
from decimal import Decimal
from .models import Loan, Installment # Ensure User is imported if using for officers

from decimal import Decimal
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum
from django.db.models.functions import Coalesce

def portfolio_status_report(request):
    """
    Comprehensive Portfolio Status Report - Format
    """
    today = timezone.now().date()

    # Optimized Query
    loans = Loan.objects.select_related('member', 'officer').filter(
        status__in=['approved', 'active', 'closed']
    ).order_by('member__member_number')

    report_data = []

    for loan in loans:
        # 1. Paid Amounts (from paid installments)
        paid_stats = loan.installments.filter(paid=True).aggregate(
            p_paid=Coalesce(Sum('principal_portion'), Decimal('0.00')),
            i_paid=Coalesce(Sum('interest_portion'), Decimal('0.00')),
            penalty_paid=Coalesce(Sum('penalty_amount'), Decimal('0.00')),
        )

        # 2. Arrears (Unpaid installments due before today)
        arrears_stats = loan.installments.filter(
            paid=False,
            due_date__lt=today
        ).aggregate(
            p_due=Coalesce(Sum('principal_portion'), Decimal('0.00')),
            i_due=Coalesce(Sum('interest_portion'), Decimal('0.00')),
            pen_due=Coalesce(Sum('penalty_amount'), Decimal('0.00')),
        )

        # 3. Total Due (including today's due date)
        total_due_stats = loan.installments.filter(
            paid=False,
            due_date__lte=today
        ).aggregate(
            total_due=Coalesce(Sum('principal_portion') + Sum('interest_portion'), Decimal('0.00'))
        )

        # 4. Aging Classification
        oldest_unpaid = loan.installments.filter(
            paid=False, 
            due_date__lt=today
        ).order_by('due_date').first()

        classification = "Performing"
        if oldest_unpaid:
            days_past_due = (today - oldest_unpaid.due_date).days
            if days_past_due > 180:
                classification = "Loss"
            elif days_past_due > 90:
                classification = "Doubtful"
            elif days_past_due > 30:
                classification = "Substandard"
            else:
                classification = "Watch"

        report_data.append({
            'member_no': loan.member.member_number or str(loan.member.id),
            'name': f"{loan.member.first_name} {loan.member.last_name}",
            'loan_no': loan.loan_reference or loan.id,
            'disbursed_amount': Decimal(str(loan.principal_amount or 0)),
            'disbursed_date': loan.disbursed_date or loan.start_date,
            'principal_paid': paid_stats['p_paid'],
            'interest_paid': paid_stats['i_paid'],
            'penalty_paid': paid_stats['penalty_paid'],
            'principal_due': arrears_stats['p_due'],
            'interest_due': arrears_stats['i_due'],
            'penalty_due': arrears_stats['pen_due'],
            'total_due': arrears_stats['p_due'] + arrears_stats['i_due'] + arrears_stats['pen_due'],
            'principal_balance': Decimal(str(loan.principal_balance or 0)),
            'interest_balance': Decimal(str(loan.interest_balance or 0)),
            'classification': classification,
            'sector': getattr(loan.member, 'economic_sector', 'N/A'),
        })

    # Calculate Grand Totals for Footer
    grand_total_disbursed = sum(item['disbursed_amount'] for item in report_data)
    grand_total_prin_paid = sum(item['principal_paid'] for item in report_data)
    grand_total_int_paid = sum(item['interest_paid'] for item in report_data)
    grand_total_penalty_paid = sum(item['penalty_paid'] for item in report_data)

    grand_total_prin_due = sum(item['principal_due'] for item in report_data)
    grand_total_int_due = sum(item['interest_due'] for item in report_data)
    grand_total_penalty_due = sum(item['penalty_due'] for item in report_data)
    grand_total_due = sum(item['total_due'] for item in report_data)

    grand_total_prin_bal = sum(item['principal_balance'] for item in report_data)
    grand_total_int_bal = sum(item['interest_balance'] for item in report_data)
    grand_total_exposure = grand_total_prin_bal + grand_total_int_bal

    context = {
        'report_data': report_data,
        'today': today,

        # Summary Totals
        'total_principal_bal': grand_total_prin_bal,
        'total_arrears': grand_total_prin_due + grand_total_int_due + grand_total_penalty_due,

        # Grand Totals for Table Footer
        'grand_total_disbursed': grand_total_disbursed,
        'grand_total_prin_paid': grand_total_prin_paid,
        'grand_total_int_paid': grand_total_int_paid,
        'grand_total_penalty': grand_total_penalty_due,        # Usually we show unpaid penalty
        'grand_total_prin_due': grand_total_prin_due,
        'grand_total_int_due': grand_total_int_due,
        'grand_total_due': grand_total_due,
        'grand_total_prin_bal': grand_total_prin_bal,
        'grand_total_int_bal': grand_total_int_bal,
        'grand_total_exposure': grand_total_exposure,
    }

    return render(request, 'finance/reports/portfolio_status.html', context)

def calculate_aging(loan, today):
    """Determines classification based on days past due (DPD)"""
    oldest_unpaid = loan.installments.filter(paid=False, due_date__lt=today).order_by('due_date').first()
    if not oldest_unpaid:
        return "Performing"
    
    days_past_due = (today - oldest_unpaid.due_date).days
    if days_past_due <= 30: return "Watch"
    if days_past_due <= 90: return "Substandard"
    if days_past_due <= 180: return "Doubtful"
    return "Loss"

def savings_report(request):
    """Savings Report"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    txs = Transaction.objects.select_related('member')

    if start_date:
        txs = txs.filter(timestamp__date__gte=start_date)
    if end_date:
        txs = txs.filter(timestamp__date__lte=end_date)

    total_deposits = txs.filter(type='deposit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    total_withdrawals = txs.filter(type='withdrawal').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    current_total_savings = SavingsAccount.objects.aggregate(Sum('balance'))['balance__sum'] or Decimal('0')

    context = {
        'transactions': txs.order_by('-timestamp'),
        'total_deposits': total_deposits,
        'total_withdrawals': total_withdrawals,
        'current_total_savings': current_total_savings,
        'net_movement': total_deposits - total_withdrawals,
    }

    return render(request, 'finance/reports/savings_report.html', context)


from django.shortcuts import render
from django.db.models import Sum, Q
from decimal import Decimal

from django.shortcuts import render
from django.db.models import Sum
from decimal import Decimal
from .models import Transaction  # Ensure your imports are correct
from django.shortcuts import render
from django.db.models import Sum
from decimal import Decimal
from .models import Transaction

def cash_flow_statement(request):
    """
    Cash Flow Statement using double-entry paths:
    Transaction -> GeneralLedger -> ChartOfAccount
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # 1. Filter Transactions by type
    inflow_qs = Transaction.objects.filter(type__in=['deposit', 'repayment'])
    outflow_qs = Transaction.objects.filter(type__in=['withdrawal', 'disbursement'])

    # 2. Apply Date Filters
    if start_date:
        inflow_qs = inflow_qs.filter(timestamp__date__gte=start_date)
        outflow_qs = outflow_qs.filter(timestamp__date__gte=start_date)
    if end_date:
        inflow_qs = inflow_qs.filter(timestamp__date__lte=end_date)
        outflow_qs = outflow_qs.filter(timestamp__date__lte=end_date)

    # 3. Calculate Summary Totals
    total_inflows = inflow_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    total_outflows = outflow_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    net_cash_flow = total_inflows - total_outflows

    # 4. Group by ChartOfAccount via GeneralLedger
    # We use the full path: generalledger__account__name
    inflow_accounts = inflow_qs.values(
        'generalledger__account__code', 
        'generalledger__account__name'
    ).annotate(
        total_amount=Sum('amount')
    ).order_by('-total_amount')

    outflow_accounts = outflow_qs.values(
        'generalledger__account__code', 
        'generalledger__account__name'
    ).annotate(
        total_amount=Sum('amount')
    ).order_by('-total_amount')

    context = {
        'total_inflows': total_inflows,
        'total_outflows': total_outflows,
        'net_cash_flow': net_cash_flow,
        'inflow_accounts': inflow_accounts,
        'outflow_accounts': outflow_accounts,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'finance/reports/cash_flow.html', context)


def chart_of_accounts(request):
    """Chart of Accounts"""
    from .models import ChartOfAccount
    accounts = ChartOfAccount.objects.filter(is_active=True).order_by('code')
    
    context = {'accounts': accounts}
    return render(request, 'finance/reports/chart_of_accounts.html', context)



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from .models import Transaction, TransactionReversal, SavingsAccount

@login_required
@transaction.atomic
def reverse_transaction(request, transaction_id):
    # 1. Permission Check
    if not request.user.is_staff:
        messages.error(request, "You do not have permission to reverse transactions.")
        return redirect('dashboard')

    # 2. Get Transaction with Row-Level Locking
    txn = get_object_or_404(Transaction.objects.select_for_update(), id=transaction_id)
    
    if txn.is_reversed:
        messages.warning(request, "This transaction has already been reversed.")
        return redirect('member_profile', member_id=txn.member.id)

    if request.method == "POST":
        reason = request.POST.get('reason')
        if not reason:
            messages.error(request, "A reason for reversal is required.")
            return render(request, 'finance/reverse_confirm.html', {'txn': txn})

        # 3. Update Member Savings Balance
        savings = get_object_or_404(SavingsAccount.objects.select_for_update(), member=txn.member)
        
        # Determine the contra-type based on the original transaction
        # If they deposited, we must "withdraw" to fix balance, and vice-versa.
        if txn.type == 'deposit':
            savings.balance -= txn.amount
            contra_type = 'withdrawal'
        elif txn.type == 'withdrawal':
            savings.balance += txn.amount
            contra_type = 'deposit'
        else:
            # Handle other types like 'repayment' or 'penalty' if necessary
            messages.error(request, f"Reversal for type {txn.type} not configured.")
            return redirect('member_profile', member_id=txn.member.id)
        
        savings.save()

        # 4. Mark original transaction as reversed
        txn.is_reversed = True
        txn.save()

        # 5. Create the Contra-Entry (The "Correction" Transaction)
        # Note: We removed 'notes' because it's not in your model.
        # We use your existing choices ('deposit'/'withdrawal') for the type.
        Transaction.objects.create(
            member=txn.member,
            amount=txn.amount,
            type=contra_type, 
            reference=f"REV-{txn.id}",
            created_by=request.user
        )

        # 6. Create the Audit Log (The 'reason' is stored here)
        TransactionReversal.objects.create(
            original_transaction=txn,
            reversed_by=request.user,
            reason=reason
        )

        messages.success(request, f"Transaction reversed successfully. Balance updated.")
        return redirect('member_profile', member_id=txn.member.id)

    return render(request, 'finance/reverse_confirm.html', {'txn': txn})



from django.shortcuts import render, get_object_or_404
from .models import Loan, Transaction
from decimal import Decimal
from django.db.models import Sum, Q
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from dateutil.relativedelta import relativedelta

from .models import Loan, Transaction, ManualPenalty
from .utils import generate_schedule
from finance.penalties import calculate_penalty

@login_required
def loan_details(request, loan_id):
    """
    Comprehensive loan detail view – provides full financial summary,
    amortization schedule with combined penalties, manual penalties list,
    and recent repayment transactions.
    """
    loan = get_object_or_404(Loan.objects.select_related('member', 'officer'), id=loan_id)
    today = timezone.now().date()

    # 1. Generate schedule if missing
    if not loan.installments.exists():
        generate_schedule(loan)

    # 2. Compute due amounts for the banner
    active_due = loan.installments.filter(paid=False, due_date__lte=today).aggregate(
        total_interest=Sum('interest_portion'),
        total_principal=Sum('principal_portion')
    )
    interest_due = active_due['total_interest'] or Decimal('0.00')
    principal_due = active_due['total_principal'] or Decimal('0.00')
    total_due_now = (interest_due + principal_due).quantize(Decimal('0.01'))

    # 3. Total paid & total payable
    total_paid = loan.repayments.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    total_payable = loan.total_payable or Decimal('0')

    # 4. Disbursement & maturity dates
    disbursement_date = loan.disbursed_date or loan.start_date
    end_date = disbursement_date + relativedelta(months=loan.period_months) if disbursement_date else None

    # 5. Build schedule_data (combines calculated + manual penalties)
    schedule_data = []
    for inst in loan.installments.all().order_by('due_date'):
        calc_penalty = calculate_penalty(inst) or Decimal('0.00')
        manual_total = inst.manual_penalties.filter(is_waived=False).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        total_penalty = calc_penalty + manual_total

        principal_bal = inst.principal_balance
        interest_bal = inst.interest_balance
        total_balance = principal_bal + interest_bal + total_penalty

        schedule_data.append({
            'id': inst.id,
            'due_date': inst.due_date,
            'principal_portion': inst.principal_portion,
            'interest_portion': inst.interest_portion,
            'penalty_amount': total_penalty,
            'balance': total_balance,
            'paid': inst.paid,
            'is_overdue': inst.is_overdue,
        })

    # 6. Manual penalties (active, not waived)
    manual_penalties = loan.manual_penalties.filter(is_waived=False).order_by('-applied_date')
    total_manual_penalty = manual_penalties.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 7. Recent repayment transactions (for the loan or member)
    repayments = Transaction.objects.filter(
        member=loan.member,
        type='repayment',
        loan=loan
    ).order_by('-timestamp')[:10]
    sms_config = SMSConfig.objects.first()
    sms_balance = sms_config.balance if sms_config else 0
    
    # 8. Context
    context = {
        'loan': loan,
        'sms_balance': sms_balance,
        'principal_balance': loan.principal_balance,
        'interest_due': interest_due,
        'principal_due': principal_due,
        'total_due_now': total_due_now,
        'schedule_data': schedule_data,
        'repayments': repayments,
        'total_paid': total_paid.quantize(Decimal('0.01')),
        'total_payable': total_payable,
        'disbursement_date': disbursement_date,
        'end_date': end_date,
        'today': today,
        'manual_penalties': manual_penalties,
        'total_manual_penalty': total_manual_penalty,
    }

    return render(request, 'finance/loan_details.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum

# Updated imports - using GeneralLedger instead of LedgerEntry
from .models import (
    GeneralLedger, 
    ChartOfAccount,
    # AccountCategory is removed because we are using account_type
)


@login_required
def accounting_dashboard(request):
    """Simple dashboard"""
    inflows = GeneralLedger.objects.filter(
        account__code='1001', 
        debit__gt=0
    ).aggregate(total=Sum('debit'))['total'] or 0

    outflows = GeneralLedger.objects.filter(
        account__code='1001', 
        credit__gt=0
    ).aggregate(total=Sum('credit'))['total'] or 0

    context = {
        'total_inflow': inflows,
        'total_outflow': outflows,
        'net_cash': inflows - outflows,
        'recent_entries': GeneralLedger.objects.select_related('account').order_by('-date')[:10]
    }
    return render(request, 'accounting/ledger.html', context)


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from .models import GeneralLedger, ChartOfAccount

from django.db.models import Sum
from datetime import datetime

from django.db.models import Sum

from django.db.models import Sum, Window, F
from django.shortcuts import render
from .models import GeneralLedger, ChartOfAccount

from django.db.models import Sum, Window, F
from django.shortcuts import render
from .models import GeneralLedger, ChartOfAccount

from django.shortcuts import render
from django.db.models import Sum, Window, F
from django.utils import timezone
from .models import GeneralLedger, ChartOfAccount

def general_ledger(request):
    # Base QuerySet
    queryset = GeneralLedger.objects.select_related('account').all().order_by('date', 'id')

    # Get data from POST, default to empty string if not present
    start_date = request.POST.get('start_date') or None
    end_date = request.POST.get('end_date') or None
    account_id = request.POST.get('account_id') or None

    # Apply filters
    if start_date:
        queryset = queryset.filter(date__gte=start_date)
    if end_date:
        queryset = queryset.filter(date__lte=end_date)
    if account_id:
        queryset = queryset.filter(account_id=account_id)

    # Annotate transactions
    transactions = queryset.annotate(
        running_balance=Window(
            expression=Sum(F('credit') - F('debit')),
            order_by=F('date').asc(),
            partition_by=F('account_id')
        )
    )

    totals = transactions.aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )

    return render(request, 'accounting/ledger.html', {
        'transactions': transactions,
        'total_debit': totals['total_debit'] or 0,
        'total_credit': totals['total_credit'] or 0,
        'start_date': start_date or '',
        'end_date': end_date or '',
        'selected_account': account_id or '',
        'all_accounts': ChartOfAccount.objects.all(),
    })
@login_required
def chart_of_accounts(request):
    """List Chart of Accounts grouped by type"""
    grouped_accounts = {}
    for account_type, label in ChartOfAccount.ACCOUNT_TYPES:
        grouped_accounts[label] = ChartOfAccount.objects.filter(
            account_type=account_type, 
            is_active=True
        )

    context = {
        'grouped_accounts': grouped_accounts,
        'title': 'Chart of Accounts'
    }
    return render(request, 'accounting/coa_list.html', context)

from decimal import Decimal

from decimal import Decimal
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import ChartOfAccount, GeneralLedger

from decimal import Decimal
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import ChartOfAccount, GeneralLedger

@login_required
def record_expense(request):
    if request.method == "POST":
        account_id = request.POST.get('account')
        amount_raw = request.POST.get('amount')
        desc = request.POST.get('description')

        try:
            amount = Decimal(amount_raw)
            expense_account = ChartOfAccount.objects.get(id=account_id)
            cash_account = ChartOfAccount.objects.get(code='1001') 

            with transaction.atomic():
                # Entry 1: Debit the Expense
                GeneralLedger.objects.create(
                    account=expense_account,
                    debit=amount,
                    description=desc,
                )
                # Entry 2: Credit the Cash
                GeneralLedger.objects.create(
                    account=cash_account,
                    credit=amount,
                    description=f"Payment for: {desc}",
                )

            messages.success(request, f"Expense of UGX {amount:,.0f} recorded successfully.")
            return redirect('general_ledger')

        except ChartOfAccount.DoesNotExist:
            messages.error(request, "Required account missing (Check if Cash Account 1001 exists).")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    # Fixed: Filter matches the lowercase 'expense' in your ACCOUNT_TYPES
    expense_accounts = ChartOfAccount.objects.filter(account_type='expense')
    
    return render(request, 'accounting/expense_form.html', {
        'expense_accounts': expense_accounts
    })

from decimal import Decimal
from django.db import transaction

@login_required
def record_inflow(request):
    """Record Income with Double Entry (Debit Cash, Credit Income)"""
    if request.method == "POST":
        account_id = request.POST.get('account')
        amount_raw = request.POST.get('amount')
        desc = request.POST.get('description')

        try:
            amount = Decimal(amount_raw)
            income_account = ChartOfAccount.objects.get(id=account_id)
            # Ensure this code matches your 'Cash' account in the DB
            cash_account = ChartOfAccount.objects.get(code='1001') 

            with transaction.atomic():
                # CREDIT the Income Account (Increases Income)
                GeneralLedger.objects.create(
                    account=income_account,
                    credit=amount,
                    debit=0,
                    description=desc,
                )
                # DEBIT the Cash Account (Increases Asset)
                GeneralLedger.objects.create(
                    account=cash_account,
                    debit=amount,
                    credit=0,
                    description=f"Received: {desc}",
                )

            messages.success(request, f"Inflow of UGX {amount:,.0f} recorded.")
            return redirect('general_ledger')
            
        except ChartOfAccount.DoesNotExist:
            messages.error(request, "Account error: Ensure Income and Cash accounts exist.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    context = {
        # FIXED: Changed 'category' to 'account_type' and 'INCOME' to 'income'
        'income_accounts': ChartOfAccount.objects.filter(account_type='income'),
        'title': 'Record Inflow'
    }
    return render(request, 'accounting/inflow_form.html', context)
@login_required
def create_chart_of_account(request):
    if request.method == "POST":
        # Capture all fields from the POST request
        code = request.POST.get('code')
        name = request.POST.get('name')
        account_type = request.POST.get('account_type') # This was the missing piece
        parent_id = request.POST.get('parent')
        description = request.POST.get('description')

        # Basic Validation
        if not account_type:
            messages.error(request, "Please select an account type.")
        elif ChartOfAccount.objects.filter(code=code).exists():
            messages.error(request, f"Account code {code} already exists!")
        else:
            try:
                parent = ChartOfAccount.objects.get(id=parent_id) if parent_id else None
                
                ChartOfAccount.objects.create(
                    code=code,
                    name=name,
                    account_type=account_type, # Passing the string 'asset', 'income', etc.
                    parent=parent,
                    description=description
                )
                messages.success(request, f"Account '{name}' (Code: {code}) has been added.")
                return redirect('general_ledger') # Or your COA list view
            except Exception as e:
                messages.error(request, f"Error creating account: {str(e)}")

    context = {
        'account_types': ChartOfAccount.ACCOUNT_TYPES,
        'parent_accounts': ChartOfAccount.objects.filter(parent=None),
        'title': 'Add New Ledger Account'
    }
    return render(request, 'accounting/coa_form.html', context)
@login_required
def edit_chart_of_account(request, pk):
    account = get_object_or_404(ChartOfAccount, pk=pk)

    if request.method == "POST":
        code = request.POST.get('code')
        name = request.POST.get('name')
        selected_type = request.POST.get('category') # Match the HTML 'name'
        description = request.POST.get('description')

        if ChartOfAccount.objects.filter(code=code).exclude(pk=pk).exists():
            messages.error(request, f"Account code {code} is already taken!")
        else:
            account.code = code
            account.name = name
            account.account_type = selected_type
            account.description = description
            account.save()
            
            messages.success(request, f"Account '{name}' updated successfully.")
            return redirect('chart_of_accounts')

    context = {
        'account': account,
        'account_types': ChartOfAccount.ACCOUNT_TYPES,
        'title': f'Edit {account.name}'
    }
    return render(request, 'accounting/coa_edit_form.html', context)


@login_required
def accounts_hub(request):
    # Get total counts and high-level balances
    total_accounts = ChartOfAccount.objects.count()
    total_inflow = GeneralLedger.objects.filter(account__account_type='income').aggregate(Sum('credit'))['credit__sum'] or 0
    total_outflow = GeneralLedger.objects.filter(account__account_type='expense').aggregate(Sum('debit'))['debit__sum'] or 0
    
    # Recent activity for the mini-table
    recent_transactions = GeneralLedger.objects.select_related('account').order_by('-date')[:5]

    context = {
        'total_accounts': total_accounts,
        'total_inflow': total_inflow,
        'total_outflow': total_outflow,
        'recent_transactions': recent_transactions,
        'net_profit': total_inflow - total_outflow,
    }
    return render(request, 'accounting/accounts_hub.html', context)



from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django_celery_beat.models import PeriodicTask
from finance.models import AutoRepaymentSetting, AutoRepaymentLog, DailyRepaymentSummary
from finance.forms import AutoRepaymentSettingForm
from finance.services import LoanRepaymentEngineService

@login_required
def auto_repayment_dashboard(request):
    config, _ = AutoRepaymentSetting.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        if 'save_settings' in request.POST:
            form = AutoRepaymentSettingForm(request.POST, instance=config)
            if form.is_valid():
                form.instance.updated_by = request.user
                form.save()
                messages.success(request, "Scheduler settings synchronized successfully.")
                return redirect('auto_repayment_dashboard')
        
        elif 'manual_execution_trigger' in request.POST:
            # Trigger manual overrides directly inline safely
            res = LoanRepaymentEngineService.execute_bulk_auto_repayments()
            messages.success(request, f"Manual repayment routine complete. Summary output parsed.")
            return redirect('auto_repayment_dashboard')

    else:
        form = AutoRepaymentSettingForm(instance=config)

    # Fetch status tracking data for dashboards
    celery_task = PeriodicTask.objects.filter(task='finance.tasks.run_automated_loan_repayments').first()
    logs = AutoRepaymentLog.objects.all()[:15]
    summaries = DailyRepaymentSummary.objects.all()[:7]

    context = {
        'form': form,
        'config': config,
        'celery_task': celery_task,
        'logs': logs,
        'summaries': summaries
    }
    return render(request, 'finance/auto_repayment_dashboard.html', context)






##########################################################################################################################


import json
import datetime
from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from .services import FinancialReportingService
from .filters import FinancialReportFilterForm
from .exports import ReportingExportEngine

import json
import datetime
from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from .services import FinancialReportingService
from .filters import FinancialReportFilterForm
from .exports import ReportingExportEngine

class ExecutiveCEODashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'reports/dashboard.html'
    permission_required = 'reports.view_executive_dashboard'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ratios = FinancialReportingService.get_regulatory_ratios()
        aging = FinancialReportingService.get_loan_aging_summary()
        
        # Build core system metrics
        context['kpis'] = ratios
        context['aging_summary'] = aging
        context['interest_data'] = FinancialReportingService.get_interest_income_data()
        
        # Format metrics into JSON structures for serialization into Chart.js/ApexCharts interfaces
        context['chart_aging_labels'] = json.dumps([item['bucket'] for item in aging])
        context['chart_aging_volumes'] = json.dumps([float(item['volume']) for item in aging])
        return context


class InterestIncomeReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    template_name = 'reports/interest_report.html'
    permission_required = 'reports.view_financial_reports'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = FinancialReportFilterForm(self.request.GET or None)
        filters = {}
        if form.is_valid():
            filters = {k: v for k, v in form.cleaned_data.items() if v}
            
        report_data = FinancialReportingService.get_interest_income_data(filters)
        context['records'] = report_data['records']
        context['totals'] = report_data['totals']
        context['filter_form'] = form
        return context


class TreasuryDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/treasury_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        forecast = FinancialReportingService.get_treasury_liquidity_forecast()
        context['forecast_raw'] = forecast
        context['forecast_json'] = json.dumps(forecast)
        return context


import datetime
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import View
from django.http import HttpResponse

# Ensure proper relative/explicit imports matching your local app structure
from .services import FinancialReportingService


import datetime
from django.views.generic import View
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin

from .services import FinancialReportingService
from .forms import InterestIncomeFilterForm
from .exports import ReportingExportEngine  # Adjust path to your ReportingExportEngine location

import datetime
from django.views.generic import View
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin



class InterestIncomeReportView(LoginRequiredMixin, View):
    """
    Unified Ledger View Engine. Handles real-time HTML data filtering, 
    and intercepts format parameters to serve binary Excel/PDF document extensions.
    """
    template_name = 'reports/interest_report.html'

    def get(self, request, *args, **kwargs):
        # 1. Initialize filter framework with query parameters
        form = InterestIncomeFilterForm(request.GET or None)
        filters = {}
        if form.is_valid():
            filters = {k: v for k, v in form.cleaned_data.items() if v}
        
        # 2. Extract calculations dataset matrix from our reporting engine
        reporting_payload = FinancialReportingService.get_interest_income_data(filters)
        records = reporting_payload['records']
        totals = reporting_payload['totals']

        # 3. Intercept binary export requests before rendering HTML templates
        export_type = request.GET.get('format', '').lower()
        if export_type in ['excel', 'pdf']:
            
            # Unified data verification columns array layout block
            columns = [
                'Date', 'Member No', 'Customer Name', 'Loan Ref', 
                'Product', 'Principal Remaining', 'Interest Remaining', 'Total Outstanding Balance'
            ]
            
            dataset = []
            for item in records:
                # Fallback safe date capture framework execution block
                d_date = getattr(item, 'disbursed_date', None) or getattr(item, 'start_date', None)
                if isinstance(d_date, (datetime.date, datetime.datetime)):
                    date_str = d_date.strftime('%Y-%m-%d')
                else:
                    date_str = str(d_date) if d_date else "N/A"
                
                # Protect computations against Null positions using float-safe fallback operations
                principal_rem = float(item.principal_receivable) if item.principal_receivable else 0.0
                interest_rem = float(item.interest_receivable) if item.interest_receivable else 0.0
                total_outstanding = float(item.total_remaining_balance) if item.total_remaining_balance else 0.0
                
                dataset.append([
                    date_str,
                    item.member.member_number,
                    f"{item.member.first_name} {item.member.last_name}",
                    item.loan_reference or f"LN-{item.id}",
                    str(item.product_type).upper() if item.product_type else "STANDARD",
                    principal_rem,
                    interest_rem,
                    total_outstanding
                ])
            
            # Forward processed data payloads using the exact engine signatures expected
            if export_type == 'excel':
                # Takes exactly 3 positional arguments
                return ReportingExportEngine.generate_excel('interest', columns, dataset)
            elif export_type == 'pdf':
                # Takes exactly 4 positional arguments
                return ReportingExportEngine.generate_pdf('interest', columns, dataset, request.user)

        # 4. Fallback to serving standard HTML template framework if no valid format parameter intercepted
        return render(request, self.template_name, {
            'filter_form': form,
            'records': records,
            'totals': totals
        })



############################################################################
import datetime
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, F, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse

# Adjust these imports to match your structural app directory layout
from finance.models import Loan, Installment  

import datetime
from django.db.models import Q, Sum, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from finance.models import Installment  # Adjust this import to your actual app structure

import datetime
from decimal import Decimal  # <-- Added for strict type safety
from django.db.models import Q, Sum, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from finance.models import Installment 

from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, F, Value, DecimalField
from django.db.models.functions import Coalesce
from datetime import datetime, date, timedelta
from decimal import Decimal

from .models import Installment


from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, F, Value, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce
from datetime import datetime, date, timedelta
from decimal import Decimal

from .models import Installment


from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, F, Value, DecimalField, IntegerField, ExpressionWrapper
from django.db.models.functions import Coalesce
from datetime import datetime, date, timedelta
from decimal import Decimal

from .models import Installment


from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, F, Value, DecimalField, IntegerField, ExpressionWrapper
from django.db.models.functions import Coalesce
from datetime import datetime, date, timedelta
from decimal import Decimal

from .models import Installment


import datetime
from datetime import date, datetime, timedelta

from decimal import Decimal

from django.db.models import (
    Q, F, Value, Sum, DecimalField, IntegerField,
    ExpressionWrapper
)
from django.db.models.functions import Coalesce, Now
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView

from finance.models import Installment


class LoansInArrearsReportView(LoginRequiredMixin, ListView):
    """
    Fixed Arrears Report (Type-safe + Production-ready)
    """
    model = Installment
    template_name = 'finance/arrears.html'
    context_object_name = 'overdue'
    paginate_by = 50

    def get_queryset(self):
        search_query = self.request.GET.get('search_query', '').strip()
        date_at_str = self.request.GET.get('date_at', '').strip()
        sort_by = self.request.GET.get('sort_by', '')

        try:
            target_date = datetime.strptime(date_at_str, '%Y-%m-%d').date() if date_at_str else date.today()
        except ValueError:
            target_date = date.today()

        queryset = Installment.objects.filter(
            paid=False,
            due_date__lt=target_date,
            loan__status__in=['approved', 'arrears']
        ).select_related('loan', 'loan__member', 'loan__officer')

        if search_query:
            queryset = queryset.filter(
                Q(loan__member__first_name__icontains=search_query) |
                Q(loan__member__last_name__icontains=search_query) |
                Q(loan__member__member_number__icontains=search_query) |
                Q(loan__loan_reference__icontains=search_query)
            )

        # =========================================================
        # FIXED ANNOTATIONS (NO TYPE CONFLICTS)
        # =========================================================
        queryset = queryset.annotate(
            arrears_amount=ExpressionWrapper(
                Coalesce(F('principal_portion') - F('principal_paid'), Value(0)) +
                Coalesce(F('interest_portion') - F('interest_paid'), Value(0)) +
                Coalesce(F('penalty_amount') - F('penalty_paid'), Value(0)),
                output_field=DecimalField(max_digits=12, decimal_places=2)
            ),

            # SAFE DAYS OVERDUE (FIXED)
            days_overdue=ExpressionWrapper(
                Now() - F('due_date'),
                output_field=DecimalField()  # duration-safe fallback
            )
        )

        # Sorting
        if sort_by == 'days':
            queryset = queryset.order_by('-due_date')
        elif sort_by == 'amount':
            queryset = queryset.order_by('-arrears_amount')
        else:
            queryset = queryset.order_by('-due_date')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_queryset = self.get_queryset()

        date_at_str = self.request.GET.get('date_at', '').strip()

        try:
            target_date = datetime.strptime(date_at_str, '%Y-%m-%d').date() if date_at_str else date.today()
        except ValueError:
            target_date = date.today()

        watchlist_barrier = target_date - timedelta(days=30)
        substandard_barrier = target_date - timedelta(days=60)

        totals = base_queryset.aggregate(
            total=Coalesce(
                Sum('arrears_amount'),
                Value(0),
                output_field=DecimalField()
            ),

            watchlist=Coalesce(
                Sum('arrears_amount', filter=Q(due_date__gte=watchlist_barrier)),
                Value(0),
                output_field=DecimalField()
            ),

            substandard=Coalesce(
                Sum('arrears_amount', filter=Q(due_date__lt=watchlist_barrier, due_date__gte=substandard_barrier)),
                Value(0),
                output_field=DecimalField()
            ),

            doubtful=Coalesce(
                Sum('arrears_amount', filter=Q(due_date__lt=substandard_barrier)),
                Value(0),
                output_field=DecimalField()
            )
        )

        active_portfolio_total = Installment.objects.filter(
            loan__status__in=['approved', 'arrears']
        ).aggregate(
            gross=Coalesce(
                Sum(F('principal_portion') + F('interest_portion')),
                Value(0),
                output_field=DecimalField()
            )
        )['gross'] or Decimal('1.0')

        total_at_risk = totals['total']
        par_rate = (total_at_risk / active_portfolio_total * Decimal('100')) if active_portfolio_total > 0 else Decimal('0')

        context.update({
            'total_at_risk': total_at_risk,
            'watchlist_total': totals['watchlist'],
            'substandard_total': totals['substandard'],
            'doubtful_total': totals['doubtful'],
            'par_rate': round(float(par_rate), 2),
            'today': target_date,
        })

        return context


# finance/views.py
import random
from decimal import Decimal
from datetime import datetime, date, timedelta
from django.db.models import Sum, Q, F, Count, Case, When, Value, DecimalField
from django.db.models.functions import Coalesce, TruncMonth
from django.shortcuts import render
from django.http import FileResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from .models import (
    Loan, Installment, Member, SystemSetting, GeneralLedger, ChartOfAccount,
    SavingsAccount, Transaction, Company, SMSConfig
)
from .utils import generate_excel_report

User = get_user_model()


# ====================================================================
# CONTEXT BUILDERS & HELPERS
# ====================================================================

def get_report_context(request):
    """
    Builds the report context based on POST/GET filters.
    Returns a dict with columns, data, totals, KPIs, etc.
    """
    context = {}

    # 1. Extract filters
    date_from = request.POST.get('date_from') or request.GET.get('date_from')
    date_to = request.POST.get('date_to') or request.GET.get('date_to')
    officer_id = request.POST.get('officer') or request.GET.get('officer')
    status = request.POST.get('status') or request.GET.get('status')

    # 2. Base queryset – use your actual model (Loan)
    qs = Loan.objects.select_related('member', 'officer')

    if date_from:
        qs = qs.filter(disbursed_date__gte=date_from)
    if date_to:
        qs = qs.filter(disbursed_date__lte=date_to)
    if officer_id:
        qs = qs.filter(officer_id=officer_id)
    if status:
        qs = qs.filter(status=status)

    # 3. Define columns (must match the keys in data rows)
    columns = [
        {'key': 'loan_reference', 'label': 'Loan Reference'},
        {'key': 'member_name', 'label': 'Member'},
        {'key': 'principal', 'label': 'Principal', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'interest_balance', 'label': 'Interest Balance', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'total_balance', 'label': 'Total Balance', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'status', 'label': 'Status', 'type': 'status'},
        {'key': 'disbursed_date', 'label': 'Disbursed', 'type': 'date'},
        {'key': 'officer', 'label': 'Officer'},
    ]

    # 4. Build data rows
    data = []
    total_principal = Decimal('0.00')
    total_interest = Decimal('0.00')
    total_balance = Decimal('0.00')

    for loan in qs:
        principal = loan.principal_amount or Decimal('0.00')
        interest = loan.interest_balance or Decimal('0.00')
        balance = principal + interest

        data.append({
            'loan_reference': loan.loan_reference or f"LN-{loan.id}",
            'member_name': loan.member.get_full_name() if loan.member else 'N/A',
            'principal': principal,
            'interest_balance': interest,
            'total_balance': balance,
            'status': loan.get_status_display(),
            'disbursed_date': loan.disbursed_date,
            'officer': loan.officer.get_full_name() if loan.officer else 'N/A',
        })

        total_principal += principal
        total_interest += interest
        total_balance += balance

    # 5. Totals dictionary
    totals = {
        'principal': total_principal,
        'interest_balance': total_interest,
        'total_balance': total_balance,
    }

    # 6. KPIs
    record_count = len(data)
    kpi_cards = [
        {'label': 'Total Loans', 'value': record_count, 'icon': 'bi-file-earmark-text', 'type': 'info'},
        {'label': 'Total Principal', 'value': f"UGX {total_principal:,.0f}", 'icon': 'bi-cash', 'type': 'success'},
        {'label': 'Total Interest', 'value': f"UGX {total_interest:,.0f}", 'icon': 'bi-percent', 'type': 'warning'},
        {'label': 'Total Balance', 'value': f"UGX {total_balance:,.0f}", 'icon': 'bi-wallet2', 'type': 'danger'},
    ]

    # 7. Aging Summary (optional)
    aging_summary = []

    # 8. Summary totals (extra stats)
    summary_totals = {
        'total_records': record_count,
        'total_amount': total_balance,
        'total_paid': Decimal('0.00'),
        'outstanding': total_balance,
        'recovery_rate': 0,
        'par_30': 0,
    }

    # 9. Officer list for filter dropdown
    officer_list = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    # 10. Company info – FROM DATABASE
    company = Company.get_company()

    # 11. Build the final context
    context.update({
        'columns': columns,
        'data': data,
        'totals': totals,
        'has_data': bool(data),
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'aging_summary': aging_summary,
        'report_title': 'Loan Portfolio Report',
        'company': company,
        'date_from': date_from,
        'date_to': date_to,
        'selected_officer': officer_id,
        'selected_status': status,
        'officer_list': officer_list,
        'officer_name': dict(officer_list.values_list('id', 'username')).get(int(officer_id) if officer_id else None),
        'generated_date': timezone.now().strftime('%d %b %Y %H:%M'),
        'generated_by': request.user.get_full_name() if request.user.is_authenticated else 'System',
    })

    return context


def generate_excel_report(columns, data, report_title="Report", company_name="Company", totals=None):
    """
    Generate an Excel workbook from report columns and data.
    Optionally add a totals row.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = report_title[:31]

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1632af", end_color="1632af", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Optional title row
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    ws.cell(row=row, column=1).value = f"{company_name} - {report_title}"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    # Headers
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col.get('label', col.get('key', '')))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    row += 1

    # Data rows
    for data_row in data:
        for col_idx, col in enumerate(columns, start=1):
            key = col.get('key')
            value = data_row.get(key, '-')

            # Format based on type
            if col.get('type') == 'currency':
                try:
                    value = f"{float(value):,.0f}"
                except (ValueError, TypeError):
                    pass
            elif col.get('type') == 'date' and value:
                if hasattr(value, 'strftime'):
                    value = value.strftime('%d %b, %Y')
                else:
                    value = str(value)
            elif col.get('type') == 'status':
                value = str(value)

            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            if col.get('align') == 'right' or col.get('type') == 'currency':
                cell.alignment = number_alignment
            else:
                cell.alignment = cell_alignment
        row += 1

    # -------------------------
    # TOTALS ROW (if totals provided)
    # -------------------------
    if totals:
        ws.cell(row=row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        for col_idx, col in enumerate(columns, start=1):
            if col.get('total') and col.get('key') in totals:
                total_val = totals[col['key']]
                if col.get('type') == 'currency':
                    try:
                        total_val = f"{float(total_val):,.0f}"
                    except (ValueError, TypeError):
                        pass
                else:
                    total_val = str(total_val)

                cell = ws.cell(row=row, column=col_idx, value=total_val)
                cell.font = Font(bold=True)
                cell.border = border
                if col.get('align') == 'right' or col.get('type') == 'currency':
                    cell.alignment = number_alignment
                else:
                    cell.alignment = cell_alignment

        row += 1

    # Auto-size columns
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ====================================================================
# EXPORT VIEW
# ====================================================================

def export_report_excel(request):
    context = get_report_context(request)

    columns = context.get('columns', [])
    data = context.get('data', [])
    totals = context.get('totals', {})
    report_title = context.get('report_title', 'Report')
    company_name = context.get('company', {}).get('name', 'Company')

    excel_file = generate_excel_report(
        columns=columns,
        data=data,
        report_title=report_title,
        company_name=company_name,
        totals=totals
    )

    filename = f"{report_title.replace(' ', '_')}_{context.get('generated_date', 'now').replace(' ', '_').replace(':', '')}.xlsx"
    response = FileResponse(
        excel_file,
        as_attachment=True,
        filename=filename,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    return response


# ====================================================================
# REPORT VIEWS – ALL USING Company.get_company()
# ====================================================================

@login_required
def loan_report(request):
    """Professional Loan Portfolio Report with Filters"""
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    officer_id = request.GET.get('officer')
    status = request.GET.get('status')

    loans = Loan.objects.select_related('member', 'officer').prefetch_related('repayments')

    if date_from:
        loans = loans.filter(disbursed_date__gte=date_from)
    if date_to:
        loans = loans.filter(disbursed_date__lte=date_to)
    if officer_id:
        loans = loans.filter(officer_id=officer_id)
    if status:
        loans = loans.filter(status=status)

    loans = loans.annotate(
        paid_amount=Sum('repayments__amount_paid', default=0),
        total_balance=F('principal_balance') + F('interest_balance'),
    )

    data = []
    total_amount = Decimal('0')
    total_paid = Decimal('0')
    total_balance = Decimal('0')

    for loan in loans:
        balance = loan.total_balance or Decimal('0')
        paid = loan.paid_amount or Decimal('0')

        data.append({
            'member': f"{loan.member.first_name} {loan.member.last_name}".strip(),
            'member_no': loan.member.member_number or str(loan.member.id),
            'loan_ref': loan.loan_reference or f"LN-{loan.id}",
            'amount': loan.principal_amount,
            'paid': paid,
            'balance': balance,
            'status': loan.status,
            'date': loan.disbursed_date or loan.start_date,
            'officer': loan.officer.get_full_name() if loan.officer else 'System',
        })

        total_amount += loan.principal_amount or Decimal('0')
        total_paid += paid
        total_balance += balance

    recovery_rate = round((total_paid / total_amount * 100), 1) if total_amount > 0 else 0
    par_30 = round((total_balance / total_amount * 100), 1) if total_amount > 0 else 0
    total_records = len(data)

    columns = [
        {'key': 'member', 'label': 'Member Name', 'align': 'left'},
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'loan_ref', 'label': 'Loan Reference', 'align': 'left'},
        {'key': 'amount', 'label': 'Amount (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'paid', 'label': 'Paid (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'balance', 'label': 'Balance (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
        {'key': 'date', 'label': 'Date', 'align': 'center', 'type': 'date'},
        {'key': 'officer', 'label': 'Officer', 'align': 'left'},
    ]

    kpi_cards = [
        {'icon': 'bi-people', 'value': f'{total_records:,}', 'label': 'Total Loans'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_amount:,.0f}', 'label': 'Total Portfolio', 'type': 'success'},
        {'icon': 'bi-check-circle', 'value': f'{recovery_rate}%', 'label': 'Recovery Rate', 'type': 'info'},
        {'icon': 'bi-exclamation-triangle', 'value': f'{par_30}%', 'label': 'PAR 30', 'type': 'warning'},
    ]

    summary_totals = {
        'total_records': total_records,
        'total_amount': total_amount,
        'total_paid': total_paid,
        'outstanding': total_balance,
        'recovery_rate': recovery_rate,
        'par_30': par_30,
    }

    officer_list = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    context = {
        'company': Company.get_company(),
        'report_title': 'Loan Portfolio Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': date_from or 'All',
        'date_to': date_to or 'All',
        'officer': officer_id or 'All Officers',
        'status': status or 'All Status',
        'columns': columns,
        'data': data,
        'totals': {
            'amount': total_amount,
            'paid': total_paid,
            'balance': total_balance,
        },
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'officer_list': officer_list,
    }

    return render(request, 'finance/reports/base_report.html', context)


@login_required
def member_report(request):
    """Member Report"""
    members = Member.objects.all().order_by('member_number')

    data = []
    total_savings = Decimal('0')
    total_loans = Decimal('0')

    for member in members:
        savings = SavingsAccount.objects.filter(member=member).first()
        savings_balance = savings.balance if savings else Decimal('0')
        total_loans_balance = member.loans.filter(is_active=True).aggregate(
            total=Sum('principal_balance') + Sum('interest_balance')
        )['total'] or Decimal('0')

        status = getattr(member, 'status', 'active')
        status_display = 'Active' if status == 'active' else 'Inactive'

        data.append({
            'member_no': member.member_number,
            'name': f"{member.first_name} {member.last_name}",
            'phone': member.phone_number,
            'email': member.email or '—',
            'savings': savings_balance,
            'loans': total_loans_balance,
            'joined': member.date_joined,
            'status': status_display,
        })

        total_savings += savings_balance
        total_loans += total_loans_balance

    columns = [
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'name', 'label': 'Member Name', 'align': 'left'},
        {'key': 'phone', 'label': 'Phone', 'align': 'left'},
        {'key': 'email', 'label': 'Email', 'align': 'left'},
        {'key': 'savings', 'label': 'Savings (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'loans', 'label': 'Loans (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'joined', 'label': 'Joined', 'align': 'center', 'type': 'date'},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
    ]

    totals = {
        'savings': total_savings,
        'loans': total_loans,
    }

    kpi_cards = [
        {'icon': 'bi-people', 'value': f'{len(data):,}', 'label': 'Total Members'},
        {'icon': 'bi-wallet2', 'value': f'UGX {total_savings:,.0f}', 'label': 'Total Savings', 'type': 'success'},
        {'icon': 'bi-bank', 'value': f'UGX {total_loans:,.0f}', 'label': 'Total Loans', 'type': 'info'},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Member Registry Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': 'All',
        'date_to': 'All',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(data) > 0,
        'summary_totals': {
            'total_records': len(data),
            'total_amount': total_savings + total_loans,
            'total_paid': total_savings,
            'outstanding': total_loans,
            'recovery_rate': '100',
            'par_30': '0',
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


@login_required
def savings_report(request):
    """Savings Report"""
    savings_accounts = SavingsAccount.objects.select_related('member').all()

    data = []
    total_balance = Decimal('0')

    for savings in savings_accounts:
        data.append({
            'member': f"{savings.member.first_name} {savings.member.last_name}",
            'member_no': savings.member.member_number,
            'phone': savings.member.phone_number,
            'balance': savings.balance,
            'account_no': savings.account_number if hasattr(savings, 'account_number') else 'N/A',
            'status': 'Active' if savings.balance > 0 else 'Inactive',
        })
        total_balance += savings.balance

    columns = [
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'member', 'label': 'Member Name', 'align': 'left'},
        {'key': 'phone', 'label': 'Phone', 'align': 'left'},
        {'key': 'account_no', 'label': 'Account No', 'align': 'left'},
        {'key': 'balance', 'label': 'Balance (UGX)', 'align': 'right', 'type': 'currency', 'total': True},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
    ]

    totals = {'balance': total_balance}

    context = {
        'company': Company.get_company(),
        'report_title': 'Savings Summary Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': 'All',
        'date_to': 'All',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': [
            {'icon': 'bi-wallet2', 'value': f'{len(data):,}', 'label': 'Total Accounts'},
            {'icon': 'bi-currency-dollar', 'value': f'UGX {total_balance:,.0f}', 'label': 'Total Savings', 'type': 'success'},
            {'icon': 'bi-people', 'value': f'{len(data):,}', 'label': 'Active Members', 'type': 'info'},
        ],
        'summary_totals': {
            'total_records': len(data),
            'total_amount': total_balance,
            'total_paid': total_balance,
            'outstanding': 0,
            'recovery_rate': '100',
            'par_30': '0',
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


@login_required
def financial_report(request):
    """Financial Performance Report (Income/Expense summary)"""
    from datetime import date, timedelta

    if request.method == "POST":
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
    else:
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')

    if not date_from:
        date_from = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = date.today().strftime('%Y-%m-%d')

    transactions = Transaction.objects.filter(
        timestamp__date__gte=date_from,
        timestamp__date__lte=date_to
    )

    total_income = transactions.filter(
        Q(type='deposit') | Q(type='repayment') | Q(type='interest_payment')
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    total_expenses = transactions.filter(
        Q(type='withdrawal') | Q(type='disbursement') | Q(type='penalty')
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    net_profit = total_income - total_expenses

    monthly_data = transactions.annotate(
        month=TruncMonth('timestamp')
    ).values('month').annotate(
        income=Sum('amount', filter=Q(type__in=['deposit', 'repayment', 'interest_payment'])),
        expense=Sum('amount', filter=Q(type__in=['withdrawal', 'disbursement', 'penalty']))
    ).order_by('month')

    data = []
    for entry in monthly_data:
        data.append({
            'month': entry['month'].strftime('%b %Y') if entry['month'] else 'N/A',
            'income': entry['income'] or Decimal('0'),
            'expense': entry['expense'] or Decimal('0'),
            'profit': (entry['income'] or Decimal('0')) - (entry['expense'] or Decimal('0')),
        })

    columns = [
        {'key': 'month', 'label': 'Month', 'align': 'left'},
        {'key': 'income', 'label': 'Income (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'expense', 'label': 'Expenses (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'profit', 'label': 'Net Profit (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
    ]

    totals = {
        'income': total_income,
        'expense': total_expenses,
        'profit': net_profit,
    }

    kpi_cards = [
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_income:,.0f}', 'label': 'Total Income', 'type': 'success'},
        {'icon': 'bi-cash', 'value': f'UGX {total_expenses:,.0f}', 'label': 'Total Expenses', 'type': 'danger'},
        {'icon': 'bi-graph-up', 'value': f'UGX {net_profit:,.0f}', 'label': 'Net Profit', 'type': 'info'},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Financial Performance Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': date_from,
        'date_to': date_to,
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(data) > 0,
        'summary_totals': {
            'total_records': len(data),
            'total_amount': total_income + total_expenses,
            'total_paid': total_income,
            'outstanding': total_expenses,
            'recovery_rate': 'N/A',
            'par_30': 'N/A',
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


from decimal import Decimal
from datetime import date
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum
from .models import Loan, Transaction, Company

@login_required
def officer_report(request):
    """Officer Performance Report with PAR 1 & PAR 30 - Credit Officers only"""
    
    # ---- Only include active users who are in the "Credit Officer" group ----
    officers = User.objects.filter(
        is_active=True,
        groups__name='Credit Officer'
    ).order_by('first_name', 'last_name')

    data = []
    total_loans = Decimal('0')
    total_disbursed = Decimal('0')
    total_collected = Decimal('0')
    total_outstanding = Decimal('0')
    total_par_1 = Decimal('0')
    total_par_30 = Decimal('0')

    today = date.today()

    for officer in officers:
        loans = Loan.objects.filter(officer=officer)
        loan_count = loans.count()
        disbursed_amount = loans.aggregate(total=Sum('principal_amount'))['total'] or Decimal('0')

        collections = Transaction.objects.filter(
            type='repayment',
            loan__officer=officer
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        active_count = loans.filter(is_active=True).count()

        par_1_amount = Decimal('0')
        par_30_amount = Decimal('0')
        outstanding_principal = Decimal('0')

        for loan in loans:
            p_bal = loan.principal_balance or Decimal('0')
            outstanding_principal += p_bal

            unpaid_installments = loan.installments.filter(paid=False, due_date__lt=today)
            if unpaid_installments.exists():
                oldest_due = unpaid_installments.earliest('due_date').due_date
                days_overdue = (today - oldest_due).days

                if days_overdue >= 1:
                    par_1_amount += p_bal
                if days_overdue >= 30:
                    par_30_amount += p_bal

        # Calculate percentages and round to 2 decimal places
        par_1_percent = round((par_1_amount / outstanding_principal * 100), 2) if outstanding_principal > 0 else 0
        par_30_percent = round((par_30_amount / outstanding_principal * 100), 2) if outstanding_principal > 0 else 0
        performance = round((collections / disbursed_amount * 100), 2) if disbursed_amount > 0 else 0

        data.append({
            'officer': officer.get_full_name() or officer.username,
            'loan_count': loan_count,
            'active_count': active_count,
            'disbursed': disbursed_amount,
            'collected': collections,
            'outstanding': outstanding_principal,
            'par_1_amount': par_1_amount,
            'par_1_percent': par_1_percent,
            'par_30_amount': par_30_amount,
            'par_30_percent': par_30_percent,
            'performance': performance,
        })

        total_loans += loan_count
        total_disbursed += disbursed_amount
        total_collected += collections
        total_outstanding += outstanding_principal
        total_par_1 += par_1_amount
        total_par_30 += par_30_amount

    # Round totals
    total_par_1_percent = round((total_par_1 / total_outstanding * 100), 2) if total_outstanding > 0 else 0
    total_par_30_percent = round((total_par_30 / total_outstanding * 100), 2) if total_outstanding > 0 else 0
    total_performance = round((total_collected / total_disbursed * 100), 2) if total_disbursed > 0 else 0

    columns = [
        {'key': 'officer', 'label': 'Officer', 'align': 'left'},
        {'key': 'loan_count', 'label': 'Total Loans', 'align': 'center'},
        {'key': 'active_count', 'label': 'Active', 'align': 'center'},
        {'key': 'disbursed', 'label': 'Disbursed (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'collected', 'label': 'Collected (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'outstanding', 'label': 'Outstanding (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'par_1_amount', 'label': 'PAR 1 (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'par_1_percent', 'label': 'PAR 1 %', 'align': 'right', 'total': True},
        {'key': 'par_30_amount', 'label': 'PAR 30 (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'par_30_percent', 'label': 'PAR 30 %', 'align': 'right', 'total': True},
        {'key': 'performance', 'label': 'Performance %', 'align': 'right', 'total': True},
    ]

    totals = {
        'disbursed': total_disbursed,
        'collected': total_collected,
        'outstanding': total_outstanding,
        'par_1_amount': total_par_1,
        'par_1_percent': total_par_1_percent,
        'par_30_amount': total_par_30,
        'par_30_percent': total_par_30_percent,
        'performance': total_performance,
    }

    kpi_cards = [
        {'icon': 'bi-person-badge', 'value': f'{len(data)}', 'label': 'Total Officers', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_disbursed:,.0f}', 'label': 'Total Disbursed', 'type': 'success'},
        {'icon': 'bi-cash-stack', 'value': f'UGX {total_collected:,.0f}', 'label': 'Total Collected', 'type': 'info'},
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {total_par_30:,.0f}', 'label': 'Total PAR 30', 'type': 'danger'},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Officer Performance Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': 'All',
        'date_to': 'All',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(data) > 0,
        'summary_totals': {
            'total_records': len(data),
            'total_amount': total_disbursed,
            'total_paid': total_collected,
            'outstanding': total_outstanding,
            'recovery_rate': f'{total_performance:.2f}',
            'par_30': f'{total_par_30_percent:.2f}',
        },
    }

    return render(request, 'finance/reports/base_report.html', context)
@login_required
def accounting_report(request):
    """Accounting Report - General Ledger Summary"""
    ledger_entries = GeneralLedger.objects.select_related('account').all().order_by('account__code')

    data = []
    total_debit = Decimal('0')
    total_credit = Decimal('0')

    for entry in ledger_entries:
        data.append({
            'account_code': entry.account.code,
            'account_name': entry.account.name,
            'account_type': entry.account.get_account_type_display(),
            'debit': entry.debit,
            'credit': entry.credit,
            'balance': entry.balance,
        })
        total_debit += entry.debit
        total_credit += entry.credit

    columns = [
        {'key': 'account_code', 'label': 'Account Code', 'align': 'left'},
        {'key': 'account_name', 'label': 'Account Name', 'align': 'left'},
        {'key': 'account_type', 'label': 'Type', 'align': 'left'},
        {'key': 'debit', 'label': 'Debit (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'credit', 'label': 'Credit (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'balance', 'label': 'Balance (UGX)', 'align': 'right', 'type': 'currency', 'prefix': 'UGX '},
    ]

    totals = {
        'debit': total_debit,
        'credit': total_credit,
        'balance': total_debit - total_credit,
    }

    kpi_cards = [
        {'icon': 'bi-journal-text', 'value': f'{len(data)}', 'label': 'Total Accounts', 'type': 'info'},
        {'icon': 'bi-arrow-down', 'value': f'UGX {total_debit:,.0f}', 'label': 'Total Debit', 'type': 'danger'},
        {'icon': 'bi-arrow-up', 'value': f'UGX {total_credit:,.0f}', 'label': 'Total Credit', 'type': 'success'},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Accounting Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': 'All',
        'date_to': 'All',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(data) > 0,
        'summary_totals': {
            'total_records': len(data),
            'total_amount': total_debit + total_credit,
            'total_paid': total_credit,
            'outstanding': total_debit,
            'recovery_rate': 'N/A',
            'par_30': 'N/A',
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


@login_required
def audit_report(request):
    """Audit Trail Report"""
    transactions = Transaction.objects.all().order_by('-timestamp')[:100]

    data = []
    for tx in transactions:
        data.append({
            'timestamp': tx.timestamp,
            'user': tx.created_by.get_full_name() if tx.created_by else 'System',
            'type': tx.get_type_display(),
            'amount': tx.amount,
            'reference': tx.reference,
            'is_reversed': 'Yes' if tx.is_reversed else 'No',
            'status': 'Reversed' if tx.is_reversed else 'Active',
        })

    columns = [
        {'key': 'timestamp', 'label': 'Date & Time', 'align': 'center', 'type': 'date'},
        {'key': 'user', 'label': 'User', 'align': 'left'},
        {'key': 'type', 'label': 'Transaction Type', 'align': 'left'},
        {'key': 'amount', 'label': 'Amount (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'reference', 'label': 'Reference', 'align': 'left'},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
    ]

    totals = {
        'amount': transactions.aggregate(total=Sum('amount'))['total'] or Decimal('0'),
    }

    kpi_cards = [
        {'icon': 'bi-clock-history', 'value': f'{len(data)}', 'label': 'Total Transactions', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {totals["amount"]:,.0f}', 'label': 'Total Volume', 'type': 'success'},
        {'icon': 'bi-person', 'value': f'{len(set(tx.created_by_id for tx in transactions if tx.created_by))}', 'label': 'Active Users', 'type': 'secondary'},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Audit Trail Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': 'All',
        'date_to': 'All',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(data) > 0,
        'summary_totals': {
            'total_records': len(data),
            'total_amount': totals['amount'],
            'total_paid': 'N/A',
            'outstanding': 'N/A',
            'recovery_rate': 'N/A',
            'par_30': 'N/A',
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


@login_required
def inventory_report(request):
    """Inventory Report - Products and Stock Levels"""
    try:
        from hardware.models import Product, Category
        products = Product.objects.select_related('category').all()
        has_inventory = True
    except ImportError:
        products = []
        has_inventory = False

    data = []
    total_value = Decimal('0')
    total_stock = 0

    if has_inventory:
        for product in products:
            stock_value = (product.current_stock or 0) * (product.cost_price or 0)
            data.append({
                'product_code': product.product_code,
                'product_name': product.name,
                'category': product.category.name if product.category else 'Uncategorized',
                'stock': product.current_stock or 0,
                'cost_price': product.cost_price or 0,
                'selling_price': product.selling_price or 0,
                'stock_value': stock_value,
                'status': 'Low Stock' if (product.current_stock or 0) <= (product.reorder_level or 5) else 'Healthy',
            })
            total_value += stock_value
            total_stock += (product.current_stock or 0)
    else:
        data = [{'message': 'Inventory module not installed'}]

    columns = [
        {'key': 'product_code', 'label': 'Product Code', 'align': 'left'},
        {'key': 'product_name', 'label': 'Product Name', 'align': 'left'},
        {'key': 'category', 'label': 'Category', 'align': 'left'},
        {'key': 'stock', 'label': 'Stock', 'align': 'center'},
        {'key': 'cost_price', 'label': 'Cost (UGX)', 'align': 'right', 'type': 'currency', 'prefix': 'UGX '},
        {'key': 'selling_price', 'label': 'Sell (UGX)', 'align': 'right', 'type': 'currency', 'prefix': 'UGX '},
        {'key': 'stock_value', 'label': 'Value (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
    ]

    totals = {'stock_value': total_value}

    kpi_cards = [
        {'icon': 'bi-boxes', 'value': f'{len(data)}', 'label': 'Total Products', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_value:,.0f}', 'label': 'Inventory Value', 'type': 'success'},
        {'icon': 'bi-box', 'value': f'{total_stock:,}', 'label': 'Total Stock Units', 'type': 'secondary'},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Inventory Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': 'All',
        'date_to': 'All',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(data) > 0,
        'summary_totals': {
            'total_records': len(data),
            'total_amount': total_value,
            'total_paid': 'N/A',
            'outstanding': 'N/A',
            'recovery_rate': 'N/A',
            'par_30': 'N/A',
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


@login_required
def interest_report(request):
    """Interest Income Report"""
    from datetime import date, timedelta

    if request.method == "POST":
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
    else:
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')

    if not date_from:
        date_from = (date.today() - timedelta(days=365)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = date.today().strftime('%Y-%m-%d')

    loans = Loan.objects.filter(
        disbursed_date__gte=date_from,
        disbursed_date__lte=date_to,
        status__in=['approved', 'active', 'closed']
    ).select_related('member', 'officer')

    data = []
    total_principal = Decimal('0')
    total_interest_charged = Decimal('0')
    total_interest_paid = Decimal('0')
    total_interest_balance = Decimal('0')

    for loan in loans:
        interest_charged = loan.installments.aggregate(total=Sum('interest_portion'))['total'] or Decimal('0')
        interest_paid = loan.installments.aggregate(total=Sum('interest_paid'))['total'] or Decimal('0')
        interest_balance = loan.interest_balance or Decimal('0')

        data.append({
            'member': f"{loan.member.first_name} {loan.member.last_name}",
            'loan_ref': loan.loan_reference or f"LN-{loan.id}",
            'principal': loan.principal_amount,
            'interest_charged': interest_charged,
            'interest_paid': interest_paid,
            'interest_balance': interest_balance,
            'status': loan.status,
            'disbursed_date': loan.disbursed_date or loan.start_date,
        })

        total_principal += loan.principal_amount
        total_interest_charged += interest_charged
        total_interest_paid += interest_paid
        total_interest_balance += interest_balance

    columns = [
        {'key': 'member', 'label': 'Member', 'align': 'left'},
        {'key': 'loan_ref', 'label': 'Loan Reference', 'align': 'left'},
        {'key': 'principal', 'label': 'Principal (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_charged', 'label': 'Interest Charged (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_paid', 'label': 'Interest Paid (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_balance', 'label': 'Interest Balance (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
        {'key': 'disbursed_date', 'label': 'Disbursed', 'align': 'center', 'type': 'date'},
    ]

    totals = {
        'principal': total_principal,
        'interest_charged': total_interest_charged,
        'interest_paid': total_interest_paid,
        'interest_balance': total_interest_balance,
    }

    kpi_cards = [
        {'icon': 'bi-percent', 'value': f'UGX {total_interest_charged:,.0f}', 'label': 'Total Interest Charged', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_interest_paid:,.0f}', 'label': 'Interest Paid', 'type': 'success'},
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {total_interest_balance:,.0f}', 'label': 'Interest Outstanding', 'type': 'warning'},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Interest Income Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': date_from,
        'date_to': date_to,
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(data) > 0,
        'summary_totals': {
            'total_records': len(data),
            'total_amount': total_principal,
            'total_paid': total_interest_paid,
            'outstanding': total_interest_balance,
            'recovery_rate': (total_interest_paid / total_interest_charged * 100) if total_interest_charged > 0 else 0,
            'par_30': 'N/A',
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


@login_required
def loan_portfolio_reports(request):
    """Loan Portfolio Report - Comprehensive loan portfolio analytics"""
    from datetime import date

    if request.method == "POST":
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        officer_id = request.POST.get('officer')
    else:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        officer_id = request.GET.get('officer')

    loans = Loan.objects.select_related('member', 'officer').filter(
        status__in=['approved', 'active', 'closed']
    ).order_by('-disbursed_date', '-start_date')

    if start_date:
        loans = loans.filter(disbursed_date__gte=start_date)
    if end_date:
        loans = loans.filter(disbursed_date__lte=end_date)
    if officer_id:
        loans = loans.filter(officer_id=officer_id)

    today = date.today()
    report_data = []

    for loan in loans:
        p_bal = Decimal(str(loan.principal_balance or 0))
        i_bal = Decimal(str(loan.interest_balance or 0))

        overdue = loan.installments.filter(paid=False, due_date__lt=today)
        principal_in_arrears = overdue.aggregate(
            total=Coalesce(Sum('principal_portion'), Decimal('0'))
        )['total']

        total_due_today = loan.installments.filter(
            paid=False, due_date__lte=today
        ).aggregate(
            total=Coalesce(Sum(F('principal_portion') + F('interest_portion')), Decimal('0'))
        )['total']

        penalty_due = overdue.aggregate(
            total=Coalesce(Sum('penalty_amount'), Decimal('0'))
        )['total']

        report_data.append({
            'borrower': f"{loan.member.first_name} {loan.member.last_name}",
            'officer': loan.officer.get_full_name() if loan.officer else 'System',
            'account_number': loan.member.member_number,
            'contact': loan.member.phone_number,
            'loan_disbursed': Decimal(str(loan.principal_amount or 0)),
            'disbursement_date': loan.disbursed_date or loan.start_date,
            'principal_balance': p_bal,
            'interest_balance': i_bal,
            'principal_in_arrears': principal_in_arrears,
            'total_dues': total_due_today + penalty_due,
            'par': p_bal if principal_in_arrears > 0 else Decimal('0'),
        })

    total_disbursed = sum(item['loan_disbursed'] for item in report_data)
    total_outstanding = sum(item['principal_balance'] + item['interest_balance'] for item in report_data)
    total_par = sum(item['par'] for item in report_data)

    columns = [
        {'key': 'borrower', 'label': 'Borrower', 'align': 'left'},
        {'key': 'account_number', 'label': 'Account No', 'align': 'left'},
        {'key': 'loan_disbursed', 'label': 'Disbursed (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_balance', 'label': 'Principal (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_balance', 'label': 'Interest (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_in_arrears', 'label': 'Arrears (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'total_dues', 'label': 'Total Due (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'par', 'label': 'PAR (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'officer', 'label': 'Officer', 'align': 'left'},
        {'key': 'disbursement_date', 'label': 'Disbursed Date', 'align': 'center', 'type': 'date'},
    ]

    totals = {
        'loan_disbursed': total_disbursed,
        'principal_balance': total_outstanding,
        'interest_balance': sum(item['interest_balance'] for item in report_data),
        'principal_in_arrears': sum(item['principal_in_arrears'] for item in report_data),
        'total_dues': sum(item['total_dues'] for item in report_data),
        'par': total_par,
    }

    kpi_cards = [
        {'icon': 'bi-bank', 'value': f'UGX {total_disbursed:,.0f}', 'label': 'Total Disbursed', 'type': 'success'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_outstanding:,.0f}', 'label': 'Total Outstanding', 'type': 'info'},
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {total_par:,.0f}', 'label': 'Portfolio at Risk', 'type': 'danger'},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Loan Portfolio Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': start_date or 'All',
        'date_to': end_date or 'All',
        'columns': columns,
        'data': report_data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(report_data) > 0,
        'officer_list': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'summary_totals': {
            'total_records': len(report_data),
            'total_amount': total_disbursed,
            'total_paid': total_disbursed - total_outstanding,
            'outstanding': total_outstanding,
            'recovery_rate': ((total_disbursed - total_outstanding) / total_disbursed * 100) if total_disbursed > 0 else 0,
            'par_30': (total_par / total_disbursed * 100) if total_disbursed > 0 else 0,
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


@login_required
def portfolio_status_report(request):
    """Comprehensive Portfolio Status Report - with aging classification"""
    today = timezone.now().date()

    loans = Loan.objects.select_related('member', 'officer').filter(
        status__in=['approved', 'active', 'closed']
    ).order_by('member__member_number')

    report_data = []

    for loan in loans:
        paid_stats = loan.installments.filter(paid=True).aggregate(
            p_paid=Coalesce(Sum('principal_portion'), Decimal('0.00')),
            i_paid=Coalesce(Sum('interest_portion'), Decimal('0.00')),
            penalty_paid=Coalesce(Sum('penalty_amount'), Decimal('0.00')),
        )

        arrears_stats = loan.installments.filter(
            paid=False,
            due_date__lt=today
        ).aggregate(
            p_due=Coalesce(Sum('principal_portion'), Decimal('0.00')),
            i_due=Coalesce(Sum('interest_portion'), Decimal('0.00')),
            pen_due=Coalesce(Sum('penalty_amount'), Decimal('0.00')),
        )

        oldest_unpaid = loan.installments.filter(
            paid=False,
            due_date__lt=today
        ).order_by('due_date').first()

        classification = "Performing"
        if oldest_unpaid:
            days_past_due = (today - oldest_unpaid.due_date).days
            if days_past_due > 180:
                classification = "Loss"
            elif days_past_due > 90:
                classification = "Doubtful"
            elif days_past_due > 30:
                classification = "Substandard"
            else:
                classification = "Watch"

        report_data.append({
            'member_no': loan.member.member_number or str(loan.member.id),
            'name': f"{loan.member.first_name} {loan.member.last_name}",
            'loan_no': loan.loan_reference or loan.id,
            'disbursed_amount': Decimal(str(loan.principal_amount or 0)),
            'disbursed_date': loan.disbursed_date or loan.start_date,
            'principal_paid': paid_stats['p_paid'],
            'interest_paid': paid_stats['i_paid'],
            'penalty_paid': paid_stats['penalty_paid'],
            'principal_due': arrears_stats['p_due'],
            'interest_due': arrears_stats['i_due'],
            'penalty_due': arrears_stats['pen_due'],
            'total_due': arrears_stats['p_due'] + arrears_stats['i_due'] + arrears_stats['pen_due'],
            'principal_balance': Decimal(str(loan.principal_balance or 0)),
            'interest_balance': Decimal(str(loan.interest_balance or 0)),
            'classification': classification,
            'sector': getattr(loan.member, 'economic_sector', 'N/A'),
        })

    grand_total_disbursed = sum(item['disbursed_amount'] for item in report_data)
    grand_total_prin_paid = sum(item['principal_paid'] for item in report_data)
    grand_total_int_paid = sum(item['interest_paid'] for item in report_data)
    grand_total_prin_due = sum(item['principal_due'] for item in report_data)
    grand_total_int_due = sum(item['interest_due'] for item in report_data)
    grand_total_due = sum(item['total_due'] for item in report_data)
    grand_total_prin_bal = sum(item['principal_balance'] for item in report_data)
    grand_total_int_bal = sum(item['interest_balance'] for item in report_data)
    grand_total_exposure = grand_total_prin_bal + grand_total_int_bal

    columns = [
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'name', 'label': 'Member Name', 'align': 'left'},
        {'key': 'loan_no', 'label': 'Loan Ref', 'align': 'left'},
        {'key': 'disbursed_amount', 'label': 'Disbursed (UGX)', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_paid', 'label': 'Principal Paid', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_paid', 'label': 'Interest Paid', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_due', 'label': 'Principal Due', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_due', 'label': 'Interest Due', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'total_due', 'label': 'Total Due', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_balance', 'label': 'Principal Balance', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_balance', 'label': 'Interest Balance', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'classification', 'label': 'Classification', 'align': 'center'},
        {'key': 'sector', 'label': 'Sector', 'align': 'left'},
    ]

    totals = {
        'disbursed_amount': grand_total_disbursed,
        'principal_paid': grand_total_prin_paid,
        'interest_paid': grand_total_int_paid,
        'principal_due': grand_total_prin_due,
        'interest_due': grand_total_int_due,
        'total_due': grand_total_due,
        'principal_balance': grand_total_prin_bal,
        'interest_balance': grand_total_int_bal,
    }

    kpi_cards = [
        {'icon': 'bi-people', 'value': f'{len(report_data):,}', 'label': 'Total Loans', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {grand_total_exposure:,.0f}', 'label': 'Total Exposure', 'type': 'success'},
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {grand_total_due:,.0f}', 'label': 'Total Arrears', 'type': 'danger'},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Portfolio Status Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': 'All',
        'date_to': 'All',
        'columns': columns,
        'data': report_data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(report_data) > 0,
        'summary_totals': {
            'total_records': len(report_data),
            'total_amount': grand_total_disbursed,
            'total_paid': grand_total_prin_paid + grand_total_int_paid,
            'outstanding': grand_total_prin_bal + grand_total_int_bal,
            'recovery_rate': ((grand_total_prin_paid + grand_total_int_paid) / grand_total_disbursed * 100) if grand_total_disbursed > 0 else 0,
            'par_30': (grand_total_due / grand_total_disbursed * 100) if grand_total_disbursed > 0 else 0,
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


@login_required
def arrears_report(request):
    """Arrears & Delinquency Report with aging buckets"""
    from datetime import datetime

    if request.method == "POST":
        date_at_str = request.POST.get('date_at')
        search_query = request.POST.get('search_query')
    else:
        date_at_str = request.GET.get('date_at')
        search_query = request.GET.get('search_query')

    today = date.today()
    if date_at_str:
        try:
            target_date = datetime.strptime(date_at_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = today
    else:
        target_date = today

    overdue_installments = Installment.objects.filter(
        paid=False,
        due_date__lt=target_date,
        loan__status__in=['approved', 'active', 'arrears']
    ).select_related('loan', 'loan__member', 'loan__officer')

    if search_query:
        overdue_installments = overdue_installments.filter(
            Q(loan__member__first_name__icontains=search_query) |
            Q(loan__member__last_name__icontains=search_query) |
            Q(loan__member__member_number__icontains=search_query) |
            Q(loan__loan_reference__icontains=search_query)
        )

    overdue_installments = overdue_installments.annotate(
        arrears_amount=(
            Coalesce(F('principal_portion') - F('principal_paid'), Decimal('0')) +
            Coalesce(F('interest_portion') - F('interest_paid'), Decimal('0')) +
            Coalesce(F('penalty_amount') - F('penalty_paid'), Decimal('0'))
        ),
        days_overdue=(target_date - F('due_date'))
    )

    overdue_installments = overdue_installments.order_by('-days_overdue')

    data = []
    total_arrears = Decimal('0')

    for inst in overdue_installments:
        days = (target_date - inst.due_date).days
        data.append({
            'member_no': inst.loan.member.member_number or str(inst.loan.member.id),
            'member_name': f"{inst.loan.member.first_name} {inst.loan.member.last_name}",
            'loan_ref': inst.loan.loan_reference or f"LN-{inst.loan.id}",
            'phone': inst.loan.member.phone_number,
            'due_date': inst.due_date,
            'days_overdue': days,
            'principal_due': inst.principal_portion - inst.principal_paid,
            'interest_due': inst.interest_portion - inst.interest_paid,
            'penalty_due': inst.penalty_amount - inst.penalty_paid,
            'total_due': (inst.principal_portion - inst.principal_paid) +
                         (inst.interest_portion - inst.interest_paid) +
                         (inst.penalty_amount - inst.penalty_paid),
            'officer': inst.loan.officer.get_full_name() if inst.loan.officer else 'System',
        })
        total_arrears += data[-1]['total_due']

    aging_buckets = {
        '1-30_days': Decimal('0'),
        '31-60_days': Decimal('0'),
        '61-90_days': Decimal('0'),
        '91-180_days': Decimal('0'),
        '180_plus': Decimal('0'),
    }

    for item in data:
        days = item['days_overdue']
        amount = item['total_due']
        if 1 <= days <= 30:
            aging_buckets['1-30_days'] += amount
        elif 31 <= days <= 60:
            aging_buckets['31-60_days'] += amount
        elif 61 <= days <= 90:
            aging_buckets['61-90_days'] += amount
        elif 91 <= days <= 180:
            aging_buckets['91-180_days'] += amount
        else:
            aging_buckets['180_plus'] += amount

    columns = [
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'member_name', 'label': 'Member', 'align': 'left'},
        {'key': 'loan_ref', 'label': 'Loan Ref', 'align': 'left'},
        {'key': 'phone', 'label': 'Phone', 'align': 'left'},
        {'key': 'due_date', 'label': 'Due Date', 'align': 'center', 'type': 'date'},
        {'key': 'days_overdue', 'label': 'Days Overdue', 'align': 'center'},
        {'key': 'principal_due', 'label': 'Principal Due', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_due', 'label': 'Interest Due', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'penalty_due', 'label': 'Penalty Due', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'total_due', 'label': 'Total Due', 'align': 'right', 'type': 'currency', 'total': True, 'prefix': 'UGX '},
        {'key': 'officer', 'label': 'Officer', 'align': 'left'},
    ]

    totals = {
        'principal_due': sum(item['principal_due'] for item in data),
        'interest_due': sum(item['interest_due'] for item in data),
        'penalty_due': sum(item['penalty_due'] for item in data),
        'total_due': total_arrears,
    }

    total_outstanding_loans = Loan.objects.filter(
        is_active=True,
        status__in=['approved', 'active', 'arrears']
    ).count()

    kpi_cards = [
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {total_arrears:,.0f}', 'label': 'Total Arrears', 'type': 'danger'},
        {'icon': 'bi-clock-history', 'value': f'{len(data):,}', 'label': 'Overdue Installments', 'type': 'warning'},
        {'icon': 'bi-percent', 'value': f'{(total_arrears / (total_outstanding_loans + 1)):.1f}%', 'label': 'Arrears Rate', 'type': 'info'},
    ]

    aging_summary = [
        {'bucket': '1-30 Days', 'amount': aging_buckets['1-30_days']},
        {'bucket': '31-60 Days', 'amount': aging_buckets['31-60_days']},
        {'bucket': '61-90 Days', 'amount': aging_buckets['61-90_days']},
        {'bucket': '91-180 Days', 'amount': aging_buckets['91-180_days']},
        {'bucket': '180+ Days', 'amount': aging_buckets['180_plus']},
    ]

    context = {
        'company': Company.get_company(),
        'report_title': 'Arrears & Delinquency Report',
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_date': timezone.now().strftime('%d %b, %Y %H:%M'),
        'date_from': target_date.strftime('%Y-%m-%d'),
        'date_to': target_date.strftime('%Y-%m-%d'),
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'has_data': len(data) > 0,
        'aging_summary': aging_summary,
        'summary_totals': {
            'total_records': len(data),
            'total_amount': total_arrears,
            'total_paid': 'N/A',
            'outstanding': total_arrears,
            'recovery_rate': 'N/A',
            'par_30': f'{(total_arrears / (total_outstanding_loans + 1)):.1f}',
        },
    }

    return render(request, 'finance/reports/base_report.html', context)


# finance/views.py (add/replace this function)

from decimal import Decimal
from django.shortcuts import render
from django.http import FileResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from .models import GeneralLedger, ChartOfAccount, Company
from .utils import generate_excel_report

User = get_user_model()

@login_required
def general_ledger_report(request):
    """
    Professional General Ledger report with date range, account, and account type filters.
    Supports HTML display and Excel export.
    """
    # --- 1. Extract filters from GET or POST ---
    date_from = request.GET.get('date_from') or request.POST.get('date_from')
    date_to = request.GET.get('date_to') or request.POST.get('date_to')
    account_id = request.GET.get('account') or request.POST.get('account')
    account_type = request.GET.get('account_type') or request.POST.get('account_type')
    # Status is included for compatibility with base template – can be used later
    status = request.GET.get('status') or request.POST.get('status')
    # Officer is not used in GL report, but we pass an empty list for the template

    # --- 2. Base queryset ---
    qs = GeneralLedger.objects.select_related('account')

    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if account_id:
        qs = qs.filter(account_id=account_id)
    if account_type:
        qs = qs.filter(account__account_type=account_type)

    # --- 3. Build data rows ---
    data = []
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')

    for entry in qs.order_by('date', 'id'):
        data.append({
            'date': entry.date,
            'account_code': entry.account.code,
            'account_name': entry.account.name,
            'description': entry.description,
            'reference': entry.reference or '-',
            'debit': entry.debit,
            'credit': entry.credit,
            'balance': entry.balance,
        })
        total_debit += entry.debit
        total_credit += entry.credit

    # --- 4. Define columns ---
    columns = [
        {'key': 'date', 'label': 'Date', 'type': 'date', 'align': 'center'},
        {'key': 'account_code', 'label': 'Account Code', 'align': 'left'},
        {'key': 'account_name', 'label': 'Account Name', 'align': 'left'},
        {'key': 'description', 'label': 'Description', 'align': 'left'},
        {'key': 'reference', 'label': 'Reference', 'align': 'left'},
        {'key': 'debit', 'label': 'Debit (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'credit', 'label': 'Credit (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'balance', 'label': 'Balance (UGX)', 'type': 'currency', 'align': 'right', 'prefix': 'UGX '},
    ]

    totals = {'debit': total_debit, 'credit': total_credit}

    # --- 5. KPIs ---
    kpi_cards = [
        {'label': 'Total Entries', 'value': qs.count(), 'icon': 'bi-list-ul', 'type': 'info'},
        {'label': 'Total Debit', 'value': f"UGX {total_debit:,.0f}", 'icon': 'bi-arrow-down', 'type': 'danger'},
        {'label': 'Total Credit', 'value': f"UGX {total_credit:,.0f}", 'icon': 'bi-arrow-up', 'type': 'success'},
        {'label': 'Net Movement', 'value': f"UGX {abs(total_debit - total_credit):,.0f}", 'icon': 'bi-arrows-vertical', 'type': 'warning'},
    ]

    # --- 6. Summary totals (bottom section) ---
    summary_totals = {
        'total_records': qs.count(),
        'total_amount': total_debit + total_credit,
        'total_paid': total_credit,   # Placeholder, not really applicable
        'outstanding': total_debit,
        'recovery_rate': 'N/A',
        'par_30': 'N/A',
    }

    # --- 7. Account list for dropdown ---
    account_list = ChartOfAccount.objects.filter(is_active=True).order_by('code')

    # --- 8. Account type choices ---
    account_type_choices = ChartOfAccount.ACCOUNT_TYPES

    # --- 9. Selected values for display in filter badges ---
    selected_account_display = None
    if account_id:
        try:
            acc = ChartOfAccount.objects.get(id=account_id)
            selected_account_display = f"{acc.code} – {acc.name}"
        except ChartOfAccount.DoesNotExist:
            pass

    selected_account_type_display = dict(account_type_choices).get(account_type)

    # --- 10. Officer list (required by base template – empty for GL) ---
    officer_list = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    # --- 11. Build final context ---
    context = {
        'columns': columns,
        'data': data,
        'totals': totals,
        'has_data': bool(data),
        'kpi_cards': kpi_cards,
        'report_title': 'General Ledger Report',
        'company': Company.get_company(),
        'date_from': date_from,
        'date_to': date_to,
        'selected_account': account_id,
        'selected_account_type': account_type,
        'selected_status': status,               # optional – passed to template
        'selected_account_display': selected_account_display,
        'selected_account_type_display': selected_account_type_display,
        'account_list': account_list,
        'account_type_choices': account_type_choices,
        'officer_list': officer_list,            # for the officer dropdown
        'officer_name': None,                    # not used in GL
        'summary_totals': summary_totals,
        'generated_date': timezone.now().strftime('%d %b %Y %H:%M'),
        'generated_by': request.user.get_full_name() if request.user.is_authenticated else 'System',
    }

    # --- 12. Handle Excel export ---
    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"General_Ledger_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        response = FileResponse(
            excel_file,
            as_attachment=True,
            filename=filename,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        return response

    return render(request, 'finance/reports/base_report.html', context)
def report_view(request):
    """
    Main report view – handles both HTML display and Excel export.
    """
    date_from = request.POST.get('date_from') or request.GET.get('date_from')
    date_to = request.POST.get('date_to') or request.GET.get('date_to')
    officer_id = request.POST.get('officer') or request.GET.get('officer')
    status = request.POST.get('status') or request.GET.get('status')

    qs = Loan.objects.select_related('member', 'officer')

    if date_from:
        qs = qs.filter(disbursed_date__gte=date_from)
    if date_to:
        qs = qs.filter(disbursed_date__lte=date_to)
    if officer_id:
        qs = qs.filter(officer_id=officer_id)
    if status:
        qs = qs.filter(status=status)

    columns = [
        {'key': 'loan_reference', 'label': 'Loan Reference'},
        {'key': 'member_name', 'label': 'Member'},
        {'key': 'principal', 'label': 'Principal', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'interest_balance', 'label': 'Interest Balance', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'total_balance', 'label': 'Total Balance', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'status', 'label': 'Status', 'type': 'status'},
        {'key': 'disbursed_date', 'label': 'Disbursed', 'type': 'date'},
        {'key': 'officer', 'label': 'Officer'},
    ]

    data = []
    total_principal = Decimal('0.00')
    total_interest = Decimal('0.00')
    total_balance = Decimal('0.00')

    for loan in qs:
        principal = loan.principal_amount or Decimal('0.00')
        interest = loan.interest_balance or Decimal('0.00')
        balance = principal + interest

        data.append({
            'loan_reference': loan.loan_reference or f"LN-{loan.id}",
            'member_name': loan.member.get_full_name() if loan.member else 'N/A',
            'principal': principal,
            'interest_balance': interest,
            'total_balance': balance,
            'status': loan.get_status_display(),
            'disbursed_date': loan.disbursed_date,
            'officer': loan.officer.get_full_name() if loan.officer else 'N/A',
        })

        total_principal += principal
        total_interest += interest
        total_balance += balance

    totals = {
        'principal': total_principal,
        'interest_balance': total_interest,
        'total_balance': total_balance,
    }

    record_count = len(data)
    kpi_cards = [
        {'label': 'Total Loans', 'value': record_count, 'icon': 'bi-file-earmark-text', 'type': 'info'},
        {'label': 'Total Principal', 'value': f"UGX {total_principal:,.0f}", 'icon': 'bi-cash', 'type': 'success'},
        {'label': 'Total Interest', 'value': f"UGX {total_interest:,.0f}", 'icon': 'bi-percent', 'type': 'warning'},
        {'label': 'Total Balance', 'value': f"UGX {total_balance:,.0f}", 'icon': 'bi-wallet2', 'type': 'danger'},
    ]

    summary_totals = {
        'total_records': record_count,
        'total_amount': total_balance,
        'total_paid': Decimal('0.00'),
        'outstanding': total_balance,
        'recovery_rate': 0,
        'par_30': 0,
    }

    officer_list = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    context = {
        'columns': columns,
        'data': data,
        'totals': totals,
        'has_data': bool(data),
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'aging_summary': [],
        'report_title': 'Loan Portfolio Report',
        'company': Company.get_company(),
        'date_from': date_from,
        'date_to': date_to,
        'selected_officer': officer_id,
        'selected_status': status,
        'officer_list': officer_list,
        'officer_name': dict(officer_list.values_list('id', 'username')).get(int(officer_id) if officer_id else None),
        'generated_date': timezone.now().strftime('%d %b %Y %H:%M'),
        'generated_by': request.user.get_full_name() if request.user.is_authenticated else 'System',
    }

    if request.POST.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        response = FileResponse(
            excel_file,
            as_attachment=True,
            filename=filename,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        return response

    return render(request, 'finance/reports/base_report.html', context)


# finance/views.py
from django.contrib.auth.decorators import permission_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Loan, Installment, ManualPenalty
# finance/views.py
from decimal import Decimal
from django.contrib.auth.decorators import permission_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.utils import timezone
from .models import Loan, Installment, ManualPenalty


# finance/views.py
from decimal import Decimal
from django.contrib.auth.decorators import permission_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.utils import timezone
from .models import Loan, Installment, ManualPenalty

@permission_required('finance.can_apply_manual_penalty')
def apply_manual_penalty(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)

    if request.method == 'POST':
        amount_str = request.POST.get('amount')
        reason = request.POST.get('reason')
        installment_id = request.POST.get('installment_id')

        # Validate amount
        try:
            amount = Decimal(amount_str)
        except (ValueError, TypeError):
            messages.error(request, "Invalid amount. Please enter a valid number.")
            return redirect('loan_detail', pk=loan.id)

        if amount <= 0:
            messages.error(request, "Amount must be greater than zero.")
            return redirect('loan_detail', pk=loan.id)

        if not reason or reason.strip() == '':
            messages.error(request, "Please provide a reason for the penalty.")
            return redirect('loan_detail', pk=loan.id)

        installment = None
        if installment_id:
            installment = get_object_or_404(Installment, id=installment_id, loan=loan)

        # Create the manual penalty
        penalty = ManualPenalty.objects.create(
            loan=loan,
            installment=installment,
            amount=amount,
            reason=reason,
            applied_by=request.user,
        )

        messages.success(request, f"Manual penalty of UGX {amount:,.2f} applied to loan {loan.loan_reference}.")
        return redirect('loan_detail', pk=loan.id)

    # GET: show form
    installments = loan.installments.filter(paid=False).order_by('due_date')
    context = {
        'loan': loan,
        'installments': installments,
    }
    return render(request, 'finance/apply_manual_penalty.html', context)

@permission_required('finance.can_waive_penalty')
def waive_manual_penalty(request, penalty_id):
    """
    View to waive an active manual penalty.
    """
    penalty = get_object_or_404(ManualPenalty, id=penalty_id)

    if request.method == 'POST':
        reason = request.POST.get('reason')

        if not penalty.is_waived:
            penalty.is_waived = True
            penalty.waived_by = request.user
            penalty.waived_date = timezone.now()
            penalty.waiver_reason = reason
            penalty.save()
            messages.success(request, "Penalty waived successfully.")
        else:
            messages.warning(request, "This penalty has already been waived.")

        return redirect('loan_detail', pk=penalty.loan.id)

    # GET: show confirmation page
    context = {'penalty': penalty}
    return render(request, 'finance/waive_penalty.html', context)



from decimal import Decimal
from django.db.models import Sum, Q, F, Value, Count, Case, When, IntegerField
from django.db.models.functions import Coalesce
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Loan, Installment, Company
from decimal import Decimal
from django.db.models import Sum, Q
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Loan, Company

@login_required
def loan_portfolio_report(request):
    """
    Comprehensive Loan Portfolio Report with classification.
    Filters: date range, officer, product, status, classification, loan_status.
    """
    # Extract filters
    date_from = request.GET.get('date_from') or request.POST.get('date_from')
    date_to = request.GET.get('date_to') or request.POST.get('date_to')
    officer_id = request.GET.get('officer') or request.POST.get('officer')
    product = request.GET.get('product') or request.POST.get('product')
    status_filter = request.GET.get('status') or request.POST.get('status')  # 'active', 'pending', etc.
    classification_filter = request.GET.get('classification') or request.POST.get('classification')
    loan_status_filter = request.GET.get('loan_status') or request.POST.get('loan_status')  # 'outstanding' / 'closed'

    # Base queryset
    qs = Loan.objects.select_related('member', 'officer').prefetch_related('installments')

    if date_from:
        qs = qs.filter(disbursed_date__gte=date_from)
    if date_to:
        qs = qs.filter(disbursed_date__lte=date_to)
    if officer_id:
        qs = qs.filter(officer_id=officer_id)
    if product:
        qs = qs.filter(product_type=product)
    if status_filter:
        qs = qs.filter(status=status_filter)

    # Build data
    data = []
    total_disbursed = Decimal('0')
    total_outstanding = Decimal('0')
    total_closed = Decimal('0')

    today = timezone.now().date()

    for loan in qs:
        principal_bal = loan.principal_balance or Decimal('0')
        interest_bal = loan.interest_balance or Decimal('0')
        total_bal = principal_bal + interest_bal

        is_closed = (total_bal == Decimal('0')) or (loan.status == 'closed')

        # Apply loan_status filter (outstanding/closed)
        if loan_status_filter and loan_status_filter != 'all':
            if loan_status_filter == 'outstanding' and is_closed:
                continue
            if loan_status_filter == 'closed' and not is_closed:
                continue

        # Classification
        overdue_inst = loan.installments.filter(paid=False, due_date__lt=today).order_by('due_date').first()
        classification = 'Performing'
        days_overdue = 0
        if overdue_inst:
            days_overdue = (today - overdue_inst.due_date).days
            if days_overdue > 180:
                classification = 'Loss'
            elif days_overdue > 90:
                classification = 'Doubtful'
            elif days_overdue > 30:
                classification = 'Substandard'
            else:
                classification = 'Watch'

        if classification_filter and classification_filter != 'all':
            if classification_filter.lower() != classification.lower():
                continue

        data.append({
            'loan_ref': loan.loan_reference or f"LN-{loan.id}",
            'member': f"{loan.member.first_name} {loan.member.last_name}",
            'member_no': loan.member.member_number,
            'product': loan.get_product_type_display(),
            'disbursed_date': loan.disbursed_date or loan.start_date,
            'principal': loan.principal_amount,
            'principal_balance': principal_bal,
            'interest_balance': interest_bal,
            'total_balance': total_bal,
            'is_closed': is_closed,
            'status': 'Closed' if is_closed else loan.status.title(),
            'classification': classification,
            'days_overdue': days_overdue,
            'officer': loan.officer.get_full_name() if loan.officer else 'System',
        })

        total_disbursed += loan.principal_amount
        if not is_closed:
            total_outstanding += total_bal
        else:
            total_closed += 1

    # Columns for the table
    columns = [
        {'key': 'loan_ref', 'label': 'Loan Ref'},
        {'key': 'member', 'label': 'Member'},
        {'key': 'member_no', 'label': 'Member No'},
        {'key': 'product', 'label': 'Product'},
        {'key': 'disbursed_date', 'label': 'Disbursed', 'type': 'date'},
        {'key': 'principal', 'label': 'Principal (UGX)', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'principal_balance', 'label': 'Principal Bal (UGX)', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'interest_balance', 'label': 'Interest Bal (UGX)', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'total_balance', 'label': 'Total Balance (UGX)', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'classification', 'label': 'Classification', 'align': 'center', 'type': 'status'},
        {'key': 'days_overdue', 'label': 'Days Overdue', 'align': 'center'},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
        {'key': 'officer', 'label': 'Officer'},
    ]

    # Totals
    totals = {
        'principal': sum(item['principal'] for item in data),
        'principal_balance': sum(item['principal_balance'] for item in data),
        'interest_balance': sum(item['interest_balance'] for item in data),
        'total_balance': sum(item['total_balance'] for item in data),
    }

    # KPIs
    kpi_cards = [
        {'label': 'Total Loans', 'value': len(data), 'icon': 'bi-file-text', 'type': 'info'},
        {'label': 'Total Disbursed', 'value': f"UGX {total_disbursed:,.0f}", 'icon': 'bi-arrow-up', 'type': 'success'},
        {'label': 'Outstanding Balance', 'value': f"UGX {total_outstanding:,.0f}", 'icon': 'bi-currency-dollar', 'type': 'warning'},
        {'label': 'Closed Loans', 'value': total_closed, 'icon': 'bi-check-circle', 'type': 'secondary'},
    ]

    from django.contrib.auth import get_user_model
    User = get_user_model()
    officer_list = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    product_choices = Loan.PRODUCT_CHOICES

    context = {
        'columns': columns,
        'data': data,
        'totals': totals,
        'has_data': bool(data),
        'kpi_cards': kpi_cards,
        'report_title': 'Loan Portfolio Report',
        'company': Company.get_company(),
        'date_from': date_from,
        'date_to': date_to,
        'selected_officer': officer_id,
        'selected_product': product,
        'selected_status': status_filter,
        'selected_classification': classification_filter,
        'selected_loan_status': loan_status_filter,
        'officer_list': officer_list,
        'product_choices': product_choices,
        'classification_choices': [
            ('all', 'All Classifications'),
            ('performing', 'Performing'),
            ('watch', 'Watch'),
            ('substandard', 'Substandard'),
            ('doubtful', 'Doubtful'),
            ('loss', 'Loss'),
        ],
        'loan_status_choices': [
            ('all', 'All Loans'),
            ('outstanding', 'Outstanding'),
            ('closed', 'Closed'),
        ],
        'generated_date': timezone.now().strftime('%d %b %Y %H:%M'),
        'generated_by': request.user.get_full_name() if request.user.is_authenticated else 'System',
    }

    return render(request, 'finance/reports/loan_portfolio_report.html', context)


from django.contrib.auth.decorators import permission_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.utils import timezone
from decimal import Decimal
from .models import Installment

@permission_required('finance.can_waive_penalty')  # or any suitable permission
def waive_auto_penalty(request, installment_id):
    """
    Waive the auto-calculated penalty for a specific installment.
    Sets penalty_amount to 0 and records who/why.
    """
    installment = get_object_or_404(Installment, id=installment_id)

    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        if not reason:
            messages.error(request, "Please provide a reason for waiving the penalty.")
            return redirect('loan_detail', pk=installment.loan.id)

        # Perform the waiver
        installment.penalty_amount = Decimal('0.00')
        installment.penalty_waived = True
        installment.penalty_waived_by = request.user
        installment.penalty_waived_date = timezone.now()
        installment.penalty_waiver_reason = reason
        installment.save(update_fields=[
            'penalty_amount', 'penalty_waived', 'penalty_waived_by',
            'penalty_waived_date', 'penalty_waiver_reason'
        ])

        messages.success(request, f"Penalty for installment #{installment.id} has been waived.")
        return redirect('loan_detail', pk=installment.loan.id)

    # GET: show confirmation form
    return render(request, 'finance/waive_auto_penalty.html', {
        'installment': installment,
        'loan': installment.loan,
    })


# finance/views.py
import random
from decimal import Decimal
from datetime import datetime, date, timedelta
from django.db.models import Sum, Q, F, Count, Case, When, Value, DecimalField
from django.db.models.functions import Coalesce, TruncMonth
from django.shortcuts import render
from django.http import FileResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User, Group
from .models import (
    Loan, Installment, Member, SystemSetting, GeneralLedger, ChartOfAccount,
    SavingsAccount, Transaction, Company, SMSConfig
)
from .utils import generate_excel_report

User = get_user_model()

# ====================================================================
# HELPER: Generate common context for base_report.html
# ====================================================================
def _get_base_context(request, extra_context=None):
    """
    Returns a base context dict with company, officer list, account list, etc.
    Override with extra_context to customise per report.
    """
    context = {
        'company': Company.get_company(),
        'officer_list': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'account_list': ChartOfAccount.objects.filter(is_active=True).order_by('code'),
        'account_type_choices': ChartOfAccount.ACCOUNT_TYPES,
        'generated_date': timezone.now().strftime('%d %b %Y %H:%M'),
        'generated_by': request.user.get_full_name() if request.user.is_authenticated else 'System',
        'has_data': False,
        'data': [],
        'columns': [],
        'totals': {},
        'kpi_cards': [],
        'summary_totals': {},
        'aging_summary': [],
        # filter defaults – override in each view
        'date_from': None,
        'date_to': None,
        'selected_officer': None,
        'selected_account': None,
        'selected_account_type': None,
        'selected_status': None,
        'selected_account_display': None,
        'selected_account_type_display': None,
        'officer_name': None,
    }
    if extra_context:
        context.update(extra_context)
    return context


# ====================================================================
# 1. LOAN PORTFOLIO REPORT (loan_report)
# ====================================================================
@login_required
def loan_report(request):
    """Professional Loan Portfolio Report with Filters"""
    date_from = request.GET.get('date_from') or request.POST.get('date_from')
    date_to = request.GET.get('date_to') or request.POST.get('date_to')
    officer_id = request.GET.get('officer') or request.POST.get('officer')
    status = request.GET.get('status') or request.POST.get('status')
    account = request.GET.get('account') or request.POST.get('account')  # not used but kept for consistency

    loans = Loan.objects.select_related('member', 'officer').prefetch_related('repayments')
    if date_from:
        loans = loans.filter(disbursed_date__gte=date_from)
    if date_to:
        loans = loans.filter(disbursed_date__lte=date_to)
    if officer_id:
        loans = loans.filter(officer_id=officer_id)
    if status:
        loans = loans.filter(status=status)

    loans = loans.annotate(
        paid_amount=Sum('repayments__amount_paid', default=0),
        total_balance=F('principal_balance') + F('interest_balance'),
    )

    data = []
    total_amount = Decimal('0')
    total_paid = Decimal('0')
    total_balance = Decimal('0')

    for loan in loans:
        balance = loan.total_balance or Decimal('0')
        paid = loan.paid_amount or Decimal('0')

        data.append({
            'member': f"{loan.member.first_name} {loan.member.last_name}".strip(),
            'member_no': loan.member.member_number or str(loan.member.id),
            'loan_ref': loan.loan_reference or f"LN-{loan.id}",
            'amount': loan.principal_amount,
            'paid': paid,
            'balance': balance,
            'status': loan.get_status_display(),
            'date': loan.disbursed_date or loan.start_date,
            'officer': loan.officer.get_full_name() if loan.officer else 'System',
        })
        total_amount += loan.principal_amount or Decimal('0')
        total_paid += paid
        total_balance += balance

    recovery_rate = round((total_paid / total_amount * 100), 1) if total_amount > 0 else 0
    par_30 = round((total_balance / total_amount * 100), 1) if total_amount > 0 else 0
    total_records = len(data)

    columns = [
        {'key': 'member', 'label': 'Member Name', 'align': 'left'},
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'loan_ref', 'label': 'Loan Reference', 'align': 'left'},
        {'key': 'amount', 'label': 'Amount (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'paid', 'label': 'Paid (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'balance', 'label': 'Balance (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
        {'key': 'date', 'label': 'Disbursed', 'align': 'center', 'type': 'date'},
        {'key': 'officer', 'label': 'Officer', 'align': 'left'},
    ]

    totals = {'amount': total_amount, 'paid': total_paid, 'balance': total_balance}
    kpi_cards = [
        {'icon': 'bi-people', 'value': f'{total_records:,}', 'label': 'Total Loans', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_amount:,.0f}', 'label': 'Total Portfolio', 'type': 'success'},
        {'icon': 'bi-check-circle', 'value': f'{recovery_rate}%', 'label': 'Recovery Rate', 'type': 'info'},
        {'icon': 'bi-exclamation-triangle', 'value': f'{par_30}%', 'label': 'PAR 30', 'type': 'warning'},
    ]
    summary_totals = {
        'total_records': total_records,
        'total_amount': total_amount,
        'total_paid': total_paid,
        'outstanding': total_balance,
        'recovery_rate': recovery_rate,
        'par_30': par_30,
    }

    selected_officer_display = None
    if officer_id:
        try:
            off = User.objects.get(id=officer_id)
            selected_officer_display = off.get_full_name() or off.username
        except User.DoesNotExist:
            pass

    context = _get_base_context(request, {
        'report_title': 'Loan Portfolio Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'date_from': date_from,
        'date_to': date_to,
        'selected_officer': officer_id,
        'selected_status': status,
        'officer_name': selected_officer_display,
        'selected_account': account,  # not used but passed
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 2. MEMBER REPORT (member_report)
# ====================================================================
@login_required
def member_report(request):
    """Member Registry Report"""
    date_from = request.GET.get('date_from') or request.POST.get('date_from')
    date_to = request.GET.get('date_to') or request.POST.get('date_to')
    status_filter = request.GET.get('status') or request.POST.get('status')

    members = Member.objects.all().order_by('member_number')
    if date_from:
        members = members.filter(date_joined__gte=date_from)
    if date_to:
        members = members.filter(date_joined__lte=date_to)
    # status filter could be used if Member has a status field – here we ignore

    data = []
    total_savings = Decimal('0')
    total_loans = Decimal('0')

    for member in members:
        savings = SavingsAccount.objects.filter(member=member).first()
        savings_balance = savings.balance if savings else Decimal('0')
        total_loans_balance = member.loans.filter(is_active=True).aggregate(
            total=Sum('principal_balance') + Sum('interest_balance')
        )['total'] or Decimal('0')

        status_display = 'Active'  # placeholder
        data.append({
            'member_no': member.member_number,
            'name': f"{member.first_name} {member.last_name}",
            'phone': member.phone_number,
            'email': member.email or '—',
            'savings': savings_balance,
            'loans': total_loans_balance,
            'joined': member.date_joined,
            'status': status_display,
        })
        total_savings += savings_balance
        total_loans += total_loans_balance

    columns = [
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'name', 'label': 'Member Name', 'align': 'left'},
        {'key': 'phone', 'label': 'Phone', 'align': 'left'},
        {'key': 'email', 'label': 'Email', 'align': 'left'},
        {'key': 'savings', 'label': 'Savings (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'loans', 'label': 'Loans (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'joined', 'label': 'Joined', 'align': 'center', 'type': 'date'},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
    ]
    totals = {'savings': total_savings, 'loans': total_loans}
    kpi_cards = [
        {'icon': 'bi-people', 'value': f'{len(data):,}', 'label': 'Total Members', 'type': 'info'},
        {'icon': 'bi-wallet2', 'value': f'UGX {total_savings:,.0f}', 'label': 'Total Savings', 'type': 'success'},
        {'icon': 'bi-bank', 'value': f'UGX {total_loans:,.0f}', 'label': 'Total Loans', 'type': 'info'},
    ]
    summary_totals = {
        'total_records': len(data),
        'total_amount': total_savings + total_loans,
        'total_paid': total_savings,
        'outstanding': total_loans,
        'recovery_rate': '100',
        'par_30': '0',
    }

    context = _get_base_context(request, {
        'report_title': 'Member Registry Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'date_from': date_from,
        'date_to': date_to,
        'selected_status': status_filter,
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 3. SAVINGS REPORT (savings_report)
# ====================================================================
@login_required
def savings_report(request):
    """Savings Summary Report"""
    date_from = request.GET.get('date_from') or request.POST.get('date_from')
    date_to = request.GET.get('date_to') or request.POST.get('date_to')
    # Filter by date range on transactions? We'll filter savings accounts by member join date or ignore.

    savings_accounts = SavingsAccount.objects.select_related('member').all()
    # Could filter by date if member has date_joined, but not required

    data = []
    total_balance = Decimal('0')

    for savings in savings_accounts:
        data.append({
            'member': f"{savings.member.first_name} {savings.member.last_name}",
            'member_no': savings.member.member_number,
            'phone': savings.member.phone_number,
            'balance': savings.balance,
            'account_no': savings.account_number if hasattr(savings, 'account_number') else 'N/A',
            'status': 'Active' if savings.balance > 0 else 'Inactive',
        })
        total_balance += savings.balance

    columns = [
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'member', 'label': 'Member Name', 'align': 'left'},
        {'key': 'phone', 'label': 'Phone', 'align': 'left'},
        {'key': 'account_no', 'label': 'Account No', 'align': 'left'},
        {'key': 'balance', 'label': 'Balance (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
    ]
    totals = {'balance': total_balance}
    kpi_cards = [
        {'icon': 'bi-wallet2', 'value': f'{len(data):,}', 'label': 'Total Accounts', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_balance:,.0f}', 'label': 'Total Savings', 'type': 'success'},
        {'icon': 'bi-people', 'value': f'{len(data):,}', 'label': 'Active Members', 'type': 'info'},
    ]
    summary_totals = {
        'total_records': len(data),
        'total_amount': total_balance,
        'total_paid': total_balance,
        'outstanding': 0,
        'recovery_rate': '100',
        'par_30': '0',
    }

    context = _get_base_context(request, {
        'report_title': 'Savings Summary Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'date_from': date_from,
        'date_to': date_to,
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 4. FINANCIAL PERFORMANCE REPORT (financial_report)
# ====================================================================
@login_required
def financial_report(request):
    """Income/Expense Financial Performance Report"""
    if request.method == "POST":
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
    else:
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')

    if not date_from:
        date_from = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = date.today().strftime('%Y-%m-%d')

    transactions = Transaction.objects.filter(
        timestamp__date__gte=date_from,
        timestamp__date__lte=date_to
    )

    total_income = transactions.filter(
        Q(type='deposit') | Q(type='repayment') | Q(type='interest_payment')
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    total_expenses = transactions.filter(
        Q(type='withdrawal') | Q(type='disbursement') | Q(type='penalty')
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    net_profit = total_income - total_expenses

    monthly_data = transactions.annotate(
        month=TruncMonth('timestamp')
    ).values('month').annotate(
        income=Sum('amount', filter=Q(type__in=['deposit', 'repayment', 'interest_payment'])),
        expense=Sum('amount', filter=Q(type__in=['withdrawal', 'disbursement', 'penalty']))
    ).order_by('month')

    data = []
    for entry in monthly_data:
        data.append({
            'month': entry['month'].strftime('%b %Y') if entry['month'] else 'N/A',
            'income': entry['income'] or Decimal('0'),
            'expense': entry['expense'] or Decimal('0'),
            'profit': (entry['income'] or Decimal('0')) - (entry['expense'] or Decimal('0')),
        })

    columns = [
        {'key': 'month', 'label': 'Month', 'align': 'left'},
        {'key': 'income', 'label': 'Income (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'expense', 'label': 'Expenses (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'profit', 'label': 'Net Profit (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
    ]
    totals = {'income': total_income, 'expense': total_expenses, 'profit': net_profit}
    kpi_cards = [
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_income:,.0f}', 'label': 'Total Income', 'type': 'success'},
        {'icon': 'bi-cash', 'value': f'UGX {total_expenses:,.0f}', 'label': 'Total Expenses', 'type': 'danger'},
        {'icon': 'bi-graph-up', 'value': f'UGX {net_profit:,.0f}', 'label': 'Net Profit', 'type': 'info'},
    ]
    summary_totals = {
        'total_records': len(data),
        'total_amount': total_income + total_expenses,
        'total_paid': total_income,
        'outstanding': total_expenses,
        'recovery_rate': 'N/A',
        'par_30': 'N/A',
    }

    context = _get_base_context(request, {
        'report_title': 'Financial Performance Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'date_from': date_from,
        'date_to': date_to,
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 5. OFFICER PERFORMANCE REPORT (officer_report)
# ====================================================================
@login_required
def officer_report(request):
    """Officer Performance Report with PAR 1 & PAR 30 - Credit Officers only"""
    date_from = request.GET.get('date_from') or request.POST.get('date_from')
    date_to = request.GET.get('date_to') or request.POST.get('date_to')

    officers = User.objects.filter(
        is_active=True,
        groups__name='Credit Officer'
    ).order_by('first_name', 'last_name')

    data = []
    total_disbursed = Decimal('0')
    total_collected = Decimal('0')
    total_outstanding = Decimal('0')
    total_par_1 = Decimal('0')
    total_par_30 = Decimal('0')
    today = date.today()

    for officer in officers:
        loans = Loan.objects.filter(officer=officer)
        if date_from:
            loans = loans.filter(disbursed_date__gte=date_from)
        if date_to:
            loans = loans.filter(disbursed_date__lte=date_to)

        loan_count = loans.count()
        disbursed_amount = loans.aggregate(total=Sum('principal_amount'))['total'] or Decimal('0')

        collections = Transaction.objects.filter(
            type='repayment',
            loan__officer=officer,
            timestamp__date__gte=date_from if date_from else date.min,
            timestamp__date__lte=date_to if date_to else date.max
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        active_count = loans.filter(is_active=True).count()

        par_1_amount = Decimal('0')
        par_30_amount = Decimal('0')
        outstanding_principal = Decimal('0')

        for loan in loans:
            p_bal = loan.principal_balance or Decimal('0')
            outstanding_principal += p_bal

            unpaid_installments = loan.installments.filter(paid=False, due_date__lt=today)
            if unpaid_installments.exists():
                oldest_due = unpaid_installments.earliest('due_date').due_date
                days_overdue = (today - oldest_due).days

                if days_overdue >= 1:
                    par_1_amount += p_bal
                if days_overdue >= 30:
                    par_30_amount += p_bal

        par_1_percent = round((par_1_amount / outstanding_principal * 100), 2) if outstanding_principal > 0 else 0
        par_30_percent = round((par_30_amount / outstanding_principal * 100), 2) if outstanding_principal > 0 else 0
        performance = round((collections / disbursed_amount * 100), 2) if disbursed_amount > 0 else 0

        data.append({
            'officer': officer.get_full_name() or officer.username,
            'loan_count': loan_count,
            'active_count': active_count,
            'disbursed': disbursed_amount,
            'collected': collections,
            'outstanding': outstanding_principal,
            'par_1_amount': par_1_amount,
            'par_1_percent': par_1_percent,
            'par_30_amount': par_30_amount,
            'par_30_percent': par_30_percent,
            'performance': performance,
        })

        total_disbursed += disbursed_amount
        total_collected += collections
        total_outstanding += outstanding_principal
        total_par_1 += par_1_amount
        total_par_30 += par_30_amount

    total_par_1_percent = round((total_par_1 / total_outstanding * 100), 2) if total_outstanding > 0 else 0
    total_par_30_percent = round((total_par_30 / total_outstanding * 100), 2) if total_outstanding > 0 else 0
    total_performance = round((total_collected / total_disbursed * 100), 2) if total_disbursed > 0 else 0

    columns = [
        {'key': 'officer', 'label': 'Officer', 'align': 'left'},
        {'key': 'loan_count', 'label': 'Total Loans', 'align': 'center'},
        {'key': 'active_count', 'label': 'Active', 'align': 'center'},
        {'key': 'disbursed', 'label': 'Disbursed (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'collected', 'label': 'Collected (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'outstanding', 'label': 'Outstanding (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'par_1_amount', 'label': 'PAR 1 (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'par_1_percent', 'label': 'PAR 1 %', 'align': 'right', 'total': True},
        {'key': 'par_30_amount', 'label': 'PAR 30 (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'par_30_percent', 'label': 'PAR 30 %', 'align': 'right', 'total': True},
        {'key': 'performance', 'label': 'Performance %', 'align': 'right', 'total': True},
    ]
    totals = {
        'disbursed': total_disbursed,
        'collected': total_collected,
        'outstanding': total_outstanding,
        'par_1_amount': total_par_1,
        'par_1_percent': total_par_1_percent,
        'par_30_amount': total_par_30,
        'par_30_percent': total_par_30_percent,
        'performance': total_performance,
    }
    kpi_cards = [
        {'icon': 'bi-person-badge', 'value': f'{len(data)}', 'label': 'Total Officers', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_disbursed:,.0f}', 'label': 'Total Disbursed', 'type': 'success'},
        {'icon': 'bi-cash-stack', 'value': f'UGX {total_collected:,.0f}', 'label': 'Total Collected', 'type': 'info'},
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {total_par_30:,.0f}', 'label': 'Total PAR 30', 'type': 'danger'},
    ]
    summary_totals = {
        'total_records': len(data),
        'total_amount': total_disbursed,
        'total_paid': total_collected,
        'outstanding': total_outstanding,
        'recovery_rate': f'{total_performance:.2f}',
        'par_30': f'{total_par_30_percent:.2f}',
    }

    context = _get_base_context(request, {
        'report_title': 'Officer Performance Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'date_from': date_from or 'All',
        'date_to': date_to or 'All',
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 6. ACCOUNTING REPORT (accounting_report)
# ====================================================================
@login_required
def accounting_report(request):
    """General Ledger Summary Report"""
    date_from = request.GET.get('date_from') or request.POST.get('date_from')
    date_to = request.GET.get('date_to') or request.POST.get('date_to')
    account = request.GET.get('account') or request.POST.get('account')

    qs = GeneralLedger.objects.select_related('account')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if account:
        qs = qs.filter(account_id=account)

    data = []
    total_debit = Decimal('0')
    total_credit = Decimal('0')

    for entry in qs.order_by('account__code', 'date'):
        data.append({
            'account_code': entry.account.code,
            'account_name': entry.account.name,
            'account_type': entry.account.get_account_type_display(),
            'debit': entry.debit,
            'credit': entry.credit,
            'balance': entry.balance,
        })
        total_debit += entry.debit
        total_credit += entry.credit

    columns = [
        {'key': 'account_code', 'label': 'Account Code', 'align': 'left'},
        {'key': 'account_name', 'label': 'Account Name', 'align': 'left'},
        {'key': 'account_type', 'label': 'Type', 'align': 'left'},
        {'key': 'debit', 'label': 'Debit (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'credit', 'label': 'Credit (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'balance', 'label': 'Balance (UGX)', 'type': 'currency', 'align': 'right', 'prefix': 'UGX '},
    ]
    totals = {'debit': total_debit, 'credit': total_credit, 'balance': total_debit - total_credit}
    kpi_cards = [
        {'icon': 'bi-journal-text', 'value': f'{len(data)}', 'label': 'Total Accounts', 'type': 'info'},
        {'icon': 'bi-arrow-down', 'value': f'UGX {total_debit:,.0f}', 'label': 'Total Debit', 'type': 'danger'},
        {'icon': 'bi-arrow-up', 'value': f'UGX {total_credit:,.0f}', 'label': 'Total Credit', 'type': 'success'},
    ]
    summary_totals = {
        'total_records': len(data),
        'total_amount': total_debit + total_credit,
        'total_paid': total_credit,
        'outstanding': total_debit,
        'recovery_rate': 'N/A',
        'par_30': 'N/A',
    }

    context = _get_base_context(request, {
        'report_title': 'Accounting Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'date_from': date_from,
        'date_to': date_to,
        'selected_account': account,
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 7. AUDIT REPORT (audit_report)
# ====================================================================
@login_required
def audit_report(request):
    """Audit Trail Report (last 100 transactions)"""
    date_from = request.GET.get('date_from') or request.POST.get('date_from')
    date_to = request.GET.get('date_to') or request.POST.get('date_to')
    # We'll limit to last 100 unless date filters are applied

    qs = Transaction.objects.select_related('created_by').order_by('-timestamp')
    if date_from:
        qs = qs.filter(timestamp__date__gte=date_from)
    if date_to:
        qs = qs.filter(timestamp__date__lte=date_to)
    if not date_from and not date_to:
        qs = qs[:100]

    data = []
    for tx in qs:
        data.append({
            'timestamp': tx.timestamp,
            'user': tx.created_by.get_full_name() if tx.created_by else 'System',
            'type': tx.get_type_display(),
            'amount': tx.amount,
            'reference': tx.reference or '-',
            'is_reversed': 'Yes' if tx.is_reversed else 'No',
            'status': 'Reversed' if tx.is_reversed else 'Active',
        })

    columns = [
        {'key': 'timestamp', 'label': 'Date & Time', 'align': 'center', 'type': 'date'},
        {'key': 'user', 'label': 'User', 'align': 'left'},
        {'key': 'type', 'label': 'Transaction Type', 'align': 'left'},
        {'key': 'amount', 'label': 'Amount (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'reference', 'label': 'Reference', 'align': 'left'},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
    ]
    totals = {'amount': sum(item['amount'] for item in data) if data else Decimal('0')}
    kpi_cards = [
        {'icon': 'bi-clock-history', 'value': f'{len(data)}', 'label': 'Total Transactions', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {totals["amount"]:,.0f}', 'label': 'Total Volume', 'type': 'success'},
        {'icon': 'bi-person', 'value': f'{len(set(tx.created_by_id for tx in qs if tx.created_by))}', 'label': 'Active Users', 'type': 'secondary'},
    ]
    summary_totals = {
        'total_records': len(data),
        'total_amount': totals['amount'],
        'total_paid': 'N/A',
        'outstanding': 'N/A',
        'recovery_rate': 'N/A',
        'par_30': 'N/A',
    }

    context = _get_base_context(request, {
        'report_title': 'Audit Trail Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'date_from': date_from,
        'date_to': date_to,
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 8. INVENTORY REPORT (inventory_report)
# ====================================================================
@login_required
def inventory_report(request):
    """Inventory Report - Products and Stock Levels (from hardware app)"""
    try:
        from hardware.models import Product, Category
        products = Product.objects.select_related('category').all()
        has_inventory = True
    except ImportError:
        products = []
        has_inventory = False

    data = []
    total_value = Decimal('0')
    total_stock = 0

    if has_inventory:
        for product in products:
            stock_value = (product.current_stock or 0) * (product.cost_price or 0)
            data.append({
                'product_code': product.product_code,
                'product_name': product.name,
                'category': product.category.name if product.category else 'Uncategorized',
                'stock': product.current_stock or 0,
                'cost_price': product.cost_price or 0,
                'selling_price': product.selling_price or 0,
                'stock_value': stock_value,
                'status': 'Low Stock' if (product.current_stock or 0) <= (product.reorder_level or 5) else 'Healthy',
            })
            total_value += stock_value
            total_stock += (product.current_stock or 0)
    else:
        data = [{'message': 'Inventory module not installed'}]

    columns = [
        {'key': 'product_code', 'label': 'Product Code', 'align': 'left'},
        {'key': 'product_name', 'label': 'Product Name', 'align': 'left'},
        {'key': 'category', 'label': 'Category', 'align': 'left'},
        {'key': 'stock', 'label': 'Stock', 'align': 'center'},
        {'key': 'cost_price', 'label': 'Cost (UGX)', 'type': 'currency', 'align': 'right', 'prefix': 'UGX '},
        {'key': 'selling_price', 'label': 'Sell (UGX)', 'type': 'currency', 'align': 'right', 'prefix': 'UGX '},
        {'key': 'stock_value', 'label': 'Value (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
    ]
    totals = {'stock_value': total_value}
    kpi_cards = [
        {'icon': 'bi-boxes', 'value': f'{len(data)}', 'label': 'Total Products', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_value:,.0f}', 'label': 'Inventory Value', 'type': 'success'},
        {'icon': 'bi-box', 'value': f'{total_stock:,}', 'label': 'Total Stock Units', 'type': 'secondary'},
    ]
    summary_totals = {
        'total_records': len(data),
        'total_amount': total_value,
        'total_paid': 'N/A',
        'outstanding': 'N/A',
        'recovery_rate': 'N/A',
        'par_30': 'N/A',
    }

    context = _get_base_context(request, {
        'report_title': 'Inventory Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': has_inventory and bool(data),
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 9. INTEREST INCOME REPORT (interest_report)
# ====================================================================
@login_required
def interest_report(request):
    """Interest Income Report"""
    if request.method == "POST":
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        officer_id = request.POST.get('officer')
        product = request.POST.get('product')
    else:
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        officer_id = request.GET.get('officer')
        product = request.GET.get('product')

    if not date_from:
        date_from = (date.today() - timedelta(days=365)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = date.today().strftime('%Y-%m-%d')

    loans = Loan.objects.filter(
        disbursed_date__gte=date_from,
        disbursed_date__lte=date_to,
        status__in=['approved', 'active', 'closed']
    ).select_related('member', 'officer')

    if officer_id:
        loans = loans.filter(officer_id=officer_id)
    if product:
        loans = loans.filter(product_type=product)

    data = []
    total_principal = Decimal('0')
    total_interest_charged = Decimal('0')
    total_interest_paid = Decimal('0')
    total_interest_balance = Decimal('0')

    for loan in loans:
        interest_charged = loan.installments.aggregate(total=Sum('interest_portion'))['total'] or Decimal('0')
        interest_paid = loan.installments.aggregate(total=Sum('interest_paid'))['total'] or Decimal('0')
        interest_balance = loan.interest_balance or Decimal('0')

        data.append({
            'member': f"{loan.member.first_name} {loan.member.last_name}",
            'loan_ref': loan.loan_reference or f"LN-{loan.id}",
            'principal': loan.principal_amount,
            'interest_charged': interest_charged,
            'interest_paid': interest_paid,
            'interest_balance': interest_balance,
            'status': loan.get_status_display(),
            'disbursed_date': loan.disbursed_date or loan.start_date,
        })
        total_principal += loan.principal_amount
        total_interest_charged += interest_charged
        total_interest_paid += interest_paid
        total_interest_balance += interest_balance

    columns = [
        {'key': 'member', 'label': 'Member', 'align': 'left'},
        {'key': 'loan_ref', 'label': 'Loan Reference', 'align': 'left'},
        {'key': 'principal', 'label': 'Principal (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_charged', 'label': 'Interest Charged (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_paid', 'label': 'Interest Paid (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_balance', 'label': 'Interest Balance (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'status', 'label': 'Status', 'align': 'center', 'type': 'status'},
        {'key': 'disbursed_date', 'label': 'Disbursed', 'align': 'center', 'type': 'date'},
    ]
    totals = {
        'principal': total_principal,
        'interest_charged': total_interest_charged,
        'interest_paid': total_interest_paid,
        'interest_balance': total_interest_balance,
    }
    kpi_cards = [
        {'icon': 'bi-percent', 'value': f'UGX {total_interest_charged:,.0f}', 'label': 'Total Interest Charged', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_interest_paid:,.0f}', 'label': 'Interest Paid', 'type': 'success'},
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {total_interest_balance:,.0f}', 'label': 'Interest Outstanding', 'type': 'warning'},
    ]
    summary_totals = {
        'total_records': len(data),
        'total_amount': total_principal,
        'total_paid': total_interest_paid,
        'outstanding': total_interest_balance,
        'recovery_rate': (total_interest_paid / total_interest_charged * 100) if total_interest_charged > 0 else 0,
        'par_30': 'N/A',
    }

    selected_officer_display = None
    if officer_id:
        try:
            off = User.objects.get(id=officer_id)
            selected_officer_display = off.get_full_name() or off.username
        except User.DoesNotExist:
            pass

    context = _get_base_context(request, {
        'report_title': 'Interest Income Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'date_from': date_from,
        'date_to': date_to,
        'selected_officer': officer_id,
        'officer_name': selected_officer_display,
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 10. LOAN PORTFOLIO REPORTS (loan_portfolio_reports)
# ====================================================================
@login_required
def loan_portfolio_reports(request):
    """Comprehensive Loan Portfolio Report with PAR and arrears"""
    if request.method == "POST":
        start_date = request.POST.get('date_from')
        end_date = request.POST.get('date_to')
        officer_id = request.POST.get('officer')
        status = request.POST.get('status')
    else:
        start_date = request.GET.get('date_from')
        end_date = request.GET.get('date_to')
        officer_id = request.GET.get('officer')
        status = request.GET.get('status')

    loans = Loan.objects.select_related('member', 'officer').filter(
        status__in=['approved', 'active', 'closed', 'arrears']
    ).order_by('-disbursed_date', '-start_date')

    if start_date:
        loans = loans.filter(disbursed_date__gte=start_date)
    if end_date:
        loans = loans.filter(disbursed_date__lte=end_date)
    if officer_id:
        loans = loans.filter(officer_id=officer_id)
    if status:
        loans = loans.filter(status=status)

    today = date.today()
    report_data = []

    for loan in loans:
        p_bal = Decimal(str(loan.principal_balance or 0))
        i_bal = Decimal(str(loan.interest_balance or 0))

        overdue = loan.installments.filter(paid=False, due_date__lt=today)
        principal_in_arrears = overdue.aggregate(
            total=Coalesce(Sum('principal_portion'), Decimal('0'))
        )['total']

        total_due_today = loan.installments.filter(
            paid=False, due_date__lte=today
        ).aggregate(
            total=Coalesce(Sum(F('principal_portion') + F('interest_portion')), Decimal('0'))
        )['total']

        penalty_due = overdue.aggregate(
            total=Coalesce(Sum('penalty_amount'), Decimal('0'))
        )['total']

        report_data.append({
            'borrower': f"{loan.member.first_name} {loan.member.last_name}",
            'officer': loan.officer.get_full_name() if loan.officer else 'System',
            'account_number': loan.member.member_number,
            'contact': loan.member.phone_number,
            'loan_disbursed': Decimal(str(loan.principal_amount or 0)),
            'disbursement_date': loan.disbursed_date or loan.start_date,
            'principal_balance': p_bal,
            'interest_balance': i_bal,
            'principal_in_arrears': principal_in_arrears,
            'total_dues': total_due_today + penalty_due,
            'par': p_bal if principal_in_arrears > 0 else Decimal('0'),
        })

    total_disbursed = sum(item['loan_disbursed'] for item in report_data)
    total_outstanding = sum(item['principal_balance'] + item['interest_balance'] for item in report_data)
    total_par = sum(item['par'] for item in report_data)

    columns = [
        {'key': 'borrower', 'label': 'Borrower', 'align': 'left'},
        {'key': 'account_number', 'label': 'Account No', 'align': 'left'},
        {'key': 'loan_disbursed', 'label': 'Disbursed (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_balance', 'label': 'Principal (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_balance', 'label': 'Interest (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_in_arrears', 'label': 'Arrears (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'total_dues', 'label': 'Total Due (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'par', 'label': 'PAR (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'officer', 'label': 'Officer', 'align': 'left'},
        {'key': 'disbursement_date', 'label': 'Disbursed Date', 'align': 'center', 'type': 'date'},
    ]
    totals = {
        'loan_disbursed': total_disbursed,
        'principal_balance': total_outstanding,
        'interest_balance': sum(item['interest_balance'] for item in report_data),
        'principal_in_arrears': sum(item['principal_in_arrears'] for item in report_data),
        'total_dues': sum(item['total_dues'] for item in report_data),
        'par': total_par,
    }
    kpi_cards = [
        {'icon': 'bi-bank', 'value': f'UGX {total_disbursed:,.0f}', 'label': 'Total Disbursed', 'type': 'success'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {total_outstanding:,.0f}', 'label': 'Total Outstanding', 'type': 'info'},
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {total_par:,.0f}', 'label': 'Portfolio at Risk', 'type': 'danger'},
    ]
    summary_totals = {
        'total_records': len(report_data),
        'total_amount': total_disbursed,
        'total_paid': total_disbursed - total_outstanding,
        'outstanding': total_outstanding,
        'recovery_rate': ((total_disbursed - total_outstanding) / total_disbursed * 100) if total_disbursed > 0 else 0,
        'par_30': (total_par / total_disbursed * 100) if total_disbursed > 0 else 0,
    }

    selected_officer_display = None
    if officer_id:
        try:
            off = User.objects.get(id=officer_id)
            selected_officer_display = off.get_full_name() or off.username
        except User.DoesNotExist:
            pass

    context = _get_base_context(request, {
        'report_title': 'Loan Portfolio Report',
        'columns': columns,
        'data': report_data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(report_data),
        'date_from': start_date or 'All',
        'date_to': end_date or 'All',
        'selected_officer': officer_id,
        'officer_name': selected_officer_display,
        'selected_status': status,
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 11. PORTFOLIO STATUS REPORT (portfolio_status_report)
# ====================================================================
@login_required
def portfolio_status_report(request):
    """Comprehensive Portfolio Status Report with aging classification"""
    today = timezone.now().date()
    loans = Loan.objects.select_related('member', 'officer').filter(
        status__in=['approved', 'active', 'closed']
    ).order_by('member__member_number')

    report_data = []
    for loan in loans:
        paid_stats = loan.installments.filter(paid=True).aggregate(
            p_paid=Coalesce(Sum('principal_portion'), Decimal('0.00')),
            i_paid=Coalesce(Sum('interest_portion'), Decimal('0.00')),
            penalty_paid=Coalesce(Sum('penalty_amount'), Decimal('0.00')),
        )

        arrears_stats = loan.installments.filter(
            paid=False,
            due_date__lt=today
        ).aggregate(
            p_due=Coalesce(Sum('principal_portion'), Decimal('0.00')),
            i_due=Coalesce(Sum('interest_portion'), Decimal('0.00')),
            pen_due=Coalesce(Sum('penalty_amount'), Decimal('0.00')),
        )

        oldest_unpaid = loan.installments.filter(
            paid=False,
            due_date__lt=today
        ).order_by('due_date').first()

        classification = "Performing"
        if oldest_unpaid:
            days_past_due = (today - oldest_unpaid.due_date).days
            if days_past_due > 180:
                classification = "Loss"
            elif days_past_due > 90:
                classification = "Doubtful"
            elif days_past_due > 30:
                classification = "Substandard"
            else:
                classification = "Watch"

        report_data.append({
            'member_no': loan.member.member_number or str(loan.member.id),
            'name': f"{loan.member.first_name} {loan.member.last_name}",
            'loan_no': loan.loan_reference or loan.id,
            'disbursed_amount': Decimal(str(loan.principal_amount or 0)),
            'disbursed_date': loan.disbursed_date or loan.start_date,
            'principal_paid': paid_stats['p_paid'],
            'interest_paid': paid_stats['i_paid'],
            'penalty_paid': paid_stats['penalty_paid'],
            'principal_due': arrears_stats['p_due'],
            'interest_due': arrears_stats['i_due'],
            'penalty_due': arrears_stats['pen_due'],
            'total_due': arrears_stats['p_due'] + arrears_stats['i_due'] + arrears_stats['pen_due'],
            'principal_balance': Decimal(str(loan.principal_balance or 0)),
            'interest_balance': Decimal(str(loan.interest_balance or 0)),
            'classification': classification,
            'sector': getattr(loan.member, 'economic_sector', 'N/A'),
        })

    grand_total_disbursed = sum(item['disbursed_amount'] for item in report_data)
    grand_total_prin_paid = sum(item['principal_paid'] for item in report_data)
    grand_total_int_paid = sum(item['interest_paid'] for item in report_data)
    grand_total_prin_due = sum(item['principal_due'] for item in report_data)
    grand_total_int_due = sum(item['interest_due'] for item in report_data)
    grand_total_due = sum(item['total_due'] for item in report_data)
    grand_total_prin_bal = sum(item['principal_balance'] for item in report_data)
    grand_total_int_bal = sum(item['interest_balance'] for item in report_data)
    grand_total_exposure = grand_total_prin_bal + grand_total_int_bal

    columns = [
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'name', 'label': 'Member Name', 'align': 'left'},
        {'key': 'loan_no', 'label': 'Loan Ref', 'align': 'left'},
        {'key': 'disbursed_amount', 'label': 'Disbursed (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_paid', 'label': 'Principal Paid', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_paid', 'label': 'Interest Paid', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_due', 'label': 'Principal Due', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_due', 'label': 'Interest Due', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'total_due', 'label': 'Total Due', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_balance', 'label': 'Principal Balance', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_balance', 'label': 'Interest Balance', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'classification', 'label': 'Classification', 'align': 'center'},
        {'key': 'sector', 'label': 'Sector', 'align': 'left'},
    ]
    totals = {
        'disbursed_amount': grand_total_disbursed,
        'principal_paid': grand_total_prin_paid,
        'interest_paid': grand_total_int_paid,
        'principal_due': grand_total_prin_due,
        'interest_due': grand_total_int_due,
        'total_due': grand_total_due,
        'principal_balance': grand_total_prin_bal,
        'interest_balance': grand_total_int_bal,
    }
    kpi_cards = [
        {'icon': 'bi-people', 'value': f'{len(report_data):,}', 'label': 'Total Loans', 'type': 'info'},
        {'icon': 'bi-currency-dollar', 'value': f'UGX {grand_total_exposure:,.0f}', 'label': 'Total Exposure', 'type': 'success'},
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {grand_total_due:,.0f}', 'label': 'Total Arrears', 'type': 'danger'},
    ]
    summary_totals = {
        'total_records': len(report_data),
        'total_amount': grand_total_disbursed,
        'total_paid': grand_total_prin_paid + grand_total_int_paid,
        'outstanding': grand_total_prin_bal + grand_total_int_bal,
        'recovery_rate': ((grand_total_prin_paid + grand_total_int_paid) / grand_total_disbursed * 100) if grand_total_disbursed > 0 else 0,
        'par_30': (grand_total_due / grand_total_disbursed * 100) if grand_total_disbursed > 0 else 0,
    }

    context = _get_base_context(request, {
        'report_title': 'Portfolio Status Report',
        'columns': columns,
        'data': report_data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(report_data),
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)


# ====================================================================
# 12. ARREARS REPORT (arrears_report)
# ====================================================================
from decimal import Decimal
from django.db.models import Q, Sum, F, Value, DecimalField, Func
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import date, datetime
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import FileResponse

from finance.models import Installment, Loan
from finance.penalties import calculate_penalty  # ensure this import exists


from decimal import Decimal
from django.db.models import Q, Sum, F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.utils import timezone
from datetime import date, datetime
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import FileResponse

from finance.models import Installment, Loan
from finance.penalties import calculate_penalty          # needed for loan_detail logic


from decimal import Decimal
from django.db.models import Q, Sum
from django.utils import timezone
from datetime import date, datetime
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import FileResponse

from finance.models import Installment, Loan
from finance.penalties import calculate_penalty          # needed for loan_detail logic


@login_required
def arrears_report(request):
    """Arrears & Delinquency Report with aging buckets – true penalty calculation"""
    from datetime import datetime

    if request.method == "POST":
        date_at_str = request.POST.get('date_at')
        search_query = request.POST.get('search_query')
    else:
        date_at_str = request.GET.get('date_at')
        search_query = request.GET.get('search_query')

    today = date.today()
    if date_at_str:
        try:
            target_date = datetime.strptime(date_at_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = today
    else:
        target_date = today

    # Fetch overdue installments (unpaid and due before target_date)
    overdue_installments = Installment.objects.filter(
        paid=False,
        due_date__lt=target_date,
        loan__status__in=['approved', 'active', 'arrears']
    ).select_related('loan', 'loan__member', 'loan__officer')

    if search_query:
        overdue_installments = overdue_installments.filter(
            Q(loan__member__first_name__icontains=search_query) |
            Q(loan__member__last_name__icontains=search_query) |
            Q(loan__member__member_number__icontains=search_query) |
            Q(loan__loan_reference__icontains=search_query)
        )

    # Build data in Python – using the same penalty calculation as loan_detail
    data = []
    total_arrears = Decimal('0')
    aging_buckets = {
        '1-30_days': Decimal('0'),
        '31-60_days': Decimal('0'),
        '61-90_days': Decimal('0'),
        '91-180_days': Decimal('0'),
        '180_plus': Decimal('0'),
    }

    for inst in overdue_installments:
        days = (target_date - inst.due_date).days

        # === Penalty calculation (mirroring loan_detail) ===
        # Calculated penalty from the rule (fresh, using today's date)
        calc_penalty = calculate_penalty(inst) or Decimal('0.00')
        # Manual penalties (not waived) for this installment
        manual_total = inst.manual_penalties.filter(is_waived=False).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        total_penalty = calc_penalty + manual_total

        # Due amounts
        principal_due = inst.principal_balance          # property: principal_portion - principal_paid
        interest_due = inst.interest_balance            # property: interest_portion - interest_paid
        penalty_due = total_penalty - inst.penalty_paid  # subtract already paid penalty

        total_due = principal_due + interest_due + penalty_due

        data.append({
            'member_no': inst.loan.member.member_number or str(inst.loan.member.id),
            'member_name': f"{inst.loan.member.first_name} {inst.loan.member.last_name}",
            'loan_ref': inst.loan.loan_reference or f"LN-{inst.loan.id}",
            'phone': inst.loan.member.phone_number,
            'due_date': inst.due_date,
            'days_overdue': days,
            'principal_due': principal_due,
            'interest_due': interest_due,
            'penalty_due': penalty_due,
            'total_due': total_due,
            'disbursed_amount': Decimal(str(inst.loan.principal_amount or 0)),
            'officer': inst.loan.officer.get_full_name() if inst.loan.officer else 'System',
        })

        total_arrears += total_due

        # Aging bucket
        if 1 <= days <= 30:
            aging_buckets['1-30_days'] += total_due
        elif 31 <= days <= 60:
            aging_buckets['31-60_days'] += total_due
        elif 61 <= days <= 90:
            aging_buckets['61-90_days'] += total_due
        elif 91 <= days <= 180:
            aging_buckets['91-180_days'] += total_due
        else:
            aging_buckets['180_plus'] += total_due

    # Columns definition
    columns = [
        {'key': 'member_no', 'label': 'Member No', 'align': 'left'},
        {'key': 'member_name', 'label': 'Member', 'align': 'left'},
        {'key': 'loan_ref', 'label': 'Loan Ref', 'align': 'left'},
        {'key': 'phone', 'label': 'Phone', 'align': 'left'},
        {'key': 'due_date', 'label': 'Due Date', 'align': 'center', 'type': 'date'},
        {'key': 'days_overdue', 'label': 'Days Overdue', 'align': 'center'},
        {'key': 'disbursed_amount', 'label': 'Amount Disbursed (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'principal_due', 'label': 'Principal Due', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'interest_due', 'label': 'Interest Due', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'penalty_due', 'label': 'Penalty Due', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'total_due', 'label': 'Total Due', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'officer', 'label': 'Officer', 'align': 'left'},
    ]

    totals = {
        'disbursed_amount': sum(item['disbursed_amount'] for item in data),
        'principal_due': sum(item['principal_due'] for item in data),
        'interest_due': sum(item['interest_due'] for item in data),
        'penalty_due': sum(item['penalty_due'] for item in data),
        'total_due': total_arrears,
    }

    total_outstanding_loans = Loan.objects.filter(
        is_active=True,
        status__in=['approved', 'active', 'arrears']
    ).count()

    kpi_cards = [
        {'icon': 'bi-exclamation-triangle', 'value': f'UGX {total_arrears:,.0f}', 'label': 'Total Arrears', 'type': 'danger'},
        {'icon': 'bi-clock-history', 'value': f'{len(data):,}', 'label': 'Overdue Installments', 'type': 'warning'},
        {'icon': 'bi-percent', 'value': f'{(total_arrears / (total_outstanding_loans + 1)):.1f}%', 'label': 'Arrears Rate', 'type': 'info'},
    ]

    aging_summary = [
        {'bucket': '1-30 Days', 'amount': aging_buckets['1-30_days']},
        {'bucket': '31-60 Days', 'amount': aging_buckets['31-60_days']},
        {'bucket': '61-90 Days', 'amount': aging_buckets['61-90_days']},
        {'bucket': '91-180 Days', 'amount': aging_buckets['91-180_days']},
        {'bucket': '180+ Days', 'amount': aging_buckets['180_plus']},
    ]

    summary_totals = {
        'total_records': len(data),
        'total_amount': total_arrears,
        'total_paid': 'N/A',
        'outstanding': total_arrears,
        'recovery_rate': 'N/A',
        'par_30': f'{(total_arrears / (total_outstanding_loans + 1)):.1f}',
    }

    context = _get_base_context(request, {
        'report_title': 'Arrears & Delinquency Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'aging_summary': aging_summary,
        'has_data': bool(data),
        'date_from': target_date.strftime('%Y-%m-%d'),
        'date_to': target_date.strftime('%Y-%m-%d'),
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)

# ====================================================================
# 13. GENERAL LEDGER REPORT (general_ledger_report)
# ====================================================================
@login_required
def general_ledger_report(request):
    """Professional General Ledger report with date, account, and account type filters."""
    date_from = request.GET.get('date_from') or request.POST.get('date_from')
    date_to = request.GET.get('date_to') or request.POST.get('date_to')
    account_id = request.GET.get('account') or request.POST.get('account')
    account_type = request.GET.get('account_type') or request.POST.get('account_type')
    status = request.GET.get('status') or request.POST.get('status')  # not used, but kept for consistency

    qs = GeneralLedger.objects.select_related('account')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if account_id:
        qs = qs.filter(account_id=account_id)
    if account_type:
        qs = qs.filter(account__account_type=account_type)

    data = []
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')

    for entry in qs.order_by('date', 'id'):
        data.append({
            'date': entry.date,
            'account_code': entry.account.code,
            'account_name': entry.account.name,
            'description': entry.description,
            'reference': entry.reference or '-',
            'debit': entry.debit,
            'credit': entry.credit,
            'balance': entry.balance,
        })
        total_debit += entry.debit
        total_credit += entry.credit

    columns = [
        {'key': 'date', 'label': 'Date', 'type': 'date', 'align': 'center'},
        {'key': 'account_code', 'label': 'Account Code', 'align': 'left'},
        {'key': 'account_name', 'label': 'Account Name', 'align': 'left'},
        {'key': 'description', 'label': 'Description', 'align': 'left'},
        {'key': 'reference', 'label': 'Reference', 'align': 'left'},
        {'key': 'debit', 'label': 'Debit (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'credit', 'label': 'Credit (UGX)', 'type': 'currency', 'align': 'right', 'total': True, 'prefix': 'UGX '},
        {'key': 'balance', 'label': 'Balance (UGX)', 'type': 'currency', 'align': 'right', 'prefix': 'UGX '},
    ]
    totals = {'debit': total_debit, 'credit': total_credit}

    kpi_cards = [
        {'label': 'Total Entries', 'value': qs.count(), 'icon': 'bi-list-ul', 'type': 'info'},
        {'label': 'Total Debit', 'value': f"UGX {total_debit:,.0f}", 'icon': 'bi-arrow-down', 'type': 'danger'},
        {'label': 'Total Credit', 'value': f"UGX {total_credit:,.0f}", 'icon': 'bi-arrow-up', 'type': 'success'},
        {'label': 'Net Movement', 'value': f"UGX {abs(total_debit - total_credit):,.0f}", 'icon': 'bi-arrows-vertical', 'type': 'warning'},
    ]

    summary_totals = {
        'total_records': qs.count(),
        'total_amount': total_debit + total_credit,
        'total_paid': total_credit,
        'outstanding': total_debit,
        'recovery_rate': 'N/A',
        'par_30': 'N/A',
    }

    selected_account_display = None
    if account_id:
        try:
            acc = ChartOfAccount.objects.get(id=account_id)
            selected_account_display = f"{acc.code} – {acc.name}"
        except ChartOfAccount.DoesNotExist:
            pass

    selected_account_type_display = dict(ChartOfAccount.ACCOUNT_TYPES).get(account_type)

    context = _get_base_context(request, {
        'report_title': 'General Ledger Report',
        'columns': columns,
        'data': data,
        'totals': totals,
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'has_data': bool(data),
        'date_from': date_from,
        'date_to': date_to,
        'selected_account': account_id,
        'selected_account_type': account_type,
        'selected_status': status,
        'selected_account_display': selected_account_display,
        'selected_account_type_display': selected_account_type_display,
    })

    if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
        excel_file = generate_excel_report(
            columns=context['columns'],
            data=context['data'],
            report_title=context['report_title'],
            company_name=context['company']['name'],
            totals=context['totals']
        )
        filename = f"General_Ledger_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
        return FileResponse(excel_file, as_attachment=True, filename=filename)

    return render(request, 'finance/reports/base_report.html', context)



# finance/views.py (or inside your existing views)
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction as db_transaction
from django.utils import timezone
from decimal import Decimal
from .models import Transaction, GeneralLedger, ChartOfAccount, AccountingEngine
from .forms import JournalEntryForm

def journal_entry(request):
    if request.method == 'POST':
        form = JournalEntryForm(request.POST)
        if form.is_valid():
            debit_acc = form.cleaned_data['debit_account']
            credit_acc = form.cleaned_data['credit_account']
            amount = form.cleaned_data['amount']
            description = form.cleaned_data['description']
            reference = form.cleaned_data['reference'] or f"JRN-{timezone.now().strftime('%Y%m%d%H%M%S')}"
            entry_date = form.cleaned_data['date'] or timezone.now().date()

            # Prevent self‑debit/credit
            if debit_acc == credit_acc:
                messages.error(request, "Debit and Credit accounts cannot be the same.")
                return render(request, 'finance/journal_entry.html', {'form': form})

            with db_transaction.atomic():
                # 1. Create a Transaction record (type='journal')
                tx = Transaction.objects.create(
                    member=None,          # Not tied to a member – or you could make it optional
                    loan=None,
                    amount=amount,
                    type='journal',
                    reference=reference,
                    timestamp=entry_date,
                    created_by=request.user if request.user.is_authenticated else None,
                )

                # 2. Post Debit entry (increase asset/expense, or decrease liability/income/equity)
                AccountingEngine.post_ledger_entry(
                    account_code=debit_acc.code,
                    description=f"{description} (Debit)",
                    reference=reference,
                    debit=amount,
                    credit=Decimal('0.00'),
                    transaction_obj=tx,
                    date_context=entry_date,
                )

                # 3. Post Credit entry (increase liability/income/equity, or decrease asset/expense)
                AccountingEngine.post_ledger_entry(
                    account_code=credit_acc.code,
                    description=f"{description} (Credit)",
                    reference=reference,
                    debit=Decimal('0.00'),
                    credit=amount,
                    transaction_obj=tx,
                    date_context=entry_date,
                )

            messages.success(request, f"Journal entry posted successfully! Ref: {reference}")
            return redirect('journal_entry')  # or to a list view

    else:
        form = JournalEntryForm()

    return render(request, 'finance/journal_entry.html', {'form': form})


# ====================================================================
# 14. GENERIC REPORT VIEW (report_view) – kept for backward compatibility
# ====================================================================
@login_required
def report_view(request):
    """Generic Loan Report view (same as loan_report, but with different defaults)"""
    # This is essentially a duplicate of loan_report; we can redirect or keep as is.
    # We'll keep it simple: just call loan_report with the same logic.
    return loan_report(request)


from django.http import JsonResponse
from django.http import JsonResponse
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal
from .models import ChartOfAccount, GeneralLedger

from django.http import JsonResponse
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal
from .models import ChartOfAccount, GeneralLedger

def account_balance_api(request, account_id):
    try:
        account = ChartOfAccount.objects.get(id=account_id)
    except ChartOfAccount.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Account not found'}, status=404)

    ledger_qs = GeneralLedger.objects.filter(account=account)
    total_debit = ledger_qs.aggregate(total=Coalesce(Sum('debit'), Value(Decimal('0.00'))))['total']
    total_credit = ledger_qs.aggregate(total=Coalesce(Sum('credit'), Value(Decimal('0.00'))))['total']

    if account.account_type in ['asset', 'expense']:
        balance = total_debit - total_credit
    else:  # liability, income, equity
        balance = total_credit - total_debit

    return JsonResponse({
        'success': True,
        'balance': float(balance),
        'account_type': account.get_account_type_display(),
    })

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import Http404

from django.shortcuts import render
from django.contrib.auth.decorators import login_required

# finance/views.py
from datetime import datetime   # add this import
# ... other imports

# finance/views.py
from datetime import datetime  # ensure import

@login_required
def view_receipt(request):
    receipt = request.session.get('deposit_receipt') or request.session.get('withdrawal_receipt')
    if not receipt or not receipt.get('show'):
        messages.warning(request, "No receipt to display.")
        return redirect('dashboard')

    # Clear session
    if 'deposit_receipt' in request.session:
        del request.session['deposit_receipt']
    if 'withdrawal_receipt' in request.session:
        del request.session['withdrawal_receipt']

    receipt_data = receipt['data']

    # Convert timestamp string to datetime object, or use current time as fallback
    timestamp_str = receipt_data.get('timestamp')
    if timestamp_str:
        try:
            receipt_data['timestamp_obj'] = datetime.fromisoformat(timestamp_str)
        except (ValueError, TypeError):
            receipt_data['timestamp_obj'] = None
    else:
        receipt_data['timestamp_obj'] = None

    context = {
        'receipt': receipt_data,
        'company': Company.get_company(),
    }
    return render(request, 'finance/receipt.html', context)




# finance/views.py
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import permission_required, login_required
from django.db import transaction as db_transaction
from django.utils import timezone
from decimal import Decimal
import requests
import logging

from .models import Loan, SMSConfig, SMSTransaction, Company
from decouple import config

logger = logging.getLogger(__name__)

# SMS cost for a reminder (set to 300 as requested)
SMS_REMINDER_COST = Decimal('300.00')

# finance/views.py – add at the top if missing
import requests
import logging
from decimal import Decimal
from django.utils import timezone
from django.db import transaction as db_transaction
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from decouple import config

from .models import Loan, SMSConfig, SMSTransaction, Company
from .models import Loan  # if not already imported

logger = logging.getLogger(__name__)

# Cost per SMS (as you already defined)
SMS_REMINDER_COST = Decimal('300.00')


@login_required
@permission_required('finance.can_send_sms', raise_exception=True)
def send_loan_reminder(request, loan_id):
    """
    Sends a payment reminder SMS to the member for a specific loan.
    Deducts 300 from SMS wallet.
    Message shows the actual overdue amount (principal + interest + penalty).
    """
    loan = get_object_or_404(Loan, id=loan_id)

    # 1. Check if loan is active (not closed)
    if loan.status in ['closed', 'rejected', 'defaulted']:
        messages.error(request, "This loan is no longer active.")
        return redirect('loan_details', loan_id=loan.id)

    # 2. Check SMS credits
    try:
        sms_config = SMSConfig.objects.get()  # assume singleton
    except SMSConfig.DoesNotExist:
        messages.error(request, "SMS service is not configured.")
        return redirect('loan_details', loan_id=loan.id)

    if sms_config.balance < SMS_REMINDER_COST:
        messages.error(
            request,
            f"Insufficient SMS credits. Required: UGX {SMS_REMINDER_COST:,.0f}, "
            f"Available: UGX {sms_config.balance:,.0f}"
        )
        return redirect('loan_details', loan_id=loan.id)

    # 3. Calculate the actual amount due (overdue installments)
    member = loan.member
    company = Company.get_company()
    today = timezone.now().date()

    # Get overdue installments (unpaid and due date <= today)
    overdue_inst = loan.installments.filter(paid=False, due_date__lte=today)

    if overdue_inst.exists():
        # Sum of balances (principal + interest + penalty) for overdue installments
        total_due = sum(inst.balance for inst in overdue_inst)
        # Earliest overdue due date
        next_due_date = overdue_inst.order_by('due_date').first().due_date
    else:
        # No overdue – fallback to next unpaid installment (if any)
        next_inst = loan.installments.filter(paid=False).order_by('due_date').first()
        if next_inst:
            total_due = next_inst.balance
            next_due_date = next_inst.due_date
        else:
            # No remaining installments – use loan balance as fallback
            total_due = loan.balance
            next_due_date = loan.start_date  # or today

    disbursed_date = loan.disbursed_date or loan.start_date

    # 4. Build the SMS message
    message = (
        f"Dear {member.first_name}, your loan of UGX {loan.principal_amount:,.0f} "
        f"disbursed on {disbursed_date.strftime('%d/%m/%Y')} has a payment due of "
        f"UGX {total_due:,.0f} by {next_due_date.strftime('%d/%m/%Y')}. "
        f"Please pay on time to avoid penalties. Thank you. - {company.name}"
    )

    # Truncate to 160 characters if needed (SpeedaMobile supports up to 160 per SMS)
    if len(message) > 160:
        message = message[:157] + "..."

    # 5. Format phone number
    raw_phone = str(member.phone_number).strip().replace(" ", "").replace("+", "")
    if raw_phone.startswith('0'):
        formatted_phone = '256' + raw_phone[1:]
    else:
        formatted_phone = raw_phone

    # 6. Send via SpeedaMobile API
    api_id = config('SPEEDA_API_ID')
    api_password = config('SPEEDA_API_PASSWORD')
    sender_id = config('SPEEDA_SENDER_ID', default='MACFinTech')

    url = "http://apidocs.speedamobile.com/api/SendSMS"
    payload = {
        "api_id": api_id,
        "api_password": api_password,
        "sms_type": "P",
        "encoding": "T",
        "sender_id": sender_id,
        "phonenumber": formatted_phone,
        "textmessage": message,
    }

    try:
        response = requests.post(url, json=payload, timeout=10)
        response.raise_for_status()
        result = response.json()

        if result.get("status") == "S":
            # Deduct credits and log transaction
            with db_transaction.atomic():
                config_obj = SMSConfig.objects.select_for_update().get(id=sms_config.id)
                config_obj.balance -= SMS_REMINDER_COST
                config_obj.save()

                SMSTransaction.objects.create(
                    amount=SMS_REMINDER_COST,
                    transaction_type='REMINDER',
                    description=f"Reminder sent for loan {loan.loan_reference} to {member.first_name}",
                    performed_by=request.user
                )

            messages.success(request, f"SMS reminder sent successfully to {member.first_name}.")
        else:
            error_msg = result.get('remarks', 'Unknown API error')
            messages.error(request, f"Failed to send SMS: {error_msg}")

    except requests.RequestException as e:
        logger.error(f"SMS API error for {formatted_phone}: {str(e)}")
        messages.error(request, "Could not connect to SMS gateway. Please try again later.")
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        messages.error(request, "An internal error occurred.")

    return redirect('loan_details', loan_id=loan.id)


@login_required
@permission_required('finance.can_send_sms', raise_exception=True)
def send_bulk_arrears_reminders(request):
    """
    Sends a bulk SMS reminder to all members with loans in arrears.
    Message shows the actual overdue amount for each member.
    """
    company = Company.get_company()
    today = timezone.now().date()

    # Get all active loans with status 'arrears'
    arrears_loans = Loan.objects.filter(
        status='arrears',
        is_active=True
    ).select_related('member')

    if not arrears_loans.exists():
        messages.info(request, "No loans are currently in arrears.")
        return redirect('loan_list')

    # Check SMS credits (needed for all messages)
    try:
        sms_config = SMSConfig.objects.get()
    except SMSConfig.DoesNotExist:
        messages.error(request, "SMS service is not configured.")
        return redirect('loan_list')

    # Count total required credits
    total_messages = arrears_loans.count()
    total_cost = total_messages * SMS_REMINDER_COST

    if sms_config.balance < total_cost:
        messages.error(
            request,
            f"Insufficient SMS credits. Need UGX {total_cost:,.0f}, "
            f"Available: UGX {sms_config.balance:,.0f}"
        )
        return redirect('loan_list')

    sent = 0
    failed = 0

    for loan in arrears_loans:
        member = loan.member
        # Calculate overdue amount for this loan
        overdue_inst = loan.installments.filter(paid=False, due_date__lte=today)
        if overdue_inst.exists():
            total_due = sum(inst.balance for inst in overdue_inst)
            next_due_date = overdue_inst.order_by('due_date').first().due_date
        else:
            # Should not happen, but fallback
            next_inst = loan.installments.filter(paid=False).order_by('due_date').first()
            if next_inst:
                total_due = next_inst.balance
                next_due_date = next_inst.due_date
            else:
                # No installments – skip this loan
                failed += 1
                continue

        disbursed_date = loan.disbursed_date or loan.start_date

        # Build personalised message
        message = (
            f"Dear {member.first_name}, your loan of UGX {loan.principal_amount:,.0f} "
            f"disbursed on {disbursed_date.strftime('%d/%m/%Y')} has a payment due of "
            f"UGX {total_due:,.0f} by {next_due_date.strftime('%d/%m/%Y')}. "
            f"Please pay on time to avoid penalties. Thank you. - {company.name}"
        )
        if len(message) > 160:
            message = message[:157] + "..."

        # Format phone number
        raw_phone = str(member.phone_number).strip().replace(" ", "").replace("+", "")
        if raw_phone.startswith('0'):
            formatted_phone = '256' + raw_phone[1:]
        else:
            formatted_phone = raw_phone

        # Call SpeedaMobile API
        api_id = config('SPEEDA_API_ID')
        api_password = config('SPEEDA_API_PASSWORD')
        sender_id = config('SPEEDA_SENDER_ID', default='MACFinTech')
        url = "http://apidocs.speedamobile.com/api/SendSMS"

        payload = {
            "api_id": api_id,
            "api_password": api_password,
            "sms_type": "P",
            "encoding": "T",
            "sender_id": sender_id,
            "phonenumber": formatted_phone,
            "textmessage": message,
        }

        try:
            response = requests.post(url, json=payload, timeout=10)
            response.raise_for_status()
            result = response.json()

            if result.get("status") == "S":
                # Deduct one SMS cost
                with db_transaction.atomic():
                    config_obj = SMSConfig.objects.select_for_update().get(id=sms_config.id)
                    config_obj.balance -= SMS_REMINDER_COST
                    config_obj.save()

                    SMSTransaction.objects.create(
                        amount=SMS_REMINDER_COST,
                        transaction_type='REMINDER',
                        description=f"Bulk reminder sent to {member.first_name} for loan {loan.loan_reference}",
                        performed_by=request.user
                    )
                sent += 1
            else:
                failed += 1
                logger.error(f"Bulk SMS failed for {formatted_phone}: {result.get('remarks')}")

        except Exception as e:
            failed += 1
            logger.error(f"Bulk SMS error for {formatted_phone}: {str(e)}")

    # Final message
    if sent > 0:
        messages.success(request, f"Bulk SMS reminders sent: {sent} successful, {failed} failed.")
    else:
        messages.error(request, "Bulk SMS reminders failed. Please try again later.")

    return redirect('loan_list')