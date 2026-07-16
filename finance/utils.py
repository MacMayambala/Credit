import requests
import logging
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from django.db import transaction
from django.contrib import messages

from decouple import config
from .models import Installment, SMSConfig, SMSTransaction

logger = logging.getLogger(__name__)

# Load from .env
SPEEDA_API_ID = config('SPEEDA_API_ID')
SPEEDA_API_PASSWORD = config('SPEEDA_API_PASSWORD')
SPEEDA_SENDER_ID = config('SPEEDA_SENDER_ID', default='MACFinTech')
SMS_COST_PER_MESSAGE = Decimal(config('SMS_COST_PER_MESSAGE', default=100))


def send_arrears_reminder_sms(request, loan):
    """
    Send arrears reminder SMS to a member using SpeedaMobile API.
    Costs SMS_COST_PER_MESSAGE per SMS.
    """
    member = loan.member
    school = getattr(request.user, 'school', None)

    if not school:
        return {"status": "F", "remarks": "School configuration missing."}

    # Get SMS Wallet
    try:
        sms_conf = SMSConfig.objects.get(school=school)
    except SMSConfig.DoesNotExist:
        return {"status": "F", "remarks": "SMS service not configured for this school."}

    if sms_conf.balance < SMS_COST_PER_MESSAGE:
        return {"status": "F", "remarks": "Insufficient SMS credits. Please top up."}

    # Format phone number (SpeedaMobile expects 256xxxxxxxxx)
    raw_phone = str(member.phone_number).strip().replace(" ", "").replace("+", "")
    if raw_phone.startswith('0'):
        formatted_phone = '256' + raw_phone[1:]
    else:
        formatted_phone = raw_phone

    # Professional SMS Message
    message = (
        f"Dear {member.first_name}, your loan repayment of UGX {loan.balance:,.0f} "
        f"is overdue. Please make payment to avoid penalties. "
        f"Thank you, MAC FinTech SACCO."
    )

    # Call SpeedaMobile API
    url = "http://apidocs.speedamobile.com/api/SendSMS"

    payload = {
        "api_id": SPEEDA_API_ID,
        "api_password": SPEEDA_API_PASSWORD,
        "sms_type": "P",
        "encoding": "T",
        "sender_id": SPEEDA_SENDER_ID,
        "phonenumber": formatted_phone,
        "textmessage": message[:160],
    }

    try:
        response = requests.post(url, json=payload, timeout=10)
        response.raise_for_status()
        result = response.json()

        if result.get("status") == "S":
            with transaction.atomic():
                # Lock the row to prevent race conditions
                wallet = SMSConfig.objects.select_for_update().get(id=sms_conf.id)
                wallet.balance -= SMS_COST_PER_MESSAGE
                wallet.save()

                SMSTransaction.objects.create(
                    school=school,
                    amount=SMS_COST_PER_MESSAGE,
                    transaction_type='REMINDER',
                    description=f"Arrears reminder sent to {member.first_name} {member.last_name}",
                    performed_by=request.user
                )

            logger.info(f"SMS sent successfully to {formatted_phone}")
            return {"status": "S", "remarks": "Reminder sent successfully."}

        else:
            return {"status": "F", "remarks": result.get('remarks', 'Unknown API error')}

    except requests.RequestException as e:
        logger.error(f"SMS API error for {formatted_phone}: {str(e)}")
        return {"status": "F", "remarks": "Failed to connect to SMS gateway."}
    except Exception as e:
        logger.error(f"Unexpected error sending SMS: {str(e)}")
        return {"status": "F", "remarks": "Internal error occurred."}


def send_bulk_arrears_reminders(request):
    """
    Send arrears reminders to all members with outstanding loans.
    """
    from .models import Loan   # Adjust import as needed

    sent = 0
    failed = 0

    loans_with_arrears = Loan.objects.filter(
        status='approved',
        balance__gt=0
    ).select_related('member')

    for loan in loans_with_arrears:
        result = send_arrears_reminder_sms(request, loan)
        if result['status'] == 'S':
            sent += 1
        else:
            failed += 1
            # Stop early if out of credits
            if "Insufficient SMS credits" in result['remarks']:
                break

    return sent, failed



# finance/models.py or utils.py
import random
import string
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from django.db import transaction
from django.utils import timezone

def generate_loan_ref(length=10):
    """Generate a unique loan reference like LN-ABC123XYZ"""
    chars = string.ascii_uppercase + string.digits
    return f"LN-{''.join(random.choices(chars, k=length))}"

def generate_schedule(loan):
    """
    Generate installments based on repayment_frequency and term_value.
    Uses flat-rate principal and interest per period.
    """
    periods = loan.term_value
    if periods <= 0:
        return

    # Compute per-period amounts
    principal_per = loan.principal_amount / Decimal(periods)
    interest_per = loan.interest_balance / Decimal(periods)

    with transaction.atomic():
        for i in range(1, periods + 1):
            # Determine due date based on frequency
            if loan.repayment_frequency == 'daily':
                due_date = loan.start_date + relativedelta(days=i)
            elif loan.repayment_frequency == 'weekly':
                due_date = loan.start_date + relativedelta(weeks=i)
            elif loan.repayment_frequency == 'monthly':
                due_date = loan.start_date + relativedelta(months=i)
            else:  # manual – no auto‑schedule
                return

            Installment.objects.create(
                loan=loan,
                principal_portion=principal_per,
                interest_portion=interest_per,
                due_date=due_date
            )

import string
import random

def generate_transaction_ref(prefix, length=8):
    """
    Generates a reference like: DEP-A1B2C3D4
    """
    chars = string.ascii_uppercase + string.digits
    code = ''.join(random.choices(chars, k=length))
    return f"{prefix}-{code}"


# reports/utils.py
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# finance/utils.py
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(columns, data, report_title="Report", company_name="Company"):
    """
    Generate an Excel workbook from report columns and data.
    Returns a BytesIO object containing the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = report_title[:31]

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1632af", end_color="1632af", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Optional title row
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    ws.cell(row=row, column=1).value = f"{company_name} - {report_title}"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    # Headers
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col.get('label', col.get('key', '')))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    row += 1

    # Data rows
    for data_row in data:
        for col_idx, col in enumerate(columns, start=1):
            key = col.get('key')
            value = data_row.get(key, '-')

            # Format based on type
            if col.get('type') == 'currency':
                try:
                    value = f"{float(value):,.0f}"
                except (ValueError, TypeError):
                    pass
            elif col.get('type') == 'date' and value:
                if hasattr(value, 'strftime'):
                    value = value.strftime('%d %b, %Y')
                else:
                    value = str(value)
            elif col.get('type') == 'status':
                value = str(value)

            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            if col.get('align') == 'right' or col.get('type') == 'currency':
                cell.alignment = number_alignment
            else:
                cell.alignment = cell_alignment
        row += 1

    # Auto-size columns
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# finance/views.py
import random
from decimal import Decimal
from datetime import datetime
from django.db.models import Sum, Q, Count, Case, When, IntegerField, Value
from django.shortcuts import render
from django.http import FileResponse
from django.contrib import messages
from django.utils import timezone
from .models import Loan, Installment, Member, SystemSetting  # adjust imports
from .utils import generate_excel_report

# -----------------------------------------------------------------
# Helper: Build the report context (shared by HTML and Excel export)
# -----------------------------------------------------------------
def get_report_context(request):
    """
    Builds the report context based on POST/GET filters.
    Returns a dict with columns, data, totals, KPIs, aging, etc.
    """
    context = {}

    # --- 1. Extract filters (works with POST and GET) ---
    date_from = request.POST.get('date_from') or request.GET.get('date_from')
    date_to = request.POST.get('date_to') or request.GET.get('date_to')
    officer_id = request.POST.get('officer') or request.GET.get('officer')
    status = request.POST.get('status') or request.GET.get('status')

    # --- 2. Base queryset (replace 'Loan' with your actual model) ---
    qs = Loan.objects.select_related('member', 'officer')

    # Apply filters
    if date_from:
        qs = qs.filter(disbursed_date__gte=date_from)
    if date_to:
        qs = qs.filter(disbursed_date__lte=date_to)
    if officer_id:
        qs = qs.filter(officer_id=officer_id)
    if status:
        qs = qs.filter(status=status)

    # --- 3. Build the data rows ---
    # We'll create a list of dicts with the fields we need.
    data = []
    for loan in qs:
        # Compute current balances (you may already have these as properties)
        principal_balance = loan.principal_balance or Decimal('0.00')
        interest_balance = loan.interest_balance or Decimal('0.00')
        total_balance = principal_balance + interest_balance

        data.append({
            'loan_reference': loan.loan_reference or f"LN-{loan.id}",
            'member_name': loan.member.get_full_name() if loan.member else 'N/A',
            'principal': loan.principal_amount,
            'interest_balance': interest_balance,
            'total_balance': total_balance,
            'status': loan.get_status_display(),
            'disbursed_date': loan.disbursed_date,
            'officer': loan.officer.get_full_name() if loan.officer else 'N/A',
            # Add more fields as needed
        })

    # --- 4. Define columns (matches the keys in data) ---
    columns = [
        {'key': 'loan_reference', 'label': 'Loan Reference'},
        {'key': 'member_name', 'label': 'Member'},
        {'key': 'principal', 'label': 'Principal', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'interest_balance', 'label': 'Interest Balance', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'total_balance', 'label': 'Total Balance', 'type': 'currency', 'align': 'right', 'total': True},
        {'key': 'status', 'label': 'Status', 'type': 'status'},
        {'key': 'disbursed_date', 'label': 'Disbursed', 'type': 'date'},
        {'key': 'officer', 'label': 'Officer'},
    ]

    # --- 5. Compute totals ---
    totals = {}
    for col in columns:
        if col.get('total'):
            # Sum the column values across data
            total_value = sum(row.get(col['key'], 0) for row in data)
            totals[col['key']] = total_value

    # --- 6. KPIs ---
    total_balance_sum = totals.get('total_balance', 0)
    total_principal_sum = totals.get('principal', 0)
    total_interest_sum = totals.get('interest_balance', 0)
    record_count = len(data)

    kpi_cards = [
        {'label': 'Total Loans', 'value': record_count, 'icon': 'bi-file-earmark-text', 'type': 'info'},
        {'label': 'Total Principal', 'value': f"UGX {total_principal_sum:,.0f}", 'icon': 'bi-cash', 'type': 'success'},
        {'label': 'Total Interest', 'value': f"UGX {total_interest_sum:,.0f}", 'icon': 'bi-percent', 'type': 'warning'},
        {'label': 'Total Balance', 'value': f"UGX {total_balance_sum:,.0f}", 'icon': 'bi-wallet2', 'type': 'danger'},
    ]

    # --- 7. Aging Summary (example: bucket by overdue days) ---
    # This is optional; you can compute from installments.
    aging_summary = []
    if data:
        # Simulate aging buckets – adapt to your actual logic
        aging_buckets = {
            '0-30 days': 0,
            '31-60 days': 0,
            '61-90 days': 0,
            '>90 days': 0,
        }
        # For demonstration, we'll just put zeros; you can compute from installments
        for key in aging_buckets:
            aging_summary.append({
                'bucket': key,
                'amount': Decimal('0.00')  # replace with actual calculations
            })

    # --- 8. Summary totals (extra stats) ---
    summary_totals = {
        'total_records': record_count,
        'total_amount': total_balance_sum,
        'total_paid': Decimal('0.00'),   # you can compute from repayments
        'outstanding': total_balance_sum,
        'recovery_rate': 0,
        'par_30': 0,
    }

    # --- 9. Officer list for the filter dropdown ---
    from django.contrib.auth import get_user_model
    User = get_user_model()
    officer_list = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    # --- 10. Company info (adjust as needed) ---
    # If you have a Company model, fetch it; else create a dummy.
    company = {
        'name': 'Your Company Ltd',
        'logo': None,
        'phone': '+256 700 000 000',
        'email': 'info@company.com',
        'website': 'www.company.com',
    }

    # --- 11. Final context ---
    context.update({
        'columns': columns,
        'data': data,
        'totals': totals,
        'has_data': bool(data),
        'kpi_cards': kpi_cards,
        'summary_totals': summary_totals,
        'aging_summary': aging_summary,
        'report_title': 'Loan Portfolio Report',
        'company': company,
        'date_from': date_from,
        'date_to': date_to,
        'selected_officer': officer_id,
        'selected_status': status,
        'officer_list': officer_list,
        'officer_name': dict(officer_list.values_list('id', 'username')).get(int(officer_id) if officer_id else None),
        'generated_date': timezone.now().strftime('%d %b %Y %H:%M'),
        'generated_by': request.user.get_full_name() if request.user.is_authenticated else 'System',
    })
    return context


# -----------------------------------------------------------------
# Main Report View – handles both HTML and Excel export
# -----------------------------------------------------------------
def report_view(request):
    # Determine if we are exporting Excel
    if request.method in ('POST', 'GET'):
        # Build the context using filters
        context = get_report_context(request)

        # Check for Excel export flag (hidden input)
        if request.POST.get('export_excel') == '1' or request.GET.get('export_excel') == '1':
            # Generate Excel file
            excel_file = generate_excel_report(
                columns=context['columns'],
                data=context['data'],
                report_title=context['report_title'],
                company_name=context['company']['name']
            )
            filename = f"{context['report_title'].replace(' ', '_')}_{context['generated_date'].replace(' ', '_').replace(':', '')}.xlsx"
            response = FileResponse(
                excel_file,
                as_attachment=True,
                filename=filename,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            return response

        # Else render HTML
        return render(request, 'reports/base_report.html', context)

    # Initial GET with no filters – show empty state
    # We still need officer_list and company info
    from django.contrib.auth import get_user_model
    User = get_user_model()
    officer_list = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    company = {
        'name': 'Your Company Ltd',
        'logo': None,
        'phone': '+256 700 000 000',
        'email': 'info@company.com',
        'website': 'www.company.com',
    }
    context = {
        'has_data': False,
        'columns': [],
        'data': [],
        'report_title': 'Loan Portfolio Report',
        'company': company,
        'officer_list': officer_list,
        'generated_date': timezone.now().strftime('%d %b %Y %H:%M'),
        'generated_by': request.user.get_full_name() if request.user.is_authenticated else 'System',
    }
    return render(request, 'reports/base_report.html', context)




# finance/utils.py
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(columns, data, report_title="Report", company_name="Company", totals=None):
    """
    Generate an Excel workbook from report columns and data.
    Optionally add a totals row.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = report_title[:31]

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1632af", end_color="1632af", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Optional title row
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    ws.cell(row=row, column=1).value = f"{company_name} - {report_title}"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 1

    # Headers
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col.get('label', col.get('key', '')))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    row += 1

    # Data rows
    for data_row in data:
        for col_idx, col in enumerate(columns, start=1):
            key = col.get('key')
            value = data_row.get(key, '-')

            # Format based on type
            if col.get('type') == 'currency':
                try:
                    value = f"{float(value):,.0f}"
                except (ValueError, TypeError):
                    pass
            elif col.get('type') == 'date' and value:
                if hasattr(value, 'strftime'):
                    value = value.strftime('%d %b, %Y')
                else:
                    value = str(value)
            elif col.get('type') == 'status':
                value = str(value)

            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            if col.get('align') == 'right' or col.get('type') == 'currency':
                cell.alignment = number_alignment
            else:
                cell.alignment = cell_alignment
        row += 1

    # -------------------------
    # TOTALS ROW (if totals provided)
    # -------------------------
    if totals:
        # Write "TOTALS" in the first column
        ws.cell(row=row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        for col_idx, col in enumerate(columns, start=1):
            if col.get('total') and col.get('key') in totals:
                total_val = totals[col['key']]
                # Format as currency
                if col.get('type') == 'currency':
                    try:
                        total_val = f"{float(total_val):,.0f}"
                    except (ValueError, TypeError):
                        pass
                else:
                    total_val = str(total_val)

                cell = ws.cell(row=row, column=col_idx, value=total_val)
                cell.font = Font(bold=True)
                cell.border = border
                if col.get('align') == 'right' or col.get('type') == 'currency':
                    cell.alignment = number_alignment
                else:
                    cell.alignment = cell_alignment

        row += 1

    # Auto-size columns (simple)
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output